from django.shortcuts import render, HttpResponse, redirect
from .forms import nuevaReservaFoms, editReservaFoms, asignaMesaForm, formBuscarFechaHistori, formIncidencia, formBuscarIncidencia, formBuscarPLaza, formIncidenciaLog, formBuscarIncidenciaLog
from django.contrib import messages
from .models import nuevaReserva, incidencia,  mozosPlaza, plazaAlmuerzo, anfitriona, plazaCena, plazaAlmuerzoMan, plazaCenaMan, incidenciaLog, controlmensaje
from django.db import models
from datetime import datetime, timedelta, date
from django.contrib.auth.decorators import login_required
from django.db.models import Sum
import openpyxl
from django.template.defaultfilters import floatformat
import requests
import random
import string
import re
import time
import pytz
import matplotlib.pyplot as ptl
from io import BytesIO
import base64
from django.db.models.functions import Extract, TruncDay
from django.db.models import Case, When, Value, IntegerField, CharField
from django.db.models import Count
# Create your views here.

@login_required
def creaNuevaReserva(request):
    fecha_actual = datetime.now().date()
    form = nuevaReservaFoms()
    context = {'form_reserva': form}

    if request.method == 'POST':
        form = nuevaReservaFoms(request.POST)
        if form.is_valid():
            a = form.save(commit=False)
            a.estado_id = 1
            cliente = a.nombre
            fecha_reserva = a.fechaReserva
            hora_reserva = a.hora
            personas = a.cantidadPersonas
            
            if nuevaReserva.objects.filter(nombre = cliente, fechaReserva = fecha_reserva, hora = hora_reserva, cantidadPersonas = personas).exists():
                  return redirect('delDia')
            else:
                a.save()
            
            ## Preparar envío por whatsapp al cliente
            id_reserva = a.id
            cliente = a.nombre
            fecha_reserva = a.fechaReserva.strftime('%d/%m/%Y')
            hora_reserva = a.hora.strftime('%I:%M %p')
            personas = a.cantidadPersonas
            telwhat = a.telefono

            # hora_cena = datetime.strptime(hora_reserva, '%I:%M %p').time()
            # hora_limite = time(17, 30)

            if not telwhat is None:
            ## Quitar las estupideces que agrega EVA CRISSEL al telefono, porque es floja, no le gusta escribir bien
            ## y se gana el sueldo facilmente. 
                telwhat = re.sub(r'[^\d]+', '', telwhat)
                telwhat = "51" + telwhat if not telwhat.startswith("51") else telwhat 
                
                # if hora_cena <= hora_limite:
                mensaje = f"""Estimado/a: *{cliente}*. 
Su reserva ha sido confirmada.\n 
*Fecha de la reservación: {fecha_reserva}* 
*Hora de la reservación: {hora_reserva}*
*Cantidad de personas: {personas}*\n 
Esperamos brindarle una experiencia gastronómica memorable en nuestro establecimiento.
¡Estamos ansiosos por darle la bienvenida!
Puede consultar nuestra carta en https://www.elcharrua.com/carta \n
Muchas gracias por elegirnos.
Te esperamos en *El Charrúa*"""
#                 else:
#                     mensaje = f"""Estimado/a: *{cliente}*. 
# Su reserva ha sido confirmada.\n 
# *Fecha de la reservación: {fecha_reserva}* 
# *Hora de la reservación: {hora_reserva}*
# *Cantidad de personas: {personas}*\n 
# Esperamos brindarle una experiencia gastonómica memorable en nuestro establecimiento.
# ¡Estamos ansiosos por darle la bienvenida!
# Puede consultar nuestra carta en https://www.elcharrua.com/carta \n
# Muchas gracias por elegirnos.
# Te esperamos en *El Charrúa*\n
# *Hoy 31/12/2023, se cobrará el derecho de corcho para todos los licores.*\n
# *Dicho monto va desde los S/50.*\n
# *Así mismo, les deseamos un prospero año nuevo.*"""
                

                letras = string.ascii_lowercase
                uid_custom = ''.join(random.choice(letras) for i in range(6))  

                url = 'https://www.waboxapp.com/api/send/chat'
                data = {
                    'token':'8f9a42d9ebc4392cca61ffd6fa13d3a6644336f382f95',
                    'uid': '51994043376',
                    'to': telwhat,
                    'custom_uid':uid_custom,
                    'text': mensaje
                }

                # comprobar si existe el id de reserva en la tabla de control de mensaje, si esta y enviado es 1 pasar por alto
                if not controlmensaje.objects.filter(id_reservacion = id_reserva, enviado = 1).exists():
                      
                      response = requests.post(url, data = data)
                      #print("what enviado")
                      guarda_control = controlmensaje(id_reservacion = id_reserva, enviado = 1)
                      guarda_control.save()
                
                #print(response.content)

            cuenta_listadia = nuevaReserva.objects.filter(estado_id=1, fecha_creacion__date=fecha_actual).count()
            #print(cuenta_listadia)
            if cuenta_listadia == 1:
                envioRecordatorio()
                       
            return redirect('/reservasDelDia/')
        else:

            return render(request, 'nuevaReserva.html', {'form_reserva': form})

    return render(request, 'nuevaReserva.html', context)

def envioRecordatorio():
    fecha_actual = datetime.now().date()
    listado_envio = nuevaReserva.objects.filter(fechaReserva=fecha_actual, estado_id=1)
    for telefono in listado_envio:
      telwhat = telefono.telefono
      if not telwhat is None:
            ## Quitar de las estupideces que agrega EVA CRISSEL al telefono, porque es floja, no le gusta escribir bien
            ## y se gana el sueldo facilmente. 
                telwhat = re.sub(r'[^\d]+', '', telwhat)
                telwhat = "51" + telwhat if not telwhat.startswith("51") else telwhat
                cliente = telefono.nombre 
                hora_reserva = telefono.hora.strftime('%I:%M %p')
                personas = telefono.cantidadPersonas
                # hora_cena = datetime.strptime(hora_reserva, '%I:%M %p').time()
                # hora_limite = time(17, 30)


                # if hora_cena <= hora_limite:
                mensaje = f"""Estimado/a: *{cliente}*. 
Le recordamos que tiene una reserva para el día de hoy en El Charrúa.\n 
*Hora de la reservación: {hora_reserva}*
*Cantidad de personas: {personas}*\n 
Esperamos brindarle una experiencia gastronómica memorable en nuestro establecimiento.
¡Estamos ansiosos por darle la bienvenida!
Puede consultar nuestra carta en https://www.elcharrua.com/carta \n
Muchas gracias por elegirnos.
Te esperamos en *El Charrúa*"""
#                 else:
#                     mensaje = f"""Estimado/a: *{cliente}*. 
# Le recordamos que tiene una reserva para el día de hoy en El Charrúa.\n 
# *Hora de la reservación: {hora_reserva}*
# *Cantidad de personas: {personas}*\n 
# Esperamos brindarle una experiencia gastronómica memorable en nuestro establecimiento.
# ¡Estamos ansiosos por darle la bienvenida!
# Puede consultar nuestra carta en https://www.elcharrua.com/carta \n
# Muchas gracias por elegirnos.
# Te esperamos en *El Charrúa*\n
# *Hoy 31/12/2023, se cobrará el derecho de corcho para todos los licores.*\n
# *Dicho monto va desde los S/50.*\n
# *Así mismo, les deseamos un prospero año nuevo.*"""
                      
                #mensaje = 'Su reserva ha sido confirmada, muchas gracias por elegirnos.Te esperamos en El Charrúa. '

                letras = string.ascii_lowercase
                uid_custom = ''.join(random.choice(letras) for i in range(6))  

                url = 'https://www.waboxapp.com/api/send/chat'
                data = {
                    'token':'8f9a42d9ebc4392cca61ffd6fa13d3a6644336f382f95',
                    'uid': '51994043376',
                    'to': telwhat,
                    'custom_uid':uid_custom,
                    'text': mensaje
                }
            
                response = requests.post(url, data = data)
                time.sleep(3)
                
@login_required
def listadoEnEspera(request):

    fecha_actual = datetime.now().date()

    cuenta_atendido = nuevaReserva.objects.filter(estado_id = 2, fechaReserva = fecha_actual).count()
    cuenta_anulado = nuevaReserva.objects.filter(estado_id = 3, fechaReserva = fecha_actual).count()
    cuenta_noshow = nuevaReserva.objects.filter(estado_id = 4, fechaReserva = fecha_actual).count()

    
    fecha_actual = datetime.now().date()

    enEspera = nuevaReserva.objects.filter(estado_id = 1).order_by('fechaReserva')
    cuenta_enEspera = nuevaReserva.objects.filter(estado_id = 1).count()
    return render(request, 'listadoEnEspera.html', {"listaEspera": enEspera, "totalEspera":cuenta_enEspera})
@login_required
def listadoDelDia(request):
    
    fecha_actual = datetime.now().date()

    cuenta_atendido = nuevaReserva.objects.filter(
        estado_id=2, fechaReserva=fecha_actual).count()
    cuenta_anulado = nuevaReserva.objects.filter(
        estado_id=3, fechaReserva=fecha_actual).count()
    cuenta_noshow = nuevaReserva.objects.filter(
        estado_id=4, fechaReserva=fecha_actual).count()
    cuenta_pendiente = nuevaReserva.objects.filter(
        estado_id=1, fechaReserva=fecha_actual).count()

    deldia = nuevaReserva.objects.filter(
        fechaReserva=fecha_actual).order_by('hora')
    
    cuenta_deldia = nuevaReserva.objects.filter(
         fechaReserva=fecha_actual).exclude(estado_id__in=[3, 4]).count()
    
    totalClientesAten = nuevaReserva.objects.filter(
        fechaReserva=fecha_actual).exclude(estado_id__in=[3, 4]).aggregate(Sum('cantidadPersonas'))
    
    totalpersonas = (totalClientesAten['cantidadPersonas__sum'])

    if request.method == "POST":
        nombre_cliente = request.POST.get('nombre')
        if len(nombre_cliente) == 0:
            return render(request, 'reservasDelDia.html', {"listaEspera": deldia, "totalDia": cuenta_deldia, 'fechaHoy': fecha_actual, 'totalAtendido': cuenta_atendido,
            'totalAnulado': cuenta_anulado, 'totalNoshow': cuenta_noshow, "totalPersonas": totalpersonas, 'totalPendiente':cuenta_pendiente})
    
        else:
            busqueda_activa = 1
            coindicenias = nuevaReserva.objects.filter(nombre__icontains = nombre_cliente, fechaReserva=fecha_actual).order_by('hora')

            return render(request, 'reservasDelDia.html', {"listaEspera": coindicenias, "totalDia": cuenta_deldia, 'fechaHoy': fecha_actual, 'totalAtendido': cuenta_atendido,
            'totalAnulado': cuenta_anulado, 'totalNoshow': cuenta_noshow, "totalPersonas": totalpersonas, 'totalPendiente':cuenta_pendiente, 'busqueda_activa':busqueda_activa})
    else:
        return render(request, 'reservasDelDia.html', {"listaEspera": deldia, "totalDia": cuenta_deldia, 'fechaHoy': fecha_actual, 'totalAtendido': cuenta_atendido,
            'totalAnulado': cuenta_anulado, 'totalNoshow': cuenta_noshow, "totalPersonas": totalpersonas, 'totalPendiente':cuenta_pendiente})  
  
@login_required
def listadoEnProceso(request):
    en_proceso = nuevaReserva.objects.filter(estado_id = 2)
    cuenta_en_proceso = nuevaReserva.objects.filter(estado_id = 2).count()
    return render(request, 'listadoEnProceso.html', {"listaEnProceso": en_proceso, "totalProceso":cuenta_en_proceso})
@login_required
def listadoCompletado(request):

    fecha_actual = datetime.now().date()
    formBuscar = formBuscarFechaHistori()
    completado = nuevaReserva.objects.filter(fechaReserva = fecha_actual).order_by('-fechaReserva', 'hora')
    cuenta_completado = nuevaReserva.objects.filter(estado_id=2).count()
    cuenta_completado_dia = nuevaReserva.objects.filter(estado_id=2, fechaReserva=fecha_actual).count()
    totalClientesAten = nuevaReserva.objects.filter(fechaReserva=fecha_actual).aggregate(Sum('cantidadPersonas'))
    totalpersonas = (totalClientesAten['cantidadPersonas__sum'])

    if request.method == "POST":
          nombre_cliente = request.POST.get('nombre')
          if len(nombre_cliente) == 0:
                return render(request, 'listadoHistorico.html', {"listaCompletado": completado, "totalCompletado": cuenta_completado, "formBusca": formBuscar, "totalCompletadodia":cuenta_completado_dia, 
                                                     "totalClientes":totalpersonas})
          else:
                
            busqueda_activa = 1
            coindicenias = nuevaReserva.objects.filter(nombre__icontains = nombre_cliente).order_by('-fechaReserva','hora')
            return render(request, 'listadoHistorico.html', {"listaCompletado": coindicenias, "totalCompletado": cuenta_completado, "formBusca": formBuscar, "totalCompletadodia":cuenta_completado_dia, 
                                                     "totalClientes":totalpersonas, 'busqueda_activa':busqueda_activa})
          

    else:
        return render(request, 'listadoHistorico.html', {"listaCompletado": completado, "totalCompletado": cuenta_completado, "formBusca": formBuscar, "totalCompletadodia":cuenta_completado_dia, 
                                                     "totalClientes":totalpersonas})

@login_required
def editarReserva(request, id):

    editCurso = nuevaReserva.objects.get(id = id)
    form = editReservaFoms(instance=editCurso)

    # if editCurso.estado_id != 1:
    
    #      return HttpResponse('Esta reserva ya ha sido procesada. No se puede modificar')

    if request.method == 'POST':

        form = editReservaFoms(request.POST,instance=editCurso)
        if form.is_valid:
            fecha_reserva = editCurso.fechaReserva.strftime('%d/%m/%Y')
            fecha_nueva = request.POST.get('fechaReserva')
            telwhat = request.POST.get('telefono')
            hora_reserva = request.POST.get('hora')
            personas = request.POST.get('cantidadPersonas')
            cliente = request.POST.get('nombre')

            if fecha_reserva != fecha_nueva:
                  if not telwhat is None:
                        ## Quitar de las estupideces que agrega EVA CRISSEL al telefono, porque es floja, no le gusta escribir bien
                        ## y se gana el sueldo facilmente. 
                    telwhat = re.sub(r'[^\d]+', '', telwhat)
                    telwhat = "51" + telwhat if not telwhat.startswith("51") else telwhat
                    mensaje = f"""Estimado/a: *{cliente}*. 
Esperamos que esté teniendo un excelente día. Nos complace informarle que su solicitud de cambiar la
fecha de su reserva en nuestro restaurante ha sido atendida. 
Su nueva fecha de reserva es el día *{fecha_nueva}* a las *{hora_reserva}*. \n
Agradecemos su solicitud y esperamos que esta nueva fecha sea aún más conveniente para usted y su grupo.
Agradecemos su preferencia y le aseguramos que estamos trabajando arduamente para garantizar una 
experiencia gastronómica memorable en su próxima visita. Si tiene alguna pregunta o inquietud, no
dude en comunicarse con nosotros al *994 043 376*. \n
Gracias por elegir *EL CHARRÚA*. \n
Saludos cordiales"""

                    letras = string.ascii_lowercase
                    uid_custom = ''.join(random.choice(letras) for i in range(6))  

                    url = 'https://www.waboxapp.com/api/send/chat'
                    data = {
                        'token':'8f9a42d9ebc4392cca61ffd6fa13d3a6644336f382f95',
                        'uid': '51994043376',
                        'to': telwhat,
                        'custom_uid':uid_custom,
                        'text': mensaje
                    }
                        
                    response = requests.post(url, data = data) 
            form.save()
            return redirect('/reservasDelDia/') 

    editCurso = nuevaReserva.objects.get(id = id)
    form = editReservaFoms(instance=editCurso)

    return render(request,'editarReserva.html',{"form":form})
@login_required
def cambioEstadoReserva(request, id):
    
    cambiaEstado = nuevaReserva.objects.get(id = id)
    cambiaEstado.estado_id = 2
    cambiaEstado.save()    
    
    return redirect('/reservasDelDia/')
@login_required
def cambiaEstadoAnulado(request, id):
    cambiaEstadoAnulado = nuevaReserva.objects.get(id = id)
    cambiaEstadoAnulado.estado_id = 3
    cambiaEstadoAnulado.save()    
    
    return redirect('/reservasDelDia/')
@login_required
def cambiaEstadoNoShow(request, id):

    cambiaEstadoShow = nuevaReserva.objects.get(id = id)
    cambiaEstadoShow.estado_id = 4
    cambiaEstadoShow.save()    
    
    return redirect('/reservasDelDia/')
@login_required
def cargaAsignarMesa(request, id):
    form = asignaMesaForm()
    reserva = nuevaReserva.objects.filter(id = id)
    return render(request,'asignaMesa.html',{"listaReserva":reserva, 'form':form})
@login_required
def estadoActual(request):

    fecha_actual = datetime.now().date()

    cuenta_enEspera = nuevaReserva.objects.filter(fechaReserva = fecha_actual).count()
    cuenta_en_proceso = nuevaReserva.objects.filter(estado_id = 2, fechaReserva = fecha_actual).count()
    cuenta_completado = nuevaReserva.objects.filter(estado_id = 3, fechaReserva = fecha_actual).count()
    cuenta_cancelada = nuevaReserva.objects.filter(estado_id = 4, fechaReserva = fecha_actual).count()

    ratio_progreso = round((cuenta_en_proceso/cuenta_enEspera)*100,2)
    ratio_completadas = round((cuenta_completado/cuenta_enEspera)*100,2)
    ratio_faltante = round(((cuenta_en_proceso + cuenta_completado)/cuenta_enEspera)*100,2)

    data = {'porcentajles': [ratio_progreso,ratio_completadas, ratio_faltante],'etiqueta':['progreso','completo','faltante']}


    return render(request, 'estadoActual.html', {'cantidadEspera':cuenta_enEspera, 'cantidadProgreso':cuenta_en_proceso, 'cantidadCompletada':cuenta_completado,
    'cantidadCancelada':cuenta_cancelada, 'ratioProgreso':ratio_progreso, 'ratioCompletado':ratio_completadas, 'ratioFaltante':ratio_faltante, 'data':data})
@login_required
def guardarMesa(request, id):
    
    if request.method == 'POST':
        mesa = request.POST.get('mesa')
        cambiaMesa = nuevaReserva.objects.get(id = id)
        cambiaMesa.mesa_asignadaa = mesa
        cambiaMesa.save()

        return redirect('delDia')
@login_required
def muentraCena(request):

    fecha_actual = datetime.now().date()

    cuenta_atendido = nuevaReserva.objects.filter(estado_id = 2, fechaReserva = fecha_actual).count()
    cuenta_anulado = nuevaReserva.objects.filter(estado_id = 3, fechaReserva = fecha_actual).count()
    cuenta_noshow = nuevaReserva.objects.filter(estado_id = 4, fechaReserva = fecha_actual).count()
    cuenta_pendiente = nuevaReserva.objects.filter(estado_id=1, fechaReserva=fecha_actual, hora__gt = '16:01').count()

    deldia = nuevaReserva.objects.filter(fechaReserva = fecha_actual, hora__gt = '16:01' ).order_by('hora')

    totalClientesAten = nuevaReserva.objects.filter(fechaReserva = fecha_actual, hora__gt = '16:01').exclude(estado_id__in=[3, 4]).aggregate(Sum('cantidadPersonas'))
    totalpersonas = (totalClientesAten['cantidadPersonas__sum'])

    cuenta_deldia = nuevaReserva.objects.filter(fechaReserva = fecha_actual, hora__gt = '16:01').exclude(estado_id__in=[3, 4]).count()

    turno = "CENA"

    return render(request, 'reservasDelDia.html', {"listaEspera": deldia, "totalDia":cuenta_deldia, 'fechaHoy':fecha_actual, 'totalAtendido':cuenta_atendido,
    'totalAnulado':cuenta_anulado, 'totalNoshow':cuenta_noshow, "totalPersonas":totalpersonas, "turno":turno, 'totalPendiente':cuenta_pendiente })

@login_required
def muentraTarde(request):

    fecha_actual = datetime.now().date()

    cuenta_atendido = nuevaReserva.objects.filter(estado_id = 2, fechaReserva = fecha_actual).count()
    cuenta_anulado = nuevaReserva.objects.filter(estado_id = 3, fechaReserva = fecha_actual).count()
    cuenta_noshow = nuevaReserva.objects.filter(estado_id = 4, fechaReserva = fecha_actual).count()
    cuenta_pendiente = nuevaReserva.objects.filter(estado_id=1, fechaReserva=fecha_actual, hora__lt = '16:01').count()

    deldia = nuevaReserva.objects.filter(fechaReserva = fecha_actual, hora__lt = '16:01' ).order_by('hora')

    totalClientesAten = nuevaReserva.objects.filter(fechaReserva = fecha_actual, hora__lt = '16:01').exclude(estado_id__in=[3, 4]).aggregate(Sum('cantidadPersonas'))
    totalpersonas = (totalClientesAten['cantidadPersonas__sum'])

    cuenta_deldia = nuevaReserva.objects.filter(fechaReserva = fecha_actual, hora__lt = '16:01' ).exclude(estado_id__in=[3, 4]).count()

    turno = "ALMUERZO"

    return render(request, 'reservasDelDia.html', {"listaEspera": deldia, "totalDia":cuenta_deldia, 'fechaHoy':fecha_actual, 'totalAtendido':cuenta_atendido,
    'totalAnulado':cuenta_anulado, 'totalNoshow':cuenta_noshow, "totalPersonas":totalpersonas, "turno":turno, 'totalPendiente':cuenta_pendiente })

@login_required
def exportaExcel(request):

    fecha_actual = datetime.now().date()
    deldia = nuevaReserva.objects.filter(fechaReserva = fecha_actual ).order_by('hora')
    
    #crea nuevo libro
    wb = openpyxl.Workbook()

    hoja = wb.active

    hoja['A1'] = 'Hora'
    hoja['B1'] = 'Cliente'
    hoja['C1'] = 'Cantidad Personas'
    hoja['D1'] = 'Observaciones'
    hoja['E1'] = 'Fecha'

    row = 2

    for i in deldia:
        hoja.cell(row, 1, i.hora)
        hoja.cell(row, 2, i.nombre)
        hoja.cell(row, 3, i.cantidadPersonas)
        hoja.cell(row, 4, i.observaciones)
        hoja.cell(row, 5, fecha_actual)
        row += 1
    
    response = HttpResponse(content_type = 'application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename="reservasdia.xlsx"'

    wb.save(response)

    return response
@login_required
def buscaHistoricoFecha(request):

    formBuscar = formBuscarFechaHistori(request.POST)

    if request.method == 'POST':
        if formBuscar.is_valid():
            decision = request.POST.get("accion")
            if decision == "Cena":
                b = formBuscar.cleaned_data['fechaBuscar']
                turno = "CENA"
                completado = nuevaReserva.objects.filter(fechaReserva=b, hora__gt='16:01').order_by('-fechaReserva', 'hora')
                cuenta_completado = nuevaReserva.objects.filter(estado_id=2, hora__gt='16:01').count()
                cuenta_completado_dia = nuevaReserva.objects.filter(estado_id=2, fechaReserva=b, hora__gt='16:01').count()
                totalClientesAten = nuevaReserva.objects.filter(fechaReserva=b, estado_id=2, hora__gt='16:01').aggregate(Sum('cantidadPersonas'))
                totalpersonas = (totalClientesAten['cantidadPersonas__sum'])
                return render(request, 'listadoHistorico.html', {"listaCompletado": completado, "totalCompletado": cuenta_completado, "formBusca": formBuscar, "totalCompletadodia":cuenta_completado_dia,
                                                                    "totalClientes":totalpersonas, "turno":turno })
            
            if decision == "Almuerzo":

                b = formBuscar.cleaned_data['fechaBuscar']
                turno = "ALMUERZO"
                completado = nuevaReserva.objects.filter(fechaReserva=b, hora__lt='16:01').order_by('-fechaReserva', 'hora')
                cuenta_completado = nuevaReserva.objects.filter(estado_id=2, hora__lt='16:01').count()
                cuenta_completado_dia = nuevaReserva.objects.filter(estado_id=2, fechaReserva=b, hora__lt='16:01').count()
                totalClientesAten = nuevaReserva.objects.filter(fechaReserva=b, estado_id=2, hora__lt='16:01').aggregate(Sum('cantidadPersonas'))
                totalpersonas = (totalClientesAten['cantidadPersonas__sum'])
                return render(request, 'listadoHistorico.html', {"listaCompletado": completado, "totalCompletado": cuenta_completado, "formBusca": formBuscar, "totalCompletadodia":cuenta_completado_dia,
                                                                    "totalClientes":totalpersonas, "turno":turno})

            else:
                b = formBuscar.cleaned_data['fechaBuscar']
                if 'menos' in request.POST:
                    c = b-timedelta(days=1)
                    completado = nuevaReserva.objects.filter(fechaReserva=(b-timedelta(days=1))).order_by('-fechaReserva', 'hora')
                    cuenta_completado_dia = nuevaReserva.objects.filter(estado_id=2, fechaReserva=c).count()
                    cuenta_completado = nuevaReserva.objects.filter(estado_id=2).count()
                    fecha_menos = b-timedelta(days=1)
                    formBuscar = formBuscarFechaHistori(initial= {'fechaBuscar': fecha_menos.strftime('%Y-%m-%d')})
                    totalClientesAten = nuevaReserva.objects.filter(fechaReserva=c, estado_id=2).aggregate(Sum('cantidadPersonas'))
                    totalpersonas = (totalClientesAten['cantidadPersonas__sum'])
                    return render(request, 'listadoHistorico.html', {"listaCompletado": completado, "totalCompletado": cuenta_completado, "formBusca": formBuscar, "totalCompletadodia":cuenta_completado_dia,
                                                                    "totalClientes":totalpersonas})
                if 'mas' in request.POST:
                    c = b+timedelta(days=1)
                    completado = nuevaReserva.objects.filter(fechaReserva=(b+timedelta(days=1))).order_by('-fechaReserva', 'hora')
                    cuenta_completado_dia = nuevaReserva.objects.filter(estado_id=2, fechaReserva=c).count()
                    cuenta_completado = nuevaReserva.objects.filter(estado_id=2).count()
                    fecha_menos = b+timedelta(days=1)
                    formBuscar = formBuscarFechaHistori(initial= {'fechaBuscar': fecha_menos.strftime('%Y-%m-%d')})
                    totalClientesAten = nuevaReserva.objects.filter(fechaReserva=c, estado_id=2).aggregate(Sum('cantidadPersonas'))
                    totalpersonas = (totalClientesAten['cantidadPersonas__sum'])
                    return render(request, 'listadoHistorico.html', {"listaCompletado": completado, "totalCompletado": cuenta_completado, "formBusca": formBuscar, "totalCompletadodia":cuenta_completado_dia,
                                                                    "totalClientes":totalpersonas})
                else:
                    
                    completado = nuevaReserva.objects.filter(fechaReserva=b).order_by('-fechaReserva', 'hora')
                    cuenta_completado = nuevaReserva.objects.filter(estado_id=2).count()
                    cuenta_completado_dia = nuevaReserva.objects.filter(estado_id=2, fechaReserva=b).count()
                    totalClientesAten = nuevaReserva.objects.filter(fechaReserva=b, estado_id=2).aggregate(Sum('cantidadPersonas'))
                    totalpersonas = (totalClientesAten['cantidadPersonas__sum'])
                    return render(request, 'listadoHistorico.html', {"listaCompletado": completado, "totalCompletado": cuenta_completado, "formBusca": formBuscar, "totalCompletadodia":cuenta_completado_dia,
                                                                    "totalClientes":totalpersonas })

@login_required
def exportaExcelHistorico(request):

    completado = nuevaReserva.objects.all().order_by('-fechaReserva')[:5000]
    
    #crea nuevo libro
    wb1 = openpyxl.Workbook()

    hoja = wb1.active

    hoja['A1'] = 'Cliente'
    hoja['B1'] = 'Teléfono'
    hoja['C1'] = 'Correo'
    hoja['D1'] = 'Fecha de Reserva'
    hoja['E1'] = 'Hora'
    hoja['F1'] = 'Cantidad de Personas'
    hoja['G1'] = 'Observaciones'
    hoja['H1'] = 'Estado'
    hoja['I1'] = 'Origen Reserva'

    row = 2

    for i in completado:
        hoja.cell(row, 1, i.nombre)
        hoja.cell(row, 2, i.telefono)
        hoja.cell(row, 3, i.emailr)
        hoja.cell(row, 4, i.fechaReserva)
        hoja.cell(row, 5, i.hora)
        hoja.cell(row, 6, i.cantidadPersonas)
        hoja.cell(row, 7, i.observaciones)
        hoja.cell(row, 8, str(i.estado))
        hoja.cell(row,9, str(i.origen_reserva))
        row += 1
    
    response = HttpResponse(content_type = 'application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename="HistoricoReservas.xlsx"'

    wb1.save(response)

    return response
@login_required
def agregarIncidencia(request):

    bi = incidencia.objects.all().order_by('-fecha_incidencia')
    form = formIncidencia()
    form_buscar = formBuscarIncidencia()

    if request.method == 'POST':
        form = formIncidencia(request.POST)
        if form.is_valid():
            form.save() 
            return redirect('/libroIncidencias/')
        else:
            
            return render(request, 'incidencia.html', {'form_inciden':form, 'formBuscarInciden':form_buscar, 'listadoInciden':bi})
    
    return render(request, 'incidencia.html', {'form_inciden':form, 'formBuscarInciden':form_buscar, 'listadoInciden':bi})
@login_required
def buscarIncidencias(request):

    form = formIncidencia()
    form_buscar = formBuscarIncidencia()

    form_buscar = formBuscarIncidencia(request.POST)
    if request.method == 'POST':
        if form_buscar.is_valid():
            b = form_buscar.cleaned_data['fechaBIncidencia']
            bi = incidencia.objects.filter(fecha_incidencia = b).order_by('-fecha_incidencia')
            return render(request, 'incidencia.html', {'form_inciden':form, 'formBuscarInciden':form_buscar, 'listadoInciden':bi})
    else:
        return redirect('/libroIncidencias/')
@login_required
def verPlaza(request):
    fecha_actual = datetime.now().date()

    hora_actual = datetime.now().strftime('%H:%M')

    ########Almuerzo##############

    if plazaAlmuerzo.objects.filter(fecha_dia=fecha_actual).exists():
        estado = 1
        p1 = plazaAlmuerzo.objects.get(fecha_dia=fecha_actual, plaza='Plaza1')
        p2 = plazaAlmuerzo.objects.get(fecha_dia=fecha_actual, plaza='Plaza2')
        p3 = plazaAlmuerzo.objects.get(fecha_dia=fecha_actual, plaza='Plaza3')
        p4 = plazaAlmuerzo.objects.get(fecha_dia=fecha_actual, plaza='Plaza4')
        p5 = plazaAlmuerzo.objects.get(fecha_dia=fecha_actual, plaza='Plaza5')
        p6 = plazaAlmuerzo.objects.get(fecha_dia=fecha_actual, plaza='Plaza6')
        p7 = plazaAlmuerzo.objects.get(fecha_dia=fecha_actual, plaza='Plaza7')
        p8 = plazaAlmuerzo.objects.get(fecha_dia=fecha_actual, plaza='Plaza8')
        p9 = plazaAlmuerzo.objects.get(fecha_dia=fecha_actual, plaza='Plaza9')
        p10 = plazaAlmuerzo.objects.get(
            fecha_dia=fecha_actual, plaza='Plaza10')
        anfi_nombre = plazaAlmuerzo.objects.get(
            fecha_dia=fecha_actual, plaza='Plaza1')

        plaza1 = p1
        plaza2 = p2
        plaza3 = p3
        plaza4 = p4
        plaza5 = p5
        plaza6 = p6
        plaza7 = p7
        plaza8 = p8
        plaza9 = p9
        plaza10 = p10
        nombre_anfi = anfi_nombre.nombre_anfitriona
        mozoplaza1 = plaza1.mozo_nombre
        mozoplaza2 = plaza2.mozo_nombre
        mozoplaza3 = plaza3.mozo_nombre
        mozoplaza4 = plaza4.mozo_nombre
        mozoplaza5 = plaza5.mozo_nombre
        mozoplaza6 = plaza6.mozo_nombre
        mozoplaza7 = plaza7.mozo_nombre
        mozoplaza8 = plaza8.mozo_nombre
        mozoplaza9 = plaza9.mozo_nombre
        mozoplaza10 = plaza10.mozo_nombre
    else:
        estado = 2
        mozoplaza1 = ""
        mozoplaza2 = ""
        mozoplaza3 = ""
        mozoplaza4 = ""
        mozoplaza5 = ""
        mozoplaza6 = ""
        mozoplaza7 = ""
        mozoplaza8 = ""
        mozoplaza9 = ""
        mozoplaza10 = ""
        nombre_anfi = ""

    mozoz_p = mozosPlaza.objects.all()
    anfitrionas = anfitriona.objects.all()

    ########CENA##########

    if plazaCena.objects.filter(fecha_dia=fecha_actual).exists():
        estadoc = 1
        p1c = plazaCena.objects.get(fecha_dia=fecha_actual, plaza='Plaza1')
        p2c = plazaCena.objects.get(fecha_dia=fecha_actual, plaza='Plaza2')
        p3c = plazaCena.objects.get(fecha_dia=fecha_actual, plaza='Plaza3')
        p4c = plazaCena.objects.get(fecha_dia=fecha_actual, plaza='Plaza4')
        p5c = plazaCena.objects.get(fecha_dia=fecha_actual, plaza='Plaza5')
        p6c = plazaCena.objects.get(fecha_dia=fecha_actual, plaza='Plaza6')
        p7c = plazaCena.objects.get(fecha_dia=fecha_actual, plaza='Plaza7')
        p8c = plazaCena.objects.get(fecha_dia=fecha_actual, plaza='Plaza8')
        p9c = plazaCena.objects.get(fecha_dia=fecha_actual, plaza='Plaza9')
        p10c = plazaCena.objects.get(fecha_dia=fecha_actual, plaza='Plaza10')
        anfi_nombrec = plazaCena.objects.get(fecha_dia=fecha_actual, plaza='Plaza1')

        plaza1c = p1c
        plaza2c = p2c
        plaza3c = p3c
        plaza4c = p4c
        plaza5c = p5c
        plaza6c = p6c
        plaza7c = p7c
        plaza8c = p8c
        plaza9c = p9c
        plaza10c = p10c
        nombre_anfic = anfi_nombrec.nombre_anfitriona
        mozoplaza1c = plaza1c.mozo_nombre
        mozoplaza2c = plaza2c.mozo_nombre
        mozoplaza3c = plaza3c.mozo_nombre
        mozoplaza4c = plaza4c.mozo_nombre
        mozoplaza5c = plaza5c.mozo_nombre
        mozoplaza6c = plaza6c.mozo_nombre
        mozoplaza7c = plaza7c.mozo_nombre
        mozoplaza8c = plaza8c.mozo_nombre
        mozoplaza9c = plaza9c.mozo_nombre
        mozoplaza10c = plaza10c.mozo_nombre
    else:
        estadoc = 2
        mozoplaza1c = ""
        mozoplaza2c = ""
        mozoplaza3c = ""
        mozoplaza4c= ""
        mozoplaza5c = ""
        mozoplaza6c = ""
        mozoplaza7c = ""
        mozoplaza8c = ""
        mozoplaza9c = ""
        mozoplaza10c = ""
        nombre_anfic = ""

    mozoz_pc = mozosPlaza.objects.all()
    anfitrionasc = anfitriona.objects.all()

    # plaza1
    mesa1 = nuevaReserva.objects.filter(
        fechaReserva=fecha_actual, mesa_asignadaa='1', hora__lt='16:01').count()
    mesa2 = nuevaReserva.objects.filter(
        fechaReserva=fecha_actual, mesa_asignadaa='2', hora__lt='16:01').count()
    mesa3 = nuevaReserva.objects.filter(
        fechaReserva=fecha_actual, mesa_asignadaa='3', hora__lt='16:01').count()
    mesa6 = nuevaReserva.objects.filter(
        fechaReserva=fecha_actual, mesa_asignadaa='6', hora__lt='16:01').count()
    mesa6A = nuevaReserva.objects.filter(
        fechaReserva=fecha_actual, mesa_asignadaa='6A', hora__lt='16:01').count()
    tota_plaza1 = mesa1 + mesa2 + mesa3 + mesa6 + mesa6A

    # plaza2
    mesa4 = nuevaReserva.objects.filter(
        fechaReserva=fecha_actual, mesa_asignadaa='4', hora__lt='16:01').count()
    mesa5 = nuevaReserva.objects.filter(
        fechaReserva=fecha_actual, mesa_asignadaa='5', hora__lt='16:01').count()
    mesa10 = nuevaReserva.objects.filter(
        fechaReserva=fecha_actual, mesa_asignadaa='10', hora__lt='16:01').count()
    mesa11 = nuevaReserva.objects.filter(
        fechaReserva=fecha_actual, mesa_asignadaa='11', hora__lt='16:01').count()
    mesa12 = nuevaReserva.objects.filter(
        fechaReserva=fecha_actual, mesa_asignadaa='12', hora__lt='16:01').count()
    tota_plaza2 = mesa4 + mesa5 + mesa10 + mesa11 + mesa12

    # plaza3
    mesa12a = nuevaReserva.objects.filter(
        fechaReserva=fecha_actual, mesa_asignadaa='12A', hora__lt='16:01').count()
    mesa14 = nuevaReserva.objects.filter(
        fechaReserva=fecha_actual, mesa_asignadaa='14', hora__lt='16:01').count()
    mesa15 = nuevaReserva.objects.filter(
        fechaReserva=fecha_actual, mesa_asignadaa='15', hora__lt='16:01').count()
    mesa16 = nuevaReserva.objects.filter(
        fechaReserva=fecha_actual, mesa_asignadaa='16', hora__lt='16:01').count()

    tota_plaza3 = mesa12a + mesa14 + mesa15 + mesa16

    # plaza4
    mesa17 = nuevaReserva.objects.filter(
        fechaReserva=fecha_actual, mesa_asignadaa='17', hora__lt='16:01').count()
    mesa18 = nuevaReserva.objects.filter(
        fechaReserva=fecha_actual, mesa_asignadaa='18', hora__lt='16:01').count()
    mesa8 = nuevaReserva.objects.filter(
        fechaReserva=fecha_actual, mesa_asignadaa='8', hora__lt='16:01').count()
    mesa9 = nuevaReserva.objects.filter(
        fechaReserva=fecha_actual, mesa_asignadaa='9', hora__lt='16:01').count()
    mesa34 = nuevaReserva.objects.filter(
        fechaReserva=fecha_actual, mesa_asignadaa='34', hora__lt='16:01').count()
    tota_plaza4 = mesa17 + mesa18 + mesa8 + mesa9 + mesa34

    # plaza5
    mesa26 = nuevaReserva.objects.filter(
        fechaReserva=fecha_actual, mesa_asignadaa='26', hora__lt='16:01').count()
    mesa19 = nuevaReserva.objects.filter(
        fechaReserva=fecha_actual, mesa_asignadaa='19', hora__lt='16:01').count()
    mesa20 = nuevaReserva.objects.filter(
        fechaReserva=fecha_actual, mesa_asignadaa='20', hora__lt='16:01').count()
    mesa21 = nuevaReserva.objects.filter(
        fechaReserva=fecha_actual, mesa_asignadaa='21', hora__lt='16:01').count()

    tota_plaza5 = mesa26 + mesa19 + mesa20 + mesa21

    # plaza6
    mesa22 = nuevaReserva.objects.filter(
        fechaReserva=fecha_actual, mesa_asignadaa='22', hora__lt='16:01').count()
    mesa30 = nuevaReserva.objects.filter(
        fechaReserva=fecha_actual, mesa_asignadaa='30', hora__lt='16:01').count()
    mesa31 = nuevaReserva.objects.filter(
        fechaReserva=fecha_actual, mesa_asignadaa='31', hora__lt='16:01').count()
    mesa32 = nuevaReserva.objects.filter(
        fechaReserva=fecha_actual, mesa_asignadaa='32', hora__lt='16:01').count()

    tota_plaza6 = mesa22 + mesa30 + mesa31 + mesa32

    # plaza7
    mesa27 = nuevaReserva.objects.filter(
        fechaReserva=fecha_actual, mesa_asignadaa='27', hora__lt='16:01').count()
    mesa28 = nuevaReserva.objects.filter(
        fechaReserva=fecha_actual, mesa_asignadaa='28', hora__lt='16:01').count()
    mesa29 = nuevaReserva.objects.filter(
        fechaReserva=fecha_actual, mesa_asignadaa='29', hora__lt='16:01').count()
    mesab3 = nuevaReserva.objects.filter(
        fechaReserva=fecha_actual, mesa_asignadaa='B3', hora__lt='16:01').count()

    tota_plaza7 = mesa27 + mesa28 + mesa29 + mesab3

    # plaza8
    mesa35 = nuevaReserva.objects.filter(
        fechaReserva=fecha_actual, mesa_asignadaa='35', hora__lt='16:01').count()
    mesa36 = nuevaReserva.objects.filter(
        fechaReserva=fecha_actual, mesa_asignadaa='36', hora__lt='16:01').count()
    mesa37 = nuevaReserva.objects.filter(
        fechaReserva=fecha_actual, mesa_asignadaa='37', hora__lt='16:01').count()
    mesa38 = nuevaReserva.objects.filter(
        fechaReserva=fecha_actual, mesa_asignadaa='38', hora__lt='16:01').count()
    mesa39 = nuevaReserva.objects.filter(
        fechaReserva=fecha_actual, mesa_asignadaa='39', hora__lt='16:01').count()
    tota_plaza8 = mesa35 + mesa36 + mesa37 + mesa38 + mesa39

    # plaza9
    mesajp = nuevaReserva.objects.filter(
        fechaReserva=fecha_actual, mesa_asignadaa='JP', hora__lt='16:01').count()
    tota_plaza9 = mesajp

    # plaza10
    mesabelua = nuevaReserva.objects.filter(
        fechaReserva=fecha_actual, mesa_asignadaa='BELUA', hora__lt='16:01').count()
    tota_plaza10 = mesabelua

    # plaza1c
    mesa1c = nuevaReserva.objects.filter(
        fechaReserva=fecha_actual, mesa_asignadaa='1', hora__gt='16:01').count()
    mesa2c = nuevaReserva.objects.filter(
        fechaReserva=fecha_actual, mesa_asignadaa='2', hora__gt='16:01').count()
    mesa3c = nuevaReserva.objects.filter(
        fechaReserva=fecha_actual, mesa_asignadaa='3', hora__gt='16:01').count()
    mesa6c = nuevaReserva.objects.filter(
        fechaReserva=fecha_actual, mesa_asignadaa='6', hora__gt='16:01').count()
    mesa6Ac = nuevaReserva.objects.filter(
        fechaReserva=fecha_actual, mesa_asignadaa='6A', hora__gt='16:01').count()
    tota_plaza1c = mesa1c + mesa2c + mesa3c + mesa6c + mesa6Ac

    # plaza2c
    mesa4c = nuevaReserva.objects.filter(
        fechaReserva=fecha_actual, mesa_asignadaa='4', hora__gt='16:01').count()
    mesa5c = nuevaReserva.objects.filter(
        fechaReserva=fecha_actual, mesa_asignadaa='5', hora__gt='16:01').count()
    mesa10c = nuevaReserva.objects.filter(
        fechaReserva=fecha_actual, mesa_asignadaa='10', hora__gt='16:01').count()
    mesa11c = nuevaReserva.objects.filter(
        fechaReserva=fecha_actual, mesa_asignadaa='11', hora__gt='16:01').count()
    mesa12c = nuevaReserva.objects.filter(
        fechaReserva=fecha_actual, mesa_asignadaa='12', hora__gt='16:01').count()
    tota_plaza2c = mesa4c + mesa5c + mesa10c + mesa11c + mesa12c

    # plaza3c
    mesa12ac = nuevaReserva.objects.filter(
        fechaReserva=fecha_actual, mesa_asignadaa='12A', hora__gt='16:01').count()
    mesa14c = nuevaReserva.objects.filter(
        fechaReserva=fecha_actual, mesa_asignadaa='14', hora__gt='16:01').count()
    mesa15c = nuevaReserva.objects.filter(
        fechaReserva=fecha_actual, mesa_asignadaa='15', hora__gt='16:01').count()
    mesa16c = nuevaReserva.objects.filter(
        fechaReserva=fecha_actual, mesa_asignadaa='16', hora__gt='16:01').count()

    tota_plaza3c = mesa12ac + mesa14c + mesa15c + mesa16c

    # plaza4c
    mesa17c = nuevaReserva.objects.filter(
        fechaReserva=fecha_actual, mesa_asignadaa='17', hora__gt='16:01').count()
    mesa18c = nuevaReserva.objects.filter(
        fechaReserva=fecha_actual, mesa_asignadaa='18', hora__gt='16:01').count()
    mesa8c = nuevaReserva.objects.filter(
        fechaReserva=fecha_actual, mesa_asignadaa='8', hora__gt='16:01').count()
    mesa9c = nuevaReserva.objects.filter(
        fechaReserva=fecha_actual, mesa_asignadaa='9', hora__gt='16:01').count()
    mesa34c = nuevaReserva.objects.filter(
        fechaReserva=fecha_actual, mesa_asignadaa='34', hora__gt='16:01').count()
    tota_plaza4c = mesa17c + mesa18c + mesa8c + mesa9c + mesa34c

    # plaza5c
    mesa26c = nuevaReserva.objects.filter(
        fechaReserva=fecha_actual, mesa_asignadaa='26', hora__gt='16:01').count()
    mesa19c = nuevaReserva.objects.filter(
        fechaReserva=fecha_actual, mesa_asignadaa='19', hora__gt='16:01').count()
    mesa20c = nuevaReserva.objects.filter(
        fechaReserva=fecha_actual, mesa_asignadaa='20', hora__gt='16:01').count()
    mesa21c = nuevaReserva.objects.filter(
        fechaReserva=fecha_actual, mesa_asignadaa='21', hora__gt='16:01').count()

    tota_plaza5c = mesa26c + mesa19c + mesa20c + mesa21c

    # plaza6c
    mesa22c = nuevaReserva.objects.filter(
        fechaReserva=fecha_actual, mesa_asignadaa='22', hora__gt='16:01').count()
    mesa30c = nuevaReserva.objects.filter(
        fechaReserva=fecha_actual, mesa_asignadaa='30', hora__gt='16:01').count()
    mesa31c = nuevaReserva.objects.filter(
        fechaReserva=fecha_actual, mesa_asignadaa='31', hora__gt='16:01').count()
    mesa32c = nuevaReserva.objects.filter(
        fechaReserva=fecha_actual, mesa_asignadaa='32', hora__gt='16:01').count()

    tota_plaza6c = mesa22c + mesa30c + mesa31c + mesa32c

    # plaza7c
    mesa27c = nuevaReserva.objects.filter(
        fechaReserva=fecha_actual, mesa_asignadaa='27', hora__gt='16:01').count()
    mesa28c = nuevaReserva.objects.filter(
        fechaReserva=fecha_actual, mesa_asignadaa='28', hora__gt='16:01').count()
    mesa29c = nuevaReserva.objects.filter(
        fechaReserva=fecha_actual, mesa_asignadaa='29', hora__gt='16:01').count()
    mesab3c = nuevaReserva.objects.filter(
        fechaReserva=fecha_actual, mesa_asignadaa='B3', hora__gt='16:01').count()

    tota_plaza7c = mesa27c + mesa28c + mesa29c + mesab3c

    # plaza8c
    mesa35c = nuevaReserva.objects.filter(
        fechaReserva=fecha_actual, mesa_asignadaa='35', hora__gt='16:01').count()
    mesa36c = nuevaReserva.objects.filter(
        fechaReserva=fecha_actual, mesa_asignadaa='36', hora__gt='16:01').count()
    mesa37c = nuevaReserva.objects.filter(
        fechaReserva=fecha_actual, mesa_asignadaa='37', hora__gt='16:01').count()
    mesa38c = nuevaReserva.objects.filter(
        fechaReserva=fecha_actual, mesa_asignadaa='38', hora__gt='16:01').count()
    mesa39c = nuevaReserva.objects.filter(
        fechaReserva=fecha_actual, mesa_asignadaa='39', hora__gt='16:01').count()
    tota_plaza8c = mesa35c + mesa36c + mesa37c + mesa38c + mesa39c

    # plaza9c
    mesajpc = nuevaReserva.objects.filter(
        fechaReserva=fecha_actual, mesa_asignadaa='JP', hora__gt='16:01').count()
    tota_plaza9c = mesajpc

    # plaza10c
    mesabeluac = nuevaReserva.objects.filter(
        fechaReserva=fecha_actual, mesa_asignadaa='BELUA', hora__gt='16:01').count()
    tota_plaza10c = mesabeluac

    ##########BUSCAR LAS MANUALES ALMUERZO#################
    if plazaAlmuerzoMan.objects.filter(fecha_dia = fecha_actual, plaza = 'Plaza1').exists():
    #plaza1
        np1 = plazaAlmuerzoMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza1')
        resultnp1 = np1.n1
        np2 = plazaAlmuerzoMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza1')
        resultnp2 = np2.n2
        np3 = plazaAlmuerzoMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza1')
        resultnp3 = np3.n3
        np6 = plazaAlmuerzoMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza1')
        resultnp6 = np6.n4
        np6A = plazaAlmuerzoMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza1')
        resultnp6A = np6A.n5

        totalnp1 = int(resultnp1) + int(resultnp2) + int(resultnp3) + int(resultnp6) + int(resultnp6A)

    #plaza2
        np4 = plazaAlmuerzoMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza2')
        resultnp4 = np4.n1
        np5 = plazaAlmuerzoMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza2')
        resultnp5 = np5.n2
        np10 = plazaAlmuerzoMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza2')
        resultnp10 = np10.n3
        np11 = plazaAlmuerzoMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza2')
        resultnp11 = np11.n4
        np12 = plazaAlmuerzoMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza2')
        resultnp12 = np12.n5

        totalnp2 = int(resultnp4) + int(resultnp5) + int(resultnp10) + int(resultnp11) + int(resultnp12)

    #plaza3
        np12A = plazaAlmuerzoMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza3')
        resultnp12A = np12A.n1
        np14 = plazaAlmuerzoMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza3')
        resultnp14 = np14.n2
        np15 = plazaAlmuerzoMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza3')
        resultnp15 = np15.n3
        np16 = plazaAlmuerzoMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza3')
        resultnp16 = np16.n4

        totalnp3 = int(resultnp12A) + int(resultnp14) + int(resultnp15) + int(resultnp16) 

    #plaza4
        np17 = plazaAlmuerzoMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza4')
        resultnp17 = np17.n1
        np18 = plazaAlmuerzoMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza4')
        resultnp18 = np18.n2
        np8 = plazaAlmuerzoMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza4')
        resultnp8 = np8.n3
        np9 = plazaAlmuerzoMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza4')
        resultnp9 = np9.n4
        np34 = plazaAlmuerzoMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza4')
        resultnp34 = np34.n5

        totalnp4 = int(resultnp17) + int(resultnp18) + int(resultnp8) + int(resultnp9) + int(resultnp34)

    #plaza5
        np26 = plazaAlmuerzoMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza5')
        resultnp26 = np26.n1
        np19 = plazaAlmuerzoMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza5')
        resultnp19 = np19.n2
        np20 = plazaAlmuerzoMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza5')
        resultnp20 = np20.n3
        np21 = plazaAlmuerzoMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza5')
        resultnp21 = np21.n4

        totalnp5 = int(resultnp26) + int(resultnp19) + int(resultnp20) + int(resultnp21) 

    #plaza6
        np22 = plazaAlmuerzoMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza6')
        resultnp22 = np22.n1
        np30 = plazaAlmuerzoMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza6')
        resultnp30 = np30.n2
        np31 = plazaAlmuerzoMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza6')
        resultnp31 = np31.n3
        np32 = plazaAlmuerzoMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza6')
        resultnp32 = np32.n4

        totalnp6 = int(resultnp22) + int(resultnp30) + int(resultnp31) + int(resultnp32) 

    #plaza7
        np27 = plazaAlmuerzoMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza7')
        resultnp27 = np27.n1
        np28 = plazaAlmuerzoMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza7')
        resultnp28 = np28.n2
        np29 = plazaAlmuerzoMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza7')
        resultnp29 = np29.n3
        npB3 = plazaAlmuerzoMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza7')
        resultnpB3 = npB3.n4

        totalnp7 = int(resultnp27) + int(resultnp28) + int(resultnp29) + int(resultnpB3) 

    #plaza8
        np35 = plazaAlmuerzoMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza8')
        resultnp35 = np35.n1
        np36 = plazaAlmuerzoMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza8')
        resultnp36 = np36.n2
        np37 = plazaAlmuerzoMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza8')
        resultnp37 = np37.n3
        np38 = plazaAlmuerzoMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza8')
        resultnp38 = np38.n4
        np39 = plazaAlmuerzoMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza8')
        resultnp39 = np39.n5

        totalnp8 = int(resultnp35) + int(resultnp36) + int(resultnp37) + int(resultnp38) + int(resultnp39)

    #plaza9
        npJP = plazaAlmuerzoMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza9')
        resultnpJP = npJP.n1 

    #plaza10
        npBELUA = plazaAlmuerzoMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza10')
        resultnpBELUA = npBELUA.n1

    else:
          resultnp1 = resultnp2 = resultnp3 = resultnp6 = resultnp6A = totalnp1 = resultnp4 = resultnp5 = resultnp10 \
            = resultnp11 = resultnp12 = totalnp2 = resultnp12A = resultnp14 = resultnp15 = resultnp16 = totalnp3 = resultnp12A \
            = resultnp14 = resultnp15 = resultnp16 = totalnp3 = resultnp17 = resultnp18 = resultnp8 = resultnp9= resultnp34\
            = totalnp4 = resultnp26 = resultnp19 = resultnp20 = resultnp21\
            = totalnp5 = resultnp22 = resultnp30 = resultnp31 = resultnp32\
            = totalnp6 = resultnp27 = resultnp28 = resultnp29 = resultnpB3\
            = totalnp7 = resultnp35 = resultnp36 = resultnp37 = resultnp38 = resultnp39\
            = totalnp8 = resultnpJP = resultnpBELUA = 0 

    ##########BUSCAR LAS MANUALES CENA#################
    if plazaCenaMan.objects.filter(fecha_dia = fecha_actual, plaza = 'Plaza1').exists():
    #plaza1
        np1c = plazaCenaMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza1')
        resultnp1c = np1c.n1
        np2c = plazaCenaMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza1')
        resultnp2c = np2c.n2
        np3c = plazaCenaMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza1')
        resultnp3c = np3c.n3
        np6c = plazaCenaMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza1')
        resultnp6c = np6c.n4
        np6Ac = plazaCenaMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza1')
        resultnp6Ac = np6Ac.n5

        totalnp1c = int(resultnp1c) + int(resultnp2c) + int(resultnp3c) + int(resultnp6c) + int(resultnp6Ac)

    #plaza2
        np4c = plazaCenaMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza2')
        resultnp4c = np4c.n1
        np5c = plazaCenaMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza2')
        resultnp5c = np5c.n2
        np10c = plazaCenaMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza2')
        resultnp10c = np10c.n3
        np11c = plazaCenaMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza2')
        resultnp11c = np11c.n4
        np12c = plazaCenaMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza2')
        resultnp12c = np12c.n5

        totalnp2c = int(resultnp4c) + int(resultnp5c) + int(resultnp10c) + int(resultnp11c) + int(resultnp12c)

    #plaza3
        np12Ac = plazaCenaMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza3')
        resultnp12Ac = np12Ac.n1
        np14c = plazaCenaMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza3')
        resultnp14c = np14c.n2
        np15c = plazaCenaMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza3')
        resultnp15c = np15c.n3
        np16c = plazaCenaMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza3')
        resultnp16c = np16c.n4

        totalnp3c = int(resultnp12Ac) + int(resultnp14c) + int(resultnp15c) + int(resultnp16c) 

    #plaza4
        np17c = plazaCenaMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza4')
        resultnp17c = np17c.n1
        np18c = plazaCenaMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza4')
        resultnp18c = np18c.n2
        np8c = plazaCenaMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza4')
        resultnp8c = np8c.n3
        np9c = plazaCenaMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza4')
        resultnp9c = np9c.n4
        np34c = plazaCenaMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza4')
        resultnp34c = np34c.n5

        totalnp4c = int(resultnp17c) + int(resultnp18c) + int(resultnp8c) + int(resultnp9c) + int(resultnp34c)

    #plaza5
        np26c = plazaCenaMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza5')
        resultnp26c = np26c.n1
        np19c = plazaCenaMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza5')
        resultnp19c = np19c.n2
        np20c = plazaCenaMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza5')
        resultnp20c = np20c.n3
        np21c = plazaCenaMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza5')
        resultnp21c = np21c.n4

        totalnp5c = int(resultnp26c) + int(resultnp19c) + int(resultnp20c) + int(resultnp21c) 

    #plaza6
        np22c = plazaCenaMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza6')
        resultnp22c = np22c.n1
        np30c = plazaCenaMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza6')
        resultnp30c = np30c.n2
        np31c = plazaCenaMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza6')
        resultnp31c = np31c.n3
        np32c = plazaCenaMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza6')
        resultnp32c = np32c.n4

        totalnp6c = int(resultnp22c) + int(resultnp30c) + int(resultnp31c) + int(resultnp32c) 

    #plaza7
        np27c = plazaCenaMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza7')
        resultnp27c = np27c.n1
        np28c = plazaCenaMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza7')
        resultnp28c = np28c.n2
        np29c = plazaCenaMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza7')
        resultnp29c = np29c.n3
        npB3c = plazaCenaMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza7')
        resultnpB3c = npB3c.n4

        totalnp7c = int(resultnp27c) + int(resultnp28c) + int(resultnp29c) + int(resultnpB3c) 

    #plaza8
        np35c = plazaCenaMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza8')
        resultnp35c = np35c.n1
        np36c = plazaCenaMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza8')
        resultnp36c = np36c.n2
        np37c = plazaCenaMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza8')
        resultnp37c = np37c.n3
        np38c = plazaCenaMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza8')
        resultnp38c = np38c.n4
        np39c = plazaCenaMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza8')
        resultnp39c = np39c.n5

        totalnp8c = int(resultnp35c) + int(resultnp36c) + int(resultnp37c) + int(resultnp38c) + int(resultnp39c)

    #plaza9
        npJPc = plazaCenaMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza9')
        resultnpJPc = npJPc.n1 

    #plaza10
        npBELUAc = plazaCenaMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza10')
        resultnpBELUAc = npBELUAc.n1
    else:
          resultnp1c = resultnp2c = resultnp3c = resultnp6c = resultnp6Ac = totalnp1c = resultnp4c = resultnp5c = resultnp10c \
            = resultnp11c = resultnp12c = totalnp2c = resultnp12Ac = resultnp14c = resultnp15c = resultnp16c = totalnp3c = resultnp12Ac \
            = resultnp14c = resultnp15c = resultnp16c = totalnp3c = resultnp17c = resultnp18c = resultnp8c = resultnp9c = resultnp34c\
            = totalnp4c = resultnp26c = resultnp19c = resultnp20c = resultnp21c\
            = totalnp5c = resultnp22c = resultnp30c = resultnp31c = resultnp32c\
            = totalnp6c = resultnp27c = resultnp28c = resultnp29c = resultnpB3c\
            = totalnp7c = resultnp35c = resultnp36c = resultnp37c = resultnp38c = resultnp39c\
            = totalnp8c = resultnpJPc = resultnpBELUAc = 0

    return render(request, 'plazas.html', {'listaMozos': mozoz_p,
                                            'meza1': mesa1, 'meza1n':resultnp1, 'meza2': mesa2,'meza2n':resultnp2, 'meza3': mesa3, 'meza3n':resultnp3,'meza6': mesa6, 'meza6n':resultnp6, 'meza6A': mesa6A, 'meza6An':resultnp6A,'totalplaza1': tota_plaza1, 'totalnp1':totalnp1,'mozoNombre': mozoplaza1, 'estado': estado,
                                          'meza4': mesa4, 'meza4n': resultnp4, 'meza5': mesa5,'meza5n': resultnp5, 'meza10': mesa10, 'meza10n': resultnp10, 'meza11': mesa11, 'meza11n': resultnp11, 'meza12': mesa12, 'meza12n': resultnp12,'totalplaza2': tota_plaza2, 'totalnp2':totalnp2, 'mozoNombre2': mozoplaza2,
                                          'meza12a': mesa12a, 'meza12An': resultnp12A, 'meza14': mesa14, 'meza14n': resultnp14, 'meza15': mesa15, 'meza15n': resultnp15, 'meza16': mesa16, 'meza16n': resultnp16, 'totalplaza3': tota_plaza3, 'totalnp3':totalnp3, 'mozoNombre3': mozoplaza3,
                                          'meza17': mesa17, 'meza17n': resultnp17, 'meza18': mesa18, 'meza18n': resultnp18, 'meza8': mesa8, 'meza8n': resultnp8, 'meza9': mesa9, 'meza9n': resultnp9, 'meza34': mesa34, 'meza34n': resultnp34,'totalplaza4': tota_plaza4, 'totalnp4':totalnp4, 'mozoNombre4': mozoplaza4,
                                          'meza26': mesa26, 'meza26n': resultnp26, 'meza19': mesa19, 'meza19n': resultnp19, 'meza20': mesa20, 'meza20n': resultnp20, 'meza21': mesa21, 'meza21n': resultnp21,'totalplaza5': tota_plaza5, 'totalnp5':totalnp5, 'mozoNombre5': mozoplaza5, 
                                          'meza22': mesa22, 'meza22n': resultnp22, 'meza30': mesa30, 'meza30n': resultnp30, 'meza31': mesa31, 'meza31n': resultnp31, 'meza32': mesa32, 'meza32n': resultnp32, 'totalplaza6': tota_plaza6, 'totalnp6':totalnp6, 'mozoNombre6': mozoplaza6,
                                          'meza27': mesa27, 'meza27n': resultnp27, 'meza28': mesa28, 'meza28n': resultnp28, 'meza29': mesa29, 'meza29n': resultnp29, 'mezab3': mesab3, 'mezaB3n': resultnpB3, 'totalplaza7': tota_plaza7, 'totalnp7':totalnp7, 'mozoNombre7': mozoplaza7,
                                          'meza35': mesa35, 'meza35n': resultnp35, 'meza36': mesa36, 'meza36n': resultnp36, 'meza37': mesa37, 'meza37n': resultnp37, 'meza38': mesa38, 'meza38n': resultnp38, 'meza39': mesa39, 'meza39n': resultnp39, 'totalplaza8': tota_plaza8, 'totalnp8':totalnp8, 'mozoNombre8': mozoplaza8,
                                          'mezajp': mesajp, 'mezaJPn': resultnpJP,'totalplaza9': tota_plaza9, 'mozoNombre9': mozoplaza9, 
                                          'mezabelua': mesabelua, 'mezaBELUAn': resultnpBELUA, 'totalplaza10': tota_plaza10, 'mozoNombre10': mozoplaza10, 
                                          'estado': estado, 'anfialmuerzo': nombre_anfi, 'listaAnfitrionas':anfitrionas, 'horaActual':hora_actual,
                                          ############cena############
                                          'listaMozosc': mozoz_pc,
                                            'meza1c': mesa1c, 'meza1nc':resultnp1c, 'meza2c': mesa2c,'meza2nc':resultnp2c, 'meza3c': mesa3c, 'meza3nc':resultnp3c,'meza6c': mesa6c, 'meza6nc':resultnp6c, 'meza6Ac': mesa6Ac, 'meza6Anc':resultnp6Ac,'totalplaza1c': tota_plaza1c, 'totalnp1c':totalnp1c,'mozoNombrec': mozoplaza1c, 'estadoc': estadoc,
                                          'meza4c': mesa4c, 'meza4nc': resultnp4c, 'meza5c': mesa5c,'meza5nc': resultnp5c, 'meza10c': mesa10c, 'meza10nc': resultnp10c, 'meza11c': mesa11c, 'meza11nc': resultnp11c, 'meza12c': mesa12c, 'meza12nc': resultnp12c,'totalplaza2c': tota_plaza2c, 'totalnp2c':totalnp2c, 'mozoNombre2c': mozoplaza2c,
                                          'meza12ac': mesa12ac, 'meza12Anc': resultnp12Ac, 'meza14c': mesa14c, 'meza14nc': resultnp14c, 'meza15c': mesa15c, 'meza15nc': resultnp15c, 'meza16c': mesa16c, 'meza16nc': resultnp16c, 'totalplaza3c': tota_plaza3c, 'totalnp3c':totalnp3c, 'mozoNombre3c': mozoplaza3c,
                                          'meza17c': mesa17c, 'meza17nc': resultnp17c, 'meza18c': mesa18c, 'meza18nc': resultnp18c, 'meza8c': mesa8c, 'meza8nc': resultnp8c, 'meza9c': mesa9c, 'meza9nc': resultnp9c, 'meza34c': mesa34c, 'meza34nc': resultnp34c,'totalplaza4c': tota_plaza4c, 'totalnp4c':totalnp4c, 'mozoNombre4c': mozoplaza4c,
                                          'meza26c': mesa26c, 'meza26nc': resultnp26c, 'meza19c': mesa19c, 'meza19nc': resultnp19c, 'meza20c': mesa20c, 'meza20nc': resultnp20c, 'meza21c': mesa21c, 'meza21nc': resultnp21c,'totalplaza5c': tota_plaza5c, 'totalnp5c':totalnp5c, 'mozoNombre5c': mozoplaza5c, 
                                          'meza22c': mesa22c, 'meza22nc': resultnp22c, 'meza30c': mesa30c, 'meza30nc': resultnp30c, 'meza31c': mesa31c, 'meza31nc': resultnp31c, 'meza32c': mesa32c, 'meza32nc': resultnp32c, 'totalplaza6c': tota_plaza6c, 'totalnp6c':totalnp6c, 'mozoNombre6c': mozoplaza6c,
                                          'meza27c': mesa27c, 'meza27nc': resultnp27c, 'meza28c': mesa28c, 'meza28nc': resultnp28c, 'meza29c': mesa29c, 'meza29nc': resultnp29c, 'mezab3c': mesab3c, 'mezaB3nc': resultnpB3c, 'totalplaza7c': tota_plaza7c, 'totalnp7c':totalnp7c, 'mozoNombre7c': mozoplaza7c,
                                          'meza35c': mesa35c, 'meza35nc': resultnp35c, 'meza36c': mesa36c, 'meza36nc': resultnp36c, 'meza37c': mesa37c, 'meza37nc': resultnp37c, 'meza38c': mesa38c, 'meza38nc': resultnp38c, 'meza39c': mesa39c, 'meza39nc': resultnp39c, 'totalplaza8c': tota_plaza8c, 'totalnp8c':totalnp8c, 'mozoNombre8c': mozoplaza8c,
                                          'mezajpc': mesajpc, 'mezaJPnc': resultnpJPc,'totalplaza9c': tota_plaza9c, 'mozoNombre9c': mozoplaza9c, 
                                          'mezabeluac': mesabeluac, 'mezaBELUAnc': resultnpBELUAc, 'totalplaza10c': tota_plaza10c, 'mozoNombre10c': mozoplaza10c, 
                                          'estadoc': estadoc, 'anfialmuerzoc': nombre_anfic, 'listaAnfitrionasc':anfitrionasc})
@login_required
def guardaPlazaAlm(request):
    fecha_actual = datetime.now().date()
    hora_actual = datetime.now().strftime('%H:%M')

    if plazaAlmuerzo.objects.filter(fecha_dia=fecha_actual).exists():
        return HttpResponse('Ya esta guardado')

    # if hora_actual < '16:01':
    #     return HttpResponse('No puede guardar la plaza hasta finalizar el turno')

    if request.method == 'POST':

        mozo = request.POST.get('mozoplaza1')
        m1 = request.POST.get('m1')
        r1 = request.POST.get('r1')
        m2 = request.POST.get('m2')
        r2 = request.POST.get('r2')
        m3 = request.POST.get('m3')
        r3 = request.POST.get('r3')
        m6 = request.POST.get('m6')
        r6 = request.POST.get('r6')
        m6a = request.POST.get('m6A')
        r6a = request.POST.get('r6A')
        total = request.POST.get('totalpz1')
        plaza = 'Plaza1'
        anfi = request.POST.get('anfitriona')

        plaza_dia = plazaAlmuerzo(m1=m1, r1=r1, m2=m2, r2=r2, m3=m3, r3=r3,
                                   m4=m6, r4=r6, m5=m6a, r5=r6a, total=total, plaza=plaza, mozo_nombre=mozo, nombre_anfitriona=anfi)

        plaza_dia.save()

        mozo = request.POST.get('mozoplaza2')
        m4 = request.POST.get('m4')
        r4 = request.POST.get('r4')
        m5 = request.POST.get('m5')
        r5 = request.POST.get('r5')
        m10 = request.POST.get('m10')
        r10 = request.POST.get('r10')
        m11 = request.POST.get('m11')
        r11 = request.POST.get('r11')
        m12 = request.POST.get('m12')
        r12 = request.POST.get('r12')
        total = request.POST.get('totalpz2')
        plaza = 'Plaza2'
        anfi = request.POST.get('anfitriona')

        plaza_dia = plazaAlmuerzo(m1=m4, r1=r4, m2=m5, r2=r5, m3=m10, r3=r10, m4=m11,
                                   r4=r11, m5=m12, r5=r12, total=total, plaza=plaza, mozo_nombre=mozo, nombre_anfitriona=anfi)

        plaza_dia.save()

        mozo = request.POST.get('mozoplaza3')
        m12a = request.POST.get('m12A')
        r12 = request.POST.get('r12')
        m14 = request.POST.get('m14')
        r14 = request.POST.get('r14')
        m15 = request.POST.get('m15')
        r15 = request.POST.get('r15')
        m16 = request.POST.get('m16')
        r16 = request.POST.get('r16')
        total = request.POST.get('totalpz3')
        plaza = 'Plaza3'
        anfi = request.POST.get('anfitriona')

        plaza_dia = plazaAlmuerzo(m1=m12a, r1=r12, m2=m14, r2=r14, m3=m15, r3=r15, m4=m16,
                                   r4=r16, total=total, plaza=plaza, mozo_nombre=mozo, nombre_anfitriona=anfi)

        plaza_dia.save()

        mozo = request.POST.get('mozoplaza4')
        m17 = request.POST.get('m17')
        r17 = request.POST.get('r17')
        m18 = request.POST.get('m18')
        r18 = request.POST.get('r18')
        m8 = request.POST.get('m8')
        r8 = request.POST.get('r8')
        m9 = request.POST.get('m9')
        r9 = request.POST.get('r9')
        m34 = request.POST.get('m34')
        r34 = request.POST.get('r34')
        total = request.POST.get('totalpz4')
        plaza = 'Plaza4'
        anfi = request.POST.get('anfitriona')

        plaza_dia = plazaAlmuerzo(m1=m17, r1=r17, m2=m18, r2=r18, m3=m8, r3=r8, m4=m9,
                                   r4=r9, m5=m34, r5=r34, total=total, plaza=plaza, mozo_nombre=mozo, nombre_anfitriona=anfi)

        plaza_dia.save()

        mozo = request.POST.get('mozoplaza5')
        m26 = request.POST.get('m26')
        r26 = request.POST.get('r26')
        m19 = request.POST.get('m19')
        r19 = request.POST.get('r19')
        m20 = request.POST.get('m20')
        r20 = request.POST.get('r20')
        m21 = request.POST.get('m21')
        r21 = request.POST.get('r21')
        total = request.POST.get('totalpz5')
        plaza = 'Plaza5'
        anfi = request.POST.get('anfitriona')

        plaza_dia = plazaAlmuerzo(m1=m26, r1=r26, m2=m19, r2=r19, m3=m20, r3=r20, m4=m21,
                                   r4=r21, total=total, plaza=plaza, mozo_nombre=mozo, nombre_anfitriona=anfi)

        plaza_dia.save()

        mozo = request.POST.get('mozoplaza6')
        m22 = request.POST.get('m22')
        r22 = request.POST.get('r22')
        m30 = request.POST.get('m30')
        r30 = request.POST.get('r30')
        m31 = request.POST.get('m31')
        r31 = request.POST.get('r31')
        m32 = request.POST.get('m32')
        r32 = request.POST.get('r32')
        total = request.POST.get('totalpz6')
        plaza = 'Plaza6'
        anfi = request.POST.get('anfitriona')

        plaza_dia = plazaAlmuerzo(m1=m22, r1=r22, m2=m30, r2=r30, m3=m31, r3=r31, m4=m32,
                                   r4=r32, total=total, plaza=plaza, mozo_nombre=mozo, nombre_anfitriona=anfi)

        plaza_dia.save()

        mozo = request.POST.get('mozoplaza7')
        m27 = request.POST.get('m27')
        r27 = request.POST.get('r27')
        m28 = request.POST.get('m28')
        r28 = request.POST.get('r28')
        m29 = request.POST.get('m29')
        r29 = request.POST.get('r29')
        mb3 = request.POST.get('mb3')
        rb3 = request.POST.get('rb3')
        total = request.POST.get('totalpz7')
        plaza = 'Plaza7'
        anfi = request.POST.get('anfitriona')

        plaza_dia = plazaAlmuerzo(m1=m27, r1=r27, m2=m28, r2=r28, m3=m29, r3=r29, m4=mb3,
                                   r4=rb3, total=total, plaza=plaza, mozo_nombre=mozo, nombre_anfitriona=anfi)

        plaza_dia.save()

        mozo = request.POST.get('mozoplaza8')
        m35 = request.POST.get('m35')
        r35 = request.POST.get('r35')
        m36 = request.POST.get('m36')
        r36 = request.POST.get('r36')
        m37 = request.POST.get('m37')
        r37 = request.POST.get('r37')
        m38 = request.POST.get('m38')
        r38 = request.POST.get('r38')
        m39 = request.POST.get('m39')
        r39 = request.POST.get('r39')
        total = request.POST.get('totalpz8')
        plaza = 'Plaza8'
        anfi = request.POST.get('anfitriona')

        plaza_dia = plazaAlmuerzo(m1=m35, r1=r35, m2=m36, r2=r36, m3=m37, r3=r37, m4=m38,
                                   r4=r38, m5=m39, r5=r39, total=total, plaza=plaza, mozo_nombre=mozo, nombre_anfitriona=anfi)

        plaza_dia.save()

        mozo = request.POST.get('mozoplaza9') + ', ' + \
                                request.POST.get('mozoplaza9b') + ' y ' + \
                                request.POST.get('mozoplaza9cm')
        mjp = request.POST.get('mjp')
        rjp = request.POST.get('rjp')
        total = request.POST.get('totalpz9')
        plaza = 'Plaza9'
        anfi = request.POST.get('anfitriona')

        plaza_dia = plazaAlmuerzo(
            m1=mjp, r1=rjp, total=total, plaza=plaza, mozo_nombre=mozo, nombre_anfitriona=anfi)

        plaza_dia.save()

        mozo = request.POST.get('mozoplaza10') + ', ' + \
                                request.POST.get('mozoplaza10b') + ' y ' + \
                                request.POST.get('mozoplaza10cm')
        mbelua = request.POST.get('mbelua')
        rbelua = request.POST.get('rbelua')
        total = request.POST.get('totalpz10')
        plaza = 'Plaza10'
        anfi = request.POST.get('anfitriona')

        plaza_dia = plazaAlmuerzo(m1=mbelua, r1=rbelua, total=total,
                                  plaza=plaza, mozo_nombre=mozo, nombre_anfitriona=anfi)

        plaza_dia.save()

        zona_horaria_peru = pytz.timezone('America/Lima')
        fecha_actual = datetime.now(zona_horaria_peru)
        es_fds = fecha_actual.weekday() in [5, 6]

        if es_fds:
            return controlMesaManAlmfds(request, 1)
        else:
            return controlMesaManAlm(request, 1)
@login_required
def guardaPlazaCena(request):
    fecha_actual = datetime.now().date()
    hora_actual = datetime.now().strftime('%H:%M')

    if plazaCena.objects.filter(fecha_dia=fecha_actual).exists():
        return HttpResponse('La plaza ya ha sido guardada')

    # if hora_actual < '23:30':
    #     return HttpResponse('No puede guardar la plaza hasta finalizar el turno')

    if request.method == 'POST':

        mozo = request.POST.get('mozoplaza1c')
        m1 = request.POST.get('m1c')
        r1 = request.POST.get('r1c')
        m2 = request.POST.get('m2c')
        r2 = request.POST.get('r2c')
        m3 = request.POST.get('m3c')
        r3 = request.POST.get('r3c')
        m6 = request.POST.get('m6c')
        r6 = request.POST.get('r6c')
        m6a = request.POST.get('m6Ac')
        r6a = request.POST.get('r6Ac')
        total = request.POST.get('totalpz1c')
        plaza = 'Plaza1'
        anfi = request.POST.get('anfitrionac')

        plaza_cena = plazaCena(m1=m1, r1=r1, m2=m2, r2=r2, m3=m3, r3=r3,
                                   m4=m6, r4=r6, m5=m6a, r5=r6a, total=total, plaza=plaza, mozo_nombre=mozo, nombre_anfitriona=anfi)

        plaza_cena.save()

        mozo = request.POST.get('mozoplaza2c')
        m4 = request.POST.get('m4c')
        r4 = request.POST.get('r4c')
        m5 = request.POST.get('m5c')
        r5 = request.POST.get('r5c')
        m10 = request.POST.get('m10c')
        r10 = request.POST.get('r10c')
        m11 = request.POST.get('m11c')
        r11 = request.POST.get('r11c')
        m12 = request.POST.get('m12c')
        r12 = request.POST.get('r12c')
        total = request.POST.get('totalpz2c')
        plaza = 'Plaza2'
        anfi = request.POST.get('anfitrionac')

        plaza_cena = plazaCena(m1=m4, r1=r4, m2=m5, r2=r5, m3=m10, r3=r10, m4=m11,
                                   r4=r11, m5=m12, r5=r12, total=total, plaza=plaza, mozo_nombre=mozo, nombre_anfitriona=anfi)

        plaza_cena.save()

        mozo = request.POST.get('mozoplaza3c')
        m12a = request.POST.get('m12Ac')
        r12 = request.POST.get('r12c')
        m14 = request.POST.get('m14c')
        r14 = request.POST.get('r14c')
        m15 = request.POST.get('m15c')
        r15 = request.POST.get('r15c')
        m16 = request.POST.get('m16c')
        r16 = request.POST.get('r16c')
        total = request.POST.get('totalpz3c')
        plaza = 'Plaza3'
        anfi = request.POST.get('anfitrionac')

        plaza_cena = plazaCena(m1=m12a, r1=r12, m2=m14, r2=r14, m3=m15, r3=r15, m4=m16,
                                   r4=r16, total=total, plaza=plaza, mozo_nombre=mozo, nombre_anfitriona=anfi)

        plaza_cena.save()

        mozo = request.POST.get('mozoplaza4c')
        m17 = request.POST.get('m17c')
        r17 = request.POST.get('r17c')
        m18 = request.POST.get('m18c')
        r18 = request.POST.get('r18c')
        m8 = request.POST.get('m8c')
        r8 = request.POST.get('r8c')
        m9 = request.POST.get('m9c')
        r9 = request.POST.get('r9c')
        m34 = request.POST.get('m34c')
        r34 = request.POST.get('r34c')
        total = request.POST.get('totalpz4c')
        plaza = 'Plaza4'
        anfi = request.POST.get('anfitrionac')

        plaza_cena = plazaCena(m1=m17, r1=r17, m2=m18, r2=r18, m3=m8, r3=r8, m4=m9,
                                   r4=r9, m5=m34, r5=r34, total=total, plaza=plaza, mozo_nombre=mozo, nombre_anfitriona=anfi)

        plaza_cena.save()

        mozo = request.POST.get('mozoplaza5c')
        m26 = request.POST.get('m26c')
        r26 = request.POST.get('r26c')
        m19 = request.POST.get('m19c')
        r19 = request.POST.get('r19c')
        m20 = request.POST.get('m20c')
        r20 = request.POST.get('r20c')
        m21 = request.POST.get('m21c')
        r21 = request.POST.get('r21c')
        total = request.POST.get('totalpz5c')
        plaza = 'Plaza5'
        anfi = request.POST.get('anfitrionac')

        plaza_cena = plazaCena(m1=m26, r1=r26, m2=m19, r2=r19, m3=m20, r3=r20, m4=m21,
                                   r4=r21, total=total, plaza=plaza, mozo_nombre=mozo, nombre_anfitriona=anfi)

        plaza_cena.save()

        mozo = request.POST.get('mozoplaza6c')
        m22 = request.POST.get('m22c')
        r22 = request.POST.get('r22c')
        m30 = request.POST.get('m30c')
        r30 = request.POST.get('r30c')
        m31 = request.POST.get('m31c')
        r31 = request.POST.get('r31c')
        m32 = request.POST.get('m32c')
        r32 = request.POST.get('r32c')
        total = request.POST.get('totalpz6c')
        plaza = 'Plaza6'
        anfi = request.POST.get('anfitrionac')

        plaza_cena = plazaCena(m1=m22, r1=r22, m2=m30, r2=r30, m3=m31, r3=r31, m4=m32,
                                   r4=r32, total=total, plaza=plaza, mozo_nombre=mozo, nombre_anfitriona=anfi)

        plaza_cena.save()

        mozo = request.POST.get('mozoplaza7c')
        m27 = request.POST.get('m27c')
        r27 = request.POST.get('r27c')
        m28 = request.POST.get('m28c')
        r28 = request.POST.get('r28c')
        m29 = request.POST.get('m29c')
        r29 = request.POST.get('r29c')
        mb3 = request.POST.get('mb3c')
        rb3 = request.POST.get('rb3c')
        total = request.POST.get('totalpz7c')
        plaza = 'Plaza7'
        anfi = request.POST.get('anfitrionac')

        plaza_cena = plazaCena(m1=m27, r1=r27, m2=m28, r2=r28, m3=m29, r3=r29, m4=mb3,
                                   r4=rb3, total=total, plaza=plaza, mozo_nombre=mozo, nombre_anfitriona=anfi)

        plaza_cena.save()

        mozo = request.POST.get('mozoplaza8c')
        m35 = request.POST.get('m35c')
        r35 = request.POST.get('r35c')
        m36 = request.POST.get('m36c')
        r36 = request.POST.get('r36c')
        m37 = request.POST.get('m37c')
        r37 = request.POST.get('r37c')
        m38 = request.POST.get('m38c')
        r38 = request.POST.get('r38c')
        m39 = request.POST.get('m39c')
        r39 = request.POST.get('r39c')
        total = request.POST.get('totalpz8c')
        plaza = 'Plaza8'
        anfi = request.POST.get('anfitrionac')

        plaza_cena = plazaCena(m1=m35, r1=r35, m2=m36, r2=r36, m3=m37, r3=r37, m4=m38,
                                   r4=r38, m5=m39, r5=r39, total=total, plaza=plaza, mozo_nombre=mozo, nombre_anfitriona=anfi)

        plaza_cena.save()

        mozo = request.POST.get('mozoplaza9c') + ', ' + \
                                request.POST.get('mozoplaza9bc') + ' y ' + \
                                request.POST.get('mozoplaza9bcn')
        mjp = request.POST.get('mjpc')
        rjp = request.POST.get('rjpc')
        total = request.POST.get('totalpz9c')
        plaza = 'Plaza9'
        anfi = request.POST.get('anfitrionac')

        plaza_cena = plazaCena(m1=mjp, r1=rjp, total=total,
                               plaza=plaza, mozo_nombre=mozo, nombre_anfitriona=anfi)

        plaza_cena.save()

        mozo = request.POST.get('mozoplaza10c') + ', ' + \
                                request.POST.get('mozoplaza10bc') + ' y ' + \
                                request.POST.get('mozoplaza10bcn') 
        mbelua = request.POST.get('mbeluac')
        rbelua = request.POST.get('rbeluac')
        total = request.POST.get('totalpz10c')
        plaza = 'Plaza10'
        anfi = request.POST.get('anfitrionac')

        plaza_cena = plazaCena(m1=mbelua, r1=rbelua, total=total,
                               plaza=plaza, mozo_nombre=mozo, nombre_anfitriona=anfi)

        plaza_cena.save()


        zona_horaria_peru = pytz.timezone('America/Lima')
        fecha_actual = datetime.now(zona_horaria_peru)
        es_fds = fecha_actual.weekday() in [5, 6]

        if es_fds:
            return controlMesaManCenfds(request, 1)
        else:
            return controlMesaManCen(request, 1)

@login_required
def verHistoricoPlaza(request):

    form = formBuscarPLaza()

    form_buscar = formBuscarPLaza(request.POST)
    if request.method == 'POST':
        if form_buscar.is_valid():
            b = form_buscar.cleaned_data['fechaPlaza']

            if plazaAlmuerzo.objects.filter(fecha_dia=b).exists():

                # plaza1
                mesa1 = nuevaReserva.objects.filter(
                    fechaReserva=b, mesa_asignadaa='1', hora__lt='16:01').count()
                mesa2 = nuevaReserva.objects.filter(
                    fechaReserva=b, mesa_asignadaa='2', hora__lt='16:01').count()
                mesa3 = nuevaReserva.objects.filter(
                    fechaReserva=b, mesa_asignadaa='3', hora__lt='16:01').count()
                mesa6 = nuevaReserva.objects.filter(
                    fechaReserva=b, mesa_asignadaa='6', hora__lt='16:01').count()
                mesa6A = nuevaReserva.objects.filter(
                    fechaReserva=b, mesa_asignadaa='6A', hora__lt='16:01').count()
                tota_plaza1 = mesa1 + mesa2 + mesa3 + mesa6 + mesa6A

                guardarp = plazaAlmuerzo.objects.get(
                    fecha_dia=b, plaza='Plaza1')

                guardarp.r1 = mesa1
                guardarp.r2 = mesa2
                guardarp.r3 = mesa3
                guardarp.r4 = mesa6
                guardarp.r5 = mesa6A
                guardarp.total = tota_plaza1
                guardarp.save()

                # plaza2
                mesa4 = nuevaReserva.objects.filter(
                    fechaReserva=b, mesa_asignadaa='4', hora__lt='16:01').count()
                mesa5 = nuevaReserva.objects.filter(
                    fechaReserva=b, mesa_asignadaa='5', hora__lt='16:01').count()
                mesa10 = nuevaReserva.objects.filter(
                    fechaReserva=b, mesa_asignadaa='10', hora__lt='16:01').count()
                mesa11 = nuevaReserva.objects.filter(
                    fechaReserva=b, mesa_asignadaa='11', hora__lt='16:01').count()
                mesa12 = nuevaReserva.objects.filter(
                    fechaReserva=b, mesa_asignadaa='12', hora__lt='16:01').count()
                tota_plaza2 = mesa4 + mesa5 + mesa10 + mesa11 + mesa12

                guardarp2 = plazaAlmuerzo.objects.get(
                    fecha_dia=b, plaza='Plaza2')

                guardarp2.r1 = mesa4
                guardarp2.r2 = mesa4
                guardarp2.r3 = mesa10
                guardarp2.r4 = mesa11
                guardarp2.r5 = mesa12
                guardarp2.total = tota_plaza2
                guardarp2.save()

                # plaza3
                mesa12A = nuevaReserva.objects.filter(
                    fechaReserva=b, mesa_asignadaa='12A', hora__lt='16:01').count()
                mesa14 = nuevaReserva.objects.filter(
                    fechaReserva=b, mesa_asignadaa='14', hora__lt='16:01').count()
                mesa15 = nuevaReserva.objects.filter(
                    fechaReserva=b, mesa_asignadaa='15', hora__lt='16:01').count()
                mesa16 = nuevaReserva.objects.filter(
                    fechaReserva=b, mesa_asignadaa='16', hora__lt='16:01').count()

                tota_plaza3 = mesa12A + mesa14 + mesa15 + mesa16

                guardarp3 = plazaAlmuerzo.objects.get(
                    fecha_dia=b, plaza='Plaza3')

                guardarp3.r1 = mesa12A
                guardarp3.r2 = mesa14
                guardarp3.r3 = mesa15
                guardarp3.r4 = mesa16
                guardarp3.total = tota_plaza3
                guardarp3.save()

                # plaza4
                mesa17 = nuevaReserva.objects.filter(
                    fechaReserva=b, mesa_asignadaa='17', hora__lt='16:01').count()
                mesa18 = nuevaReserva.objects.filter(
                    fechaReserva=b, mesa_asignadaa='18', hora__lt='16:01').count()
                mesa8 = nuevaReserva.objects.filter(
                    fechaReserva=b, mesa_asignadaa='8', hora__lt='16:01').count()
                mesa9 = nuevaReserva.objects.filter(
                    fechaReserva=b, mesa_asignadaa='9', hora__lt='16:01').count()
                mesa34 = nuevaReserva.objects.filter(
                    fechaReserva=b, mesa_asignadaa='34', hora__lt='16:01').count()

                tota_plaza4 = mesa17 + mesa18 + mesa8 + mesa9 + mesa34

                guardarp4 = plazaAlmuerzo.objects.get(
                    fecha_dia=b, plaza='Plaza4')

                guardarp4.r1 = mesa17
                guardarp4.r2 = mesa18
                guardarp4.r3 = mesa8
                guardarp4.r4 = mesa9
                guardarp4.r5 = mesa34
                guardarp4.total = tota_plaza4
                guardarp4.save()

                # plaza5
                mesa26 = nuevaReserva.objects.filter(
                    fechaReserva=b, mesa_asignadaa='26', hora__lt='16:01').count()
                mesa19 = nuevaReserva.objects.filter(
                    fechaReserva=b, mesa_asignadaa='19', hora__lt='16:01').count()
                mesa20 = nuevaReserva.objects.filter(
                    fechaReserva=b, mesa_asignadaa='20', hora__lt='16:01').count()
                mesa21 = nuevaReserva.objects.filter(
                    fechaReserva=b, mesa_asignadaa='21', hora__lt='16:01').count()

                tota_plaza5 = mesa26 + mesa19 + mesa20 + mesa21

                guardarp5 = plazaAlmuerzo.objects.get(
                    fecha_dia=b, plaza='Plaza5')

                guardarp5.r1 = mesa26
                guardarp5.r2 = mesa19
                guardarp5.r3 = mesa20
                guardarp5.r4 = mesa21
                guardarp5.total = tota_plaza5
                guardarp5.save()

                # plaza6
                mesa22 = nuevaReserva.objects.filter(
                    fechaReserva=b, mesa_asignadaa='22', hora__lt='16:01').count()
                mesa30 = nuevaReserva.objects.filter(
                    fechaReserva=b, mesa_asignadaa='30', hora__lt='16:01').count()
                mesa31 = nuevaReserva.objects.filter(
                    fechaReserva=b, mesa_asignadaa='31', hora__lt='16:01').count()
                mesa32 = nuevaReserva.objects.filter(
                    fechaReserva=b, mesa_asignadaa='32', hora__lt='16:01').count()

                tota_plaza6 = mesa22 + mesa30 + mesa31 + mesa31

                guardarp6 = plazaAlmuerzo.objects.get(
                    fecha_dia=b, plaza='Plaza6')

                guardarp6.r1 = mesa22
                guardarp6.r2 = mesa30
                guardarp6.r3 = mesa31
                guardarp6.r4 = mesa32
                guardarp6.total = tota_plaza6
                guardarp6.save()

                # plaza7
                mesa27 = nuevaReserva.objects.filter(
                    fechaReserva=b, mesa_asignadaa='27', hora__lt='16:01').count()
                mesa28 = nuevaReserva.objects.filter(
                    fechaReserva=b, mesa_asignadaa='28', hora__lt='16:01').count()
                mesa29 = nuevaReserva.objects.filter(
                    fechaReserva=b, mesa_asignadaa='29', hora__lt='16:01').count()
                mesaB3 = nuevaReserva.objects.filter(
                    fechaReserva=b, mesa_asignadaa='B3', hora__lt='16:01').count()

                tota_plaza7 = mesa27 + mesa28 + mesa29 + mesaB3

                guardarp7 = plazaAlmuerzo.objects.get(
                    fecha_dia=b, plaza='Plaza7')

                guardarp7.r1 = mesa27
                guardarp7.r2 = mesa28
                guardarp7.r3 = mesa29
                guardarp7.r4 = mesaB3
                guardarp7.total = tota_plaza7
                guardarp7.save()

                # plaza8
                mesa35 = nuevaReserva.objects.filter(
                    fechaReserva=b, mesa_asignadaa='35', hora__lt='16:01').count()
                mesa36 = nuevaReserva.objects.filter(
                    fechaReserva=b, mesa_asignadaa='36', hora__lt='16:01').count()
                mesa37 = nuevaReserva.objects.filter(
                    fechaReserva=b, mesa_asignadaa='37', hora__lt='16:01').count()
                mesa38 = nuevaReserva.objects.filter(
                    fechaReserva=b, mesa_asignadaa='38', hora__lt='16:01').count()
                mesa39 = nuevaReserva.objects.filter(
                    fechaReserva=b, mesa_asignadaa='39', hora__lt='16:01').count()

                tota_plaza8 = mesa35 + mesa36 + mesa37 + mesa38 + mesa39

                guardarp8 = plazaAlmuerzo.objects.get(
                    fecha_dia=b, plaza='Plaza8')

                guardarp8.r1 = mesa35
                guardarp8.r2 = mesa36
                guardarp8.r3 = mesa37
                guardarp8.r4 = mesa38
                guardarp8.r4 = mesa39
                guardarp8.total = tota_plaza8
                guardarp8.save()

                # plaza9
                mesajp = nuevaReserva.objects.filter(
                    fechaReserva=b, mesa_asignadaa='JP', hora__lt='16:01').count()

                tota_plaza9 = mesajp

                guardarp9 = plazaAlmuerzo.objects.get(
                    fecha_dia=b, plaza='Plaza9')

                guardarp9.r1 = mesajp
                guardarp9.total = tota_plaza9
                guardarp9.save()

                # plaza10
                mesabelua = nuevaReserva.objects.filter(
                    fechaReserva=b, mesa_asignadaa='BELUA', hora__lt='16:01').count()

                tota_plaza10 = mesabelua

                guardarp10 = plazaAlmuerzo.objects.get(
                    fecha_dia=b, plaza='Plaza10')

                guardarp10.r1 = mesabelua
                guardarp10.total = tota_plaza10
                guardarp10.save()

            if plazaCena.objects.filter(fecha_dia=b).exists():

                # plaza1
                mesa1c = nuevaReserva.objects.filter(fechaReserva=b, mesa_asignadaa='1', hora__gt='16:01').count()
                mesa2c = nuevaReserva.objects.filter(fechaReserva=b, mesa_asignadaa='2', hora__gt='16:01').count()
                mesa3c = nuevaReserva.objects.filter(fechaReserva=b, mesa_asignadaa='3', hora__gt='16:01').count()
                mesa6c = nuevaReserva.objects.filter(fechaReserva=b, mesa_asignadaa='6', hora__gt='16:01').count()
                mesa6Ac = nuevaReserva.objects.filter(fechaReserva=b, mesa_asignadaa='6A', hora__gt='16:01').count()
                tota_plaza1c = mesa1c + mesa2c + mesa3c + mesa6c + mesa6Ac

                guardarpc = plazaCena.objects.get(fecha_dia=b, plaza='Plaza1')

                guardarpc.r1 = mesa1c
                guardarpc.r2 = mesa2c
                guardarpc.r3 = mesa3c
                guardarpc.r4 = mesa6c
                guardarpc.r5 = mesa6Ac
                guardarpc.total = tota_plaza1c
                guardarpc.save()

                # plaza2
                mesa4c = nuevaReserva.objects.filter(fechaReserva=b, mesa_asignadaa='4', hora__gt='16:01').count()
                mesa5c = nuevaReserva.objects.filter(fechaReserva=b, mesa_asignadaa='5', hora__gt='16:01').count()
                mesa10c = nuevaReserva.objects.filter(fechaReserva=b, mesa_asignadaa='10', hora__gt='16:01').count()
                mesa11c = nuevaReserva.objects.filter(fechaReserva=b, mesa_asignadaa='11', hora__gt='16:01').count()
                mesa12c = nuevaReserva.objects.filter(fechaReserva=b, mesa_asignadaa='12', hora__gt='16:01').count()
                total_plaza2c = mesa4c + mesa5c + mesa10c + mesa11c + mesa12c

                guardarp2c = plazaCena.objects.get(fecha_dia=b, plaza='Plaza2')

                guardarp2c.r1 = mesa4c
                guardarp2c.r2 = mesa5c
                guardarp2c.r3 = mesa10c
                guardarp2c.r4 = mesa11c
                guardarp2c.r5 = mesa12c
                guardarp2c.total = total_plaza2c
                guardarp2c.save()

                # plaza3
                mesa12Ac = nuevaReserva.objects.filter(
                    fechaReserva=b, mesa_asignadaa='12A', hora__gt='16:01').count()
                mesa14c = nuevaReserva.objects.filter(
                    fechaReserva=b, mesa_asignadaa='14', hora__gt='16:01').count()
                mesa15c = nuevaReserva.objects.filter(
                    fechaReserva=b, mesa_asignadaa='15', hora__gt='16:01').count()
                mesa16c = nuevaReserva.objects.filter(
                    fechaReserva=b, mesa_asignadaa='16', hora__gt='16:01').count()

                tota_plaza3c = mesa12Ac + mesa14c + mesa15c + mesa16c

                guardarp3c = plazaCena.objects.get(fecha_dia=b, plaza='Plaza3')

                guardarp3c.r1 = mesa12Ac
                guardarp3c.r2 = mesa14c
                guardarp3c.r3 = mesa15c
                guardarp3c.r4 = mesa16c
                guardarp3c.total = tota_plaza3c
                guardarp3c.save()

                # plaza4
                mesa17c = nuevaReserva.objects.filter(
                    fechaReserva=b, mesa_asignadaa='17', hora__gt='16:01').count()
                mesa18c = nuevaReserva.objects.filter(
                    fechaReserva=b, mesa_asignadaa='18', hora__gt='16:01').count()
                mesa8c = nuevaReserva.objects.filter(
                    fechaReserva=b, mesa_asignadaa='8', hora__gt='16:01').count()
                mesa9c = nuevaReserva.objects.filter(
                    fechaReserva=b, mesa_asignadaa='9', hora__gt='16:01').count()
                mesa34c = nuevaReserva.objects.filter(
                    fechaReserva=b, mesa_asignadaa='34', hora__gt='16:01').count()

                tota_plaza4c = mesa17c + mesa18c + mesa8c + mesa9c + mesa34c

                guardarp4c = plazaCena.objects.get(fecha_dia=b, plaza='Plaza4')

                guardarp4c.r1 = mesa17c
                guardarp4c.r2 = mesa18c
                guardarp4c.r3 = mesa8c
                guardarp4c.r4 = mesa9c
                guardarp4c.r5 = mesa34c
                guardarp4c.total = tota_plaza4c
                guardarp4c.save()

                 # plaza5
                mesa26c = nuevaReserva.objects.filter(
                    fechaReserva=b, mesa_asignadaa='26', hora__gt='16:01').count()
                mesa19c = nuevaReserva.objects.filter(
                    fechaReserva=b, mesa_asignadaa='19', hora__gt='16:01').count()
                mesa20c = nuevaReserva.objects.filter(
                    fechaReserva=b, mesa_asignadaa='20', hora__gt='16:01').count()
                mesa21c = nuevaReserva.objects.filter(
                    fechaReserva=b, mesa_asignadaa='21', hora__gt='16:01').count()

                tota_plaza5c = mesa26c + mesa19c + mesa20c + mesa21c

                guardarp5c = plazaCena.objects.get(
                    fecha_dia=b, plaza='Plaza5')

                guardarp5c.r1 = mesa26c
                guardarp5c.r2 = mesa19c
                guardarp5c.r3 = mesa20c
                guardarp5c.r4 = mesa21c
                guardarp5c.total = tota_plaza5c
                guardarp5c.save()

                # plaza6
                mesa22c = nuevaReserva.objects.filter(
                    fechaReserva=b, mesa_asignadaa='22', hora__gt='16:01').count()
                mesa30c = nuevaReserva.objects.filter(
                    fechaReserva=b, mesa_asignadaa='30', hora__gt='16:01').count()
                mesa31c = nuevaReserva.objects.filter(
                    fechaReserva=b, mesa_asignadaa='31', hora__gt='16:01').count()
                mesa32c = nuevaReserva.objects.filter(
                    fechaReserva=b, mesa_asignadaa='32', hora__gt='16:01').count()

                tota_plaza6c = mesa22c + mesa30c + mesa31c + mesa32c

                guardarp6c = plazaCena.objects.get(
                    fecha_dia=b, plaza='Plaza6')

                guardarp6c.r1 = mesa22c
                guardarp6c.r2 = mesa30c
                guardarp6c.r3 = mesa31c
                guardarp6c.r4 = mesa32c
                guardarp6c.total = tota_plaza6c
                guardarp6c.save()

                # plaza7
                mesa27c = nuevaReserva.objects.filter(
                    fechaReserva=b, mesa_asignadaa='27', hora__gt='16:01').count()
                mesa28c = nuevaReserva.objects.filter(
                    fechaReserva=b, mesa_asignadaa='28', hora__gt='16:01').count()
                mesa29c = nuevaReserva.objects.filter(
                    fechaReserva=b, mesa_asignadaa='29', hora__gt='16:01').count()
                mesaB3c = nuevaReserva.objects.filter(
                    fechaReserva=b, mesa_asignadaa='B3', hora__gt='16:01').count()

                tota_plaza7c = mesa27c + mesa28c + mesa29c + mesaB3c

                guardarp7c = plazaCena.objects.get(
                    fecha_dia=b, plaza='Plaza7')

                guardarp7c.r1 = mesa27c
                guardarp7c.r2 = mesa28c
                guardarp7c.r3 = mesa29c
                guardarp7c.r4 = mesaB3c
                guardarp7c.total = tota_plaza7c
                guardarp7c.save()

                 # plaza8
                mesa35c = nuevaReserva.objects.filter(
                    fechaReserva=b, mesa_asignadaa='35', hora__gt='16:01').count()
                mesa36c = nuevaReserva.objects.filter(
                    fechaReserva=b, mesa_asignadaa='36', hora__gt='16:01').count()
                mesa37c = nuevaReserva.objects.filter(
                    fechaReserva=b, mesa_asignadaa='37', hora__gt='16:01').count()
                mesa38c = nuevaReserva.objects.filter(
                    fechaReserva=b, mesa_asignadaa='38', hora__gt='16:01').count()
                mesa39c = nuevaReserva.objects.filter(
                    fechaReserva=b, mesa_asignadaa='39', hora__gt='16:01').count()

                tota_plaza8c = mesa35c + mesa36c + mesa37c + mesa38c + mesa39c

                guardarp8c = plazaCena.objects.get(
                    fecha_dia=b, plaza='Plaza8')

                guardarp8c.r1 = mesa35c
                guardarp8c.r2 = mesa36c
                guardarp8c.r3 = mesa37c
                guardarp8c.r4 = mesa38c
                guardarp8c.r5 = mesa39c
                guardarp8c.total = tota_plaza8c
                guardarp8c.save()

                # plaza9
                mesajpc = nuevaReserva.objects.filter(
                    fechaReserva=b, mesa_asignadaa='JP', hora__gt='16:01').count()

                tota_plaza9c = mesajpc

                guardarp9c = plazaCena.objects.get(
                    fecha_dia=b, plaza='Plaza9')

                guardarp9c.r1 = mesajpc
                guardarp9c.total = tota_plaza9c
                guardarp9c.save()

                # plaza10
                mesabeluac = nuevaReserva.objects.filter(
                    fechaReserva=b, mesa_asignadaa='BELUA', hora__gt='16:01').count()

                tota_plaza10c = mesabeluac

                guardarp10c = plazaCena.objects.get(
                    fecha_dia=b, plaza='Plaza10')

                guardarp10c.r1 = mesabeluac
                guardarp10c.total = tota_plaza10c
                guardarp10c.save()



            bi = plazaAlmuerzo.objects.filter(fecha_dia=b)
            anfi_dia = plazaAlmuerzo.objects.filter(fecha_dia=b).first()

            if not anfi_dia is None:
                nombre_anfi_dia = anfi_dia.nombre_anfitriona
            else:
                nombre_anfi_dia = ""

            ci = plazaCena.objects.filter(fecha_dia=b)
            anfi_noche = plazaCena.objects.filter(fecha_dia=b).first()
            if not anfi_noche is None:
                nombre_anfi_cena = anfi_noche.nombre_anfitriona
            else:
                nombre_anfi_cena = ""
                            
            n_manual = plazaAlmuerzoMan.objects.filter(fecha_dia = b)
            n_manualc = plazaCenaMan.objects.filter(fecha_dia = b)  

            data = zip(n_manual,bi)
            datac = zip(ci,n_manualc)


            return render(request, 'historicoPlazas.html', {'listadoManual':n_manual,'listadoplaza': bi,'form_buscar': form, 'formBuscarInciden': form_buscar, 'listadoplazacena': ci, 'anfi_dia': nombre_anfi_dia, 'anfi_noche': nombre_anfi_cena, 'data':data, 'datac':datac})

    return render(request, 'historicoPlazas.html', {'form_buscar': form})
@login_required
def controlMesaManAlm(request, nmesa):

    fecha_actual = datetime.now().date()

    if not plazaAlmuerzo.objects.filter(fecha_dia=fecha_actual).exists():
         return redirect('verPlaza')
         

    if not plazaAlmuerzoMan.objects.filter(fecha_dia=fecha_actual).exists():

        # plaza1
        guardap1 = plazaAlmuerzoMan()
        guardap1.n1 = 0
        guardap1.n2 = 0
        guardap1.n3 = 0
        guardap1.n4 = 0
        guardap1.n5 = 0
        guardap1.plaza = "Plaza1"
        guardap1.save()
        # plaza2
        guardap2 = plazaAlmuerzoMan()
        guardap2.n1 = 0
        guardap2.n2 = 0
        guardap2.n3 = 0
        guardap2.n4 = 0
        guardap2.n5 = 0
        guardap2.plaza = "Plaza2"
        guardap2.save()
        # plaza3
        guardap3 = plazaAlmuerzoMan()
        guardap3.n1 = 0
        guardap3.n2 = 0
        guardap3.n3 = 0
        guardap3.n4 = 0
        guardap3.n5 = 0
        guardap3.plaza = "Plaza3"
        guardap3.save()
        # plaza4
        guardap4 = plazaAlmuerzoMan()
        guardap4.n1 = 0
        guardap4.n2 = 0
        guardap4.n3 = 0
        guardap4.n4 = 0
        guardap4.n5 = 0
        guardap4.plaza = "Plaza4"
        guardap4.save()
        # plaza5
        guardap5 = plazaAlmuerzoMan()
        guardap5.n1 = 0
        guardap5.n2 = 0
        guardap5.n3 = 0
        guardap5.n4 = 0
        guardap5.n5 = 0
        guardap5.plaza = "Plaza5"
        guardap5.save()
        # plaza6
        guardap6 = plazaAlmuerzoMan()
        guardap6.n1 = 0
        guardap6.n2 = 0
        guardap6.n3 = 0
        guardap6.n4 = 0
        guardap6.n5 = 0
        guardap6.plaza = "Plaza6"
        guardap6.save()
        # plaza7
        guardap7 = plazaAlmuerzoMan()
        guardap7.n1 = 0
        guardap7.n2 = 0
        guardap7.n3 = 0
        guardap7.n4 = 0
        guardap7.n5 = 0
        guardap7.plaza = "Plaza7"
        guardap7.save()
        # plaza8
        guardap8 = plazaAlmuerzoMan()
        guardap8.n1 = 0
        guardap8.n2 = 0
        guardap8.n3 = 0
        guardap8.n4 = 0
        guardap8.n5 = 0
        guardap8.plaza = "Plaza8"
        guardap8.save()
        # plaza9
        guardap9 = plazaAlmuerzoMan()
        guardap9.n1 = 0
        guardap9.n2 = 0
        guardap9.n3 = 0
        guardap9.n4 = 0
        guardap9.n5 = 0
        guardap9.plaza = "Plaza9"
        guardap9.save()
        # plaza10
        guardap10 = plazaAlmuerzoMan()
        guardap10.n1 = 0
        guardap10.n2 = 0
        guardap10.n3 = 0
        guardap10.n4 = 0
        guardap10.n5 = 0
        guardap10.plaza = "Plaza10"
        guardap10.save()

        return redirect('verPlaza')

    n_mesa = nmesa
    #plaza1
    if n_mesa == "1":
           guardan1 = plazaAlmuerzoMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza1')
           guardan1.n1 = int(guardan1.n1) + 1
           guardan1.save()
    if n_mesa == "2":
           guardan1 = plazaAlmuerzoMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza1')
           guardan1.n2 = int(guardan1.n2) + 1
           guardan1.save()
    if n_mesa == "3":
           guardan1 = plazaAlmuerzoMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza1')
           guardan1.n3 = int(guardan1.n3) + 1
           guardan1.save()
    if n_mesa == "6":
           guardan1 = plazaAlmuerzoMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza1')
           guardan1.n4 = int(guardan1.n4) + 1
           guardan1.save()
    if n_mesa == "6A":
           guardan1 = plazaAlmuerzoMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza1')
           guardan1.n5 = int(guardan1.n5) + 1
           guardan1.save()
    #plaza2
    if n_mesa == "4":
           guardan1 = plazaAlmuerzoMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza2')
           guardan1.n1 = int(guardan1.n1) + 1
           guardan1.save()
    if n_mesa == "5":
           guardan1 = plazaAlmuerzoMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza2')
           guardan1.n2 = int(guardan1.n2) + 1
           guardan1.save()
    if n_mesa == "10":
           guardan1 = plazaAlmuerzoMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza2')
           guardan1.n3 = int(guardan1.n3) + 1
           guardan1.save()
    if n_mesa == "11":
           guardan1 = plazaAlmuerzoMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza2')
           guardan1.n4 = int(guardan1.n4) + 1
           guardan1.save()
    if n_mesa == "12":
           guardan1 = plazaAlmuerzoMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza2')
           guardan1.n5 = int(guardan1.n5) + 1
           guardan1.save()
        #plaza3
    if n_mesa == "12A":
           guardan1 = plazaAlmuerzoMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza3')
           guardan1.n1 = int(guardan1.n1) + 1
           guardan1.save()
    if n_mesa == "14":
           guardan1 = plazaAlmuerzoMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza3')
           guardan1.n2 = int(guardan1.n2) + 1
           guardan1.save()
    if n_mesa == "15":
           guardan1 = plazaAlmuerzoMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza3')
           guardan1.n3 = int(guardan1.n3) + 1
           guardan1.save()
    if n_mesa == "16":
           guardan1 = plazaAlmuerzoMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza3')
           guardan1.n4 = int(guardan1.n4) + 1
           guardan1.save()
    #plaza4
    if n_mesa == "17":
           guardan1 = plazaAlmuerzoMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza4')
           guardan1.n1 = int(guardan1.n1) + 1
           guardan1.save()
    if n_mesa == "18":
           guardan1 = plazaAlmuerzoMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza4')
           guardan1.n2 = int(guardan1.n2) + 1
           guardan1.save()
    if n_mesa == "8":
           guardan1 = plazaAlmuerzoMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza4')
           guardan1.n3 = int(guardan1.n3) + 1
           guardan1.save()
    if n_mesa == "9":
           guardan1 = plazaAlmuerzoMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza4')
           guardan1.n4 = int(guardan1.n4) + 1
           guardan1.save()
    if n_mesa == "34":
           guardan1 = plazaAlmuerzoMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza4')
           guardan1.n5 = int(guardan1.n5) + 1
           guardan1.save()
    #plaza5
    if n_mesa == "26":
           guardan1 = plazaAlmuerzoMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza5')
           guardan1.n1 = int(guardan1.n1) + 1
           guardan1.save()
    if n_mesa == "19":
           guardan1 = plazaAlmuerzoMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza5')
           guardan1.n2 = int(guardan1.n2) + 1
           guardan1.save()
    if n_mesa == "20":
           guardan1 = plazaAlmuerzoMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza5')
           guardan1.n3 = int(guardan1.n3) + 1
           guardan1.save()
    if n_mesa == "21":
           guardan1 = plazaAlmuerzoMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza5')
           guardan1.n4 = int(guardan1.n4) + 1
           guardan1.save()
    #plaza6
    if n_mesa == "22":
           guardan1 = plazaAlmuerzoMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza6')
           guardan1.n1 = int(guardan1.n1) + 1
           guardan1.save()
    if n_mesa == "30":
           guardan1 = plazaAlmuerzoMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza6')
           guardan1.n2 = int(guardan1.n2) + 1
           guardan1.save()
    if n_mesa == "31":
           guardan1 = plazaAlmuerzoMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza6')
           guardan1.n3 = int(guardan1.n3) + 1
           guardan1.save()
    if n_mesa == "32":
           guardan1 = plazaAlmuerzoMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza6')
           guardan1.n4 = int(guardan1.n4) + 1
           guardan1.save()
    #plaza7
    if n_mesa == "27":
           guardan1 = plazaAlmuerzoMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza7')
           guardan1.n1 = int(guardan1.n1) + 1
           guardan1.save()
    if n_mesa == "28":
           guardan1 = plazaAlmuerzoMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza7')
           guardan1.n2 = int(guardan1.n2) + 1
           guardan1.save()
    if n_mesa == "29":
           guardan1 = plazaAlmuerzoMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza7')
           guardan1.n3 = int(guardan1.n3) + 1
           guardan1.save()
    if n_mesa == "B3":
           guardan1 = plazaAlmuerzoMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza7')
           guardan1.n4 = int(guardan1.n4) + 1
           guardan1.save()
    #plaza8
    if n_mesa == "35":
           guardan1 = plazaAlmuerzoMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza8')
           guardan1.n1 = int(guardan1.n1) + 1
           guardan1.save()
    if n_mesa == "36":
           guardan1 = plazaAlmuerzoMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza8')
           guardan1.n2 = int(guardan1.n2) + 1
           guardan1.save()
    if n_mesa == "37":
           guardan1 = plazaAlmuerzoMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza8')
           guardan1.n3 = int(guardan1.n3) + 1
           guardan1.save()
    if n_mesa == "38":
           guardan1 = plazaAlmuerzoMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza8')
           guardan1.n4 = int(guardan1.n4) + 1
           guardan1.save()
    if n_mesa == "39":
           guardan1 = plazaAlmuerzoMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza8')
           guardan1.n5 = int(guardan1.n5) + 1
           guardan1.save()
    #plaza9
    if n_mesa == "JP":
           guardan1 = plazaAlmuerzoMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza9')
           guardan1.n1 = int(guardan1.n1) + 1
           guardan1.save()
    #plaza10
    if n_mesa == "BELUA":
           guardan1 = plazaAlmuerzoMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza10')
           guardan1.n1 = int(guardan1.n1) + 1
           guardan1.save()

    return redirect('verPlaza')
@login_required
def controlMesaManCen(request, nmesa):
     
    fecha_actual = datetime.now().date()

    if not plazaCena.objects.filter(fecha_dia=fecha_actual).exists():
         return redirect('verPlaza')
         

    if not  plazaCenaMan.objects.filter(fecha_dia=fecha_actual).exists():

        # plaza1
        guardap1 =  plazaCenaMan()
        guardap1.n1 = 0
        guardap1.n2 = 0
        guardap1.n3 = 0
        guardap1.n4 = 0
        guardap1.n5 = 0
        guardap1.plaza = "Plaza1"
        guardap1.save()
        # plaza2
        guardap2 =  plazaCenaMan()
        guardap2.n1 = 0
        guardap2.n2 = 0
        guardap2.n3 = 0
        guardap2.n4 = 0
        guardap2.n5 = 0
        guardap2.plaza = "Plaza2"
        guardap2.save()
        # plaza3
        guardap3 =  plazaCenaMan()
        guardap3.n1 = 0
        guardap3.n2 = 0
        guardap3.n3 = 0
        guardap3.n4 = 0
        guardap3.n5 = 0
        guardap3.plaza = "Plaza3"
        guardap3.save()
        # plaza4
        guardap4 =  plazaCenaMan()
        guardap4.n1 = 0
        guardap4.n2 = 0
        guardap4.n3 = 0
        guardap4.n4 = 0
        guardap4.n5 = 0
        guardap4.plaza = "Plaza4"
        guardap4.save()
        # plaza5
        guardap5 =  plazaCenaMan()
        guardap5.n1 = 0
        guardap5.n2 = 0
        guardap5.n3 = 0
        guardap5.n4 = 0
        guardap5.n5 = 0
        guardap5.plaza = "Plaza5"
        guardap5.save()
        # plaza6
        guardap6 =  plazaCenaMan()
        guardap6.n1 = 0
        guardap6.n2 = 0
        guardap6.n3 = 0
        guardap6.n4 = 0
        guardap6.n5 = 0
        guardap6.plaza = "Plaza6"
        guardap6.save()
        # plaza7
        guardap7 =  plazaCenaMan()
        guardap7.n1 = 0
        guardap7.n2 = 0
        guardap7.n3 = 0
        guardap7.n4 = 0
        guardap7.n5 = 0
        guardap7.plaza = "Plaza7"
        guardap7.save()
        # plaza8
        guardap8 =  plazaCenaMan()
        guardap8.n1 = 0
        guardap8.n2 = 0
        guardap8.n3 = 0
        guardap8.n4 = 0
        guardap8.n5 = 0
        guardap8.plaza = "Plaza8"
        guardap8.save()
        # plaza9
        guardap9 =  plazaCenaMan()
        guardap9.n1 = 0
        guardap9.n2 = 0
        guardap9.n3 = 0
        guardap9.n4 = 0
        guardap9.n5 = 0
        guardap9.plaza = "Plaza9"
        guardap9.save()
        # plaza10
        guardap10 =  plazaCenaMan()
        guardap10.n1 = 0
        guardap10.n2 = 0
        guardap10.n3 = 0
        guardap10.n4 = 0
        guardap10.n5 = 0
        guardap10.plaza = "Plaza10"
        guardap10.save()

        return redirect('verPlaza')

    n_mesa = nmesa
    #plaza1
    if n_mesa == "1":
           guardan1 =  plazaCenaMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza1')
           guardan1.n1 = int(guardan1.n1) + 1
           guardan1.save()
    if n_mesa == "2":
           guardan1 =  plazaCenaMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza1')
           guardan1.n2 = int(guardan1.n2) + 1
           guardan1.save()
    if n_mesa == "3":
           guardan1 =  plazaCenaMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza1')
           guardan1.n3 = int(guardan1.n3) + 1
           guardan1.save()
    if n_mesa == "6":
           guardan1 =  plazaCenaMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza1')
           guardan1.n4 = int(guardan1.n4) + 1
           guardan1.save()
    if n_mesa == "6A":
           guardan1 =  plazaCenaMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza1')
           guardan1.n5 = int(guardan1.n5) + 1
           guardan1.save()
    #plaza2
    if n_mesa == "4":
           guardan1 =  plazaCenaMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza2')
           guardan1.n1 = int(guardan1.n1) + 1
           guardan1.save()
    if n_mesa == "5":
           guardan1 =  plazaCenaMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza2')
           guardan1.n2 = int(guardan1.n2) + 1
           guardan1.save()
    if n_mesa == "10":
           guardan1 =  plazaCenaMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza2')
           guardan1.n3 = int(guardan1.n3) + 1
           guardan1.save()
    if n_mesa == "11":
           guardan1 =  plazaCenaMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza2')
           guardan1.n4 = int(guardan1.n4) + 1
           guardan1.save()
    if n_mesa == "12":
           guardan1 =  plazaCenaMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza2')
           guardan1.n5 = int(guardan1.n5) + 1
           guardan1.save()
        #plaza3
    if n_mesa == "12A":
           guardan1 =  plazaCenaMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza3')
           guardan1.n1 = int(guardan1.n1) + 1
           guardan1.save()
    if n_mesa == "14":
           guardan1 =  plazaCenaMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza3')
           guardan1.n2 = int(guardan1.n2) + 1
           guardan1.save()
    if n_mesa == "15":
           guardan1 =  plazaCenaMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza3')
           guardan1.n3 = int(guardan1.n3) + 1
           guardan1.save()
    if n_mesa == "16":
           guardan1 =  plazaCenaMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza3')
           guardan1.n4 = int(guardan1.n4) + 1
           guardan1.save()
    #plaza4
    if n_mesa == "17":
           guardan1 =  plazaCenaMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza4')
           guardan1.n1 = int(guardan1.n1) + 1
           guardan1.save()
    if n_mesa == "18":
           guardan1 =  plazaCenaMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza4')
           guardan1.n2 = int(guardan1.n2) + 1
           guardan1.save()
    if n_mesa == "8":
           guardan1 =  plazaCenaMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza4')
           guardan1.n3 = int(guardan1.n3) + 1
           guardan1.save()
    if n_mesa == "9":
           guardan1 =  plazaCenaMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza4')
           guardan1.n4 = int(guardan1.n4) + 1
           guardan1.save()
    if n_mesa == "34":
           guardan1 =  plazaCenaMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza4')
           guardan1.n5 = int(guardan1.n5) + 1
           guardan1.save()
    #plaza5
    if n_mesa == "26":
           guardan1 =  plazaCenaMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza5')
           guardan1.n1 = int(guardan1.n1) + 1
           guardan1.save()
    if n_mesa == "19":
           guardan1 =  plazaCenaMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza5')
           guardan1.n2 = int(guardan1.n2) + 1
           guardan1.save()
    if n_mesa == "20":
           guardan1 =  plazaCenaMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza5')
           guardan1.n3 = int(guardan1.n3) + 1
           guardan1.save()
    if n_mesa == "21":
           guardan1 =  plazaCenaMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza5')
           guardan1.n4 = int(guardan1.n4) + 1
           guardan1.save()
    #plaza6
    if n_mesa == "22":
           guardan1 =  plazaCenaMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza6')
           guardan1.n1 = int(guardan1.n1) + 1
           guardan1.save()
    if n_mesa == "30":
           guardan1 =  plazaCenaMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza6')
           guardan1.n2 = int(guardan1.n2) + 1
           guardan1.save()
    if n_mesa == "31":
           guardan1 =  plazaCenaMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza6')
           guardan1.n3 = int(guardan1.n3) + 1
           guardan1.save()
    if n_mesa == "32":
           guardan1 =  plazaCenaMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza6')
           guardan1.n4 = int(guardan1.n4) + 1
           guardan1.save()
    #plaza7
    if n_mesa == "27":
           guardan1 =  plazaCenaMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza7')
           guardan1.n1 = int(guardan1.n1) + 1
           guardan1.save()
    if n_mesa == "28":
           guardan1 =  plazaCenaMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza7')
           guardan1.n2 = int(guardan1.n2) + 1
           guardan1.save()
    if n_mesa == "29":
           guardan1 =  plazaCenaMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza7')
           guardan1.n3 = int(guardan1.n3) + 1
           guardan1.save()
    if n_mesa == "B3":
           guardan1 =  plazaCenaMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza7')
           guardan1.n4 = int(guardan1.n4) + 1
           guardan1.save()
    #plaza8
    if n_mesa == "35":
           guardan1 =  plazaCenaMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza8')
           guardan1.n1 = int(guardan1.n1) + 1
           guardan1.save()
    if n_mesa == "36":
           guardan1 =  plazaCenaMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza8')
           guardan1.n2 = int(guardan1.n2) + 1
           guardan1.save()
    if n_mesa == "37":
           guardan1 =  plazaCenaMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza8')
           guardan1.n3 = int(guardan1.n3) + 1
           guardan1.save()
    if n_mesa == "38":
           guardan1 =  plazaCenaMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza8')
           guardan1.n4 = int(guardan1.n4) + 1
           guardan1.save()
    if n_mesa == "39":
           guardan1 =  plazaCenaMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza8')
           guardan1.n5 = int(guardan1.n5) + 1
           guardan1.save()
    #plaza9
    if n_mesa == "JP":
           guardan1 =  plazaCenaMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza9')
           guardan1.n1 = int(guardan1.n1) + 1
           guardan1.save()
    #plaza10
    if n_mesa == "BELUA":
           guardan1 = plazaCenaMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza10')
           guardan1.n1 = int(guardan1.n1) + 1
           guardan1.save()
    return redirect('verPlaza')
@login_required
def agregarIncidenciaLog(request):

    bi = incidenciaLog.objects.all().order_by('-fecha_incidencia')
    form = formIncidenciaLog()
    form_buscar = formBuscarIncidenciaLog()

    if request.method == 'POST':
        form = formIncidenciaLog(request.POST)
        if form.is_valid():
            form.save()
            return redirect('/libroIncidenciasLog/')
        else:

            return render(request, 'incidenciaLog.html', {'form_inciden': form, 'formBuscarInciden': form_buscar, 'listadoInciden': bi})

    return render(request, 'incidenciaLog.html', {'form_inciden': form, 'formBuscarInciden': form_buscar, 'listadoInciden': bi})
@login_required     
def buscarIncidenciasLog(request):

    form = formIncidenciaLog()
    form_buscar = formBuscarIncidenciaLog()

    form_buscar = formBuscarIncidenciaLog(request.POST)
    if request.method == 'POST':
        if form_buscar.is_valid():
            b = form_buscar.cleaned_data['fechaBIncidenciaLog']
            bi = incidenciaLog.objects.filter(
                fecha_incidencia=b).order_by('-fecha_incidencia')
            return render(request, 'incidenciaLog.html', {'form_inciden': form, 'formBuscarInciden': form_buscar, 'listadoInciden': bi})
    else:
        return redirect('/libroIncidenciasLog/')
@login_required
def agregarIncidenciaReserva(request, id):

    bi = incidencia.objects.all().order_by('-fecha_incidencia')
    form = formIncidencia()
    form_buscar = formBuscarIncidencia()
    id_reserva = id

    return render(request, 'incidenciaReserva.html', {'form_inciden': form, 'formBuscarInciden': form_buscar, 'listadoInciden': bi, 'id_reser':id_reserva})
@login_required
def guardarIncidenciaReserva(request):

    bi = incidencia.objects.all().order_by('-fecha_incidencia')
    form = formIncidencia()
    form_buscar = formBuscarIncidencia()

    if request.method == 'POST':
        form = formIncidencia(request.POST)
        if form.is_valid():
            a = form.save(commit=False)
            reserva = request.POST.get('reserva')
            a.id_reservacion = reserva
            a.save()
            return redirect('/libroIncidencias/')
        else:

            return render(request, 'incidencia.html', {'form_inciden': form, 'formBuscarInciden': form_buscar, 'listadoInciden': bi})

    return render(request, 'incidencia.html', {'form_inciden': form, 'formBuscarInciden': form_buscar, 'listadoInciden': bi})

def calculardia(request):
    zona_horaria_peru = pytz.timezone('America/Lima')
    fecha_actual = datetime.now(zona_horaria_peru)
    es_fds = fecha_actual.weekday() in [5, 6]

    if es_fds:
        return redirect('verPlazasfds')
    else:
          return redirect('verPlaza')

def verplazafds(request):
    fecha_actual = datetime.now().date()

    hora_actual = datetime.now().strftime('%H')

        ## si es sabado o domingo debe cargar otra plantilla  

    ########Almuerzo##############

    if plazaAlmuerzo.objects.filter(fecha_dia=fecha_actual).exists():
        estado = 1
        p1 = plazaAlmuerzo.objects.get(fecha_dia=fecha_actual, plaza='Plaza1')
        p2 = plazaAlmuerzo.objects.get(fecha_dia=fecha_actual, plaza='Plaza2')
        p3 = plazaAlmuerzo.objects.get(fecha_dia=fecha_actual, plaza='Plaza3')
        p4 = plazaAlmuerzo.objects.get(fecha_dia=fecha_actual, plaza='Plaza4')
        p5 = plazaAlmuerzo.objects.get(fecha_dia=fecha_actual, plaza='Plaza5')
        p6 = plazaAlmuerzo.objects.get(fecha_dia=fecha_actual, plaza='Plaza6')
        p7 = plazaAlmuerzo.objects.get(fecha_dia=fecha_actual, plaza='Plaza7')
        p8 = plazaAlmuerzo.objects.get(fecha_dia=fecha_actual, plaza='Plaza8')
        p9 = plazaAlmuerzo.objects.get(fecha_dia=fecha_actual, plaza='Plaza9')
        p10 = plazaAlmuerzo.objects.get(fecha_dia=fecha_actual, plaza='Plaza10')
        p11 = plazaAlmuerzo.objects.get(fecha_dia=fecha_actual, plaza='Plaza11')
        anfi_nombre = plazaAlmuerzo.objects.get(fecha_dia=fecha_actual, plaza='Plaza1')

        plaza1 = p1
        plaza2 = p2
        plaza3 = p3
        plaza4 = p4
        plaza5 = p5
        plaza6 = p6
        plaza7 = p7
        plaza8 = p8
        plaza9 = p9
        plaza10 = p10
        plaza11 = p11
        nombre_anfi = anfi_nombre.nombre_anfitriona
        mozoplaza1 = plaza1.mozo_nombre
        mozoplaza2 = plaza2.mozo_nombre
        mozoplaza3 = plaza3.mozo_nombre
        mozoplaza4 = plaza4.mozo_nombre
        mozoplaza5 = plaza5.mozo_nombre
        mozoplaza6 = plaza6.mozo_nombre
        mozoplaza7 = plaza7.mozo_nombre
        mozoplaza8 = plaza8.mozo_nombre
        mozoplaza9 = plaza9.mozo_nombre
        mozoplaza10 = plaza10.mozo_nombre
        mozoplaza11 = plaza11.mozo_nombre
    else:
        estado = 2
        mozoplaza1 = ""
        mozoplaza2 = ""
        mozoplaza3 = ""
        mozoplaza4 = ""
        mozoplaza5 = ""
        mozoplaza6 = ""
        mozoplaza7 = ""
        mozoplaza8 = ""
        mozoplaza9 = ""
        mozoplaza10 = ""
        mozoplaza11 = ""
        nombre_anfi = ""

    mozoz_p = mozosPlaza.objects.all()
    anfitrionas = anfitriona.objects.all()

    ########CENA##########

    if plazaCena.objects.filter(fecha_dia=fecha_actual).exists():
        estadoc = 1
        p1c = plazaCena.objects.get(fecha_dia=fecha_actual, plaza='Plaza1')
        p2c = plazaCena.objects.get(fecha_dia=fecha_actual, plaza='Plaza2')
        p3c = plazaCena.objects.get(fecha_dia=fecha_actual, plaza='Plaza3')
        p4c = plazaCena.objects.get(fecha_dia=fecha_actual, plaza='Plaza4')
        p5c = plazaCena.objects.get(fecha_dia=fecha_actual, plaza='Plaza5')
        p6c = plazaCena.objects.get(fecha_dia=fecha_actual, plaza='Plaza6')
        p7c = plazaCena.objects.get(fecha_dia=fecha_actual, plaza='Plaza7')
        p8c = plazaCena.objects.get(fecha_dia=fecha_actual, plaza='Plaza8')
        p9c = plazaCena.objects.get(fecha_dia=fecha_actual, plaza='Plaza9')
        p10c = plazaCena.objects.get(fecha_dia=fecha_actual, plaza='Plaza10')
        p11c = plazaCena.objects.get(fecha_dia=fecha_actual, plaza='Plaza11')
        anfi_nombrec = plazaCena.objects.get(fecha_dia=fecha_actual, plaza='Plaza1')

        plaza1c = p1c
        plaza2c = p2c
        plaza3c = p3c
        plaza4c = p4c
        plaza5c = p5c
        plaza6c = p6c
        plaza7c = p7c
        plaza8c = p8c
        plaza9c = p9c
        plaza10c = p10c
        plaza11c = p11c
        nombre_anfic = anfi_nombrec.nombre_anfitriona
        mozoplaza1c = plaza1c.mozo_nombre
        mozoplaza2c = plaza2c.mozo_nombre
        mozoplaza3c = plaza3c.mozo_nombre
        mozoplaza4c = plaza4c.mozo_nombre
        mozoplaza5c = plaza5c.mozo_nombre
        mozoplaza6c = plaza6c.mozo_nombre
        mozoplaza7c = plaza7c.mozo_nombre
        mozoplaza8c = plaza8c.mozo_nombre
        mozoplaza9c = plaza9c.mozo_nombre
        mozoplaza10c = plaza10c.mozo_nombre
        mozoplaza11c = plaza11c.mozo_nombre
    else:
        estadoc = 2
        mozoplaza1c = ""
        mozoplaza2c = ""
        mozoplaza3c = ""
        mozoplaza4c= ""
        mozoplaza5c = ""
        mozoplaza6c = ""
        mozoplaza7c = ""
        mozoplaza8c = ""
        mozoplaza9c = ""
        mozoplaza10c = ""
        mozoplaza11c = ""
        nombre_anfic = ""

    mozoz_pc = mozosPlaza.objects.all()
    anfitrionasc = anfitriona.objects.all()

    # plaza1
    mesa1 = nuevaReserva.objects.filter(
        fechaReserva=fecha_actual, mesa_asignadaa='1', hora__lt='16:01').count()
    mesa2 = nuevaReserva.objects.filter(
        fechaReserva=fecha_actual, mesa_asignadaa='2', hora__lt='16:01').count()
    mesa3 = nuevaReserva.objects.filter(
        fechaReserva=fecha_actual, mesa_asignadaa='3', hora__lt='16:01').count()
    mesa6 = nuevaReserva.objects.filter(
        fechaReserva=fecha_actual, mesa_asignadaa='6', hora__lt='16:01').count()
    # mesa6A = nuevaReserva.objects.filter(
    #     fechaReserva=fecha_actual, mesa_asignadaa='6A', hora__lt='16:01').count()
    tota_plaza1 = mesa1 + mesa2 + mesa3 + mesa6 

    # plaza2
    mesa4 = nuevaReserva.objects.filter(
        fechaReserva=fecha_actual, mesa_asignadaa='4', hora__lt='16:01').count()
    mesa5 = nuevaReserva.objects.filter(
        fechaReserva=fecha_actual, mesa_asignadaa='5', hora__lt='16:01').count()
    # mesa10 = nuevaReserva.objects.filter(
    #     fechaReserva=fecha_actual, mesa_asignadaa='10', hora__lt='16:01').count()
    mesa11 = nuevaReserva.objects.filter(
        fechaReserva=fecha_actual, mesa_asignadaa='11', hora__lt='16:01').count()
    mesa12 = nuevaReserva.objects.filter(
        fechaReserva=fecha_actual, mesa_asignadaa='12', hora__lt='16:01').count()
    tota_plaza2 = mesa4 + mesa5 +  mesa11 + mesa12

    # plaza3
    mesa12a = nuevaReserva.objects.filter(
        fechaReserva=fecha_actual, mesa_asignadaa='12A', hora__lt='16:01').count()
    mesa14 = nuevaReserva.objects.filter(
        fechaReserva=fecha_actual, mesa_asignadaa='14', hora__lt='16:01').count()
    mesa15 = nuevaReserva.objects.filter(
        fechaReserva=fecha_actual, mesa_asignadaa='15', hora__lt='16:01').count()
    mesa21 = nuevaReserva.objects.filter(
        fechaReserva=fecha_actual, mesa_asignadaa='21', hora__lt='16:01').count()

    tota_plaza3 = mesa12a + mesa14 + mesa15 + mesa21

    # plaza4
    mesa6A = nuevaReserva.objects.filter(
        fechaReserva=fecha_actual, mesa_asignadaa='6A', hora__lt='16:01').count()
    mesa9 = nuevaReserva.objects.filter(
        fechaReserva=fecha_actual, mesa_asignadaa='9', hora__lt='16:01').count()
    mesa10 = nuevaReserva.objects.filter(
        fechaReserva=fecha_actual, mesa_asignadaa='10', hora__lt='16:01').count()
    mesa16 = nuevaReserva.objects.filter(
        fechaReserva=fecha_actual, mesa_asignadaa='16', hora__lt='16:01').count()
    # mesa34 = nuevaReserva.objects.filter(
    #     fechaReserva=fecha_actual, mesa_asignadaa='34', hora__lt='16:01').count()
    tota_plaza4 = mesa6A + mesa9 + mesa10 + mesa16 

    # plaza5
    mesa8 = nuevaReserva.objects.filter(
        fechaReserva=fecha_actual, mesa_asignadaa='8', hora__lt='16:01').count()
    mesa17 = nuevaReserva.objects.filter(
        fechaReserva=fecha_actual, mesa_asignadaa='17', hora__lt='16:01').count()
    mesa18 = nuevaReserva.objects.filter(
        fechaReserva=fecha_actual, mesa_asignadaa='18', hora__lt='16:01').count()
    mesa34 = nuevaReserva.objects.filter(
        fechaReserva=fecha_actual, mesa_asignadaa='34', hora__lt='16:01').count()

    tota_plaza5 = mesa8 + mesa17 + mesa18 + mesa34

    # plaza6
    mesa19 = nuevaReserva.objects.filter(
        fechaReserva=fecha_actual, mesa_asignadaa='19', hora__lt='16:01').count()
    mesa20 = nuevaReserva.objects.filter(
        fechaReserva=fecha_actual, mesa_asignadaa='20', hora__lt='16:01').count()
    mesa24 = nuevaReserva.objects.filter(
        fechaReserva=fecha_actual, mesa_asignadaa='24', hora__lt='16:01').count()
    mesa26 = nuevaReserva.objects.filter(
        fechaReserva=fecha_actual, mesa_asignadaa='26', hora__lt='16:01').count()

    tota_plaza6 = mesa19 + mesa20 + mesa24 + mesa26

    # plaza8
    mesa27 = nuevaReserva.objects.filter(
        fechaReserva=fecha_actual, mesa_asignadaa='27', hora__lt='16:01').count()
    mesa28 = nuevaReserva.objects.filter(
        fechaReserva=fecha_actual, mesa_asignadaa='28', hora__lt='16:01').count()
    mesa29 = nuevaReserva.objects.filter(
        fechaReserva=fecha_actual, mesa_asignadaa='29', hora__lt='16:01').count()
    mesab3 = nuevaReserva.objects.filter(
        fechaReserva=fecha_actual, mesa_asignadaa='B3', hora__lt='16:01').count()

    tota_plaza8 = mesa27 + mesa28 + mesa29 + mesab3

    # plaza7
    mesa22 = nuevaReserva.objects.filter(
        fechaReserva=fecha_actual, mesa_asignadaa='22', hora__lt='16:01').count()
    mesa23 = nuevaReserva.objects.filter(
        fechaReserva=fecha_actual, mesa_asignadaa='23', hora__lt='16:01').count()
    mesa30 = nuevaReserva.objects.filter(
        fechaReserva=fecha_actual, mesa_asignadaa='30', hora__lt='16:01').count()
    mesa31 = nuevaReserva.objects.filter(
        fechaReserva=fecha_actual, mesa_asignadaa='31', hora__lt='16:01').count()
    mesa32 = nuevaReserva.objects.filter(
        fechaReserva=fecha_actual, mesa_asignadaa='32', hora__lt='16:01').count()
    tota_plaza7 = mesa22 + mesa23 + mesa30 + mesa31 + mesa32

    # plaza9
    mesa35 = nuevaReserva.objects.filter(
        fechaReserva=fecha_actual, mesa_asignadaa='35', hora__lt='16:01').count()
    mesa36 = nuevaReserva.objects.filter(
        fechaReserva=fecha_actual, mesa_asignadaa='36', hora__lt='16:01').count()
    mesa37 = nuevaReserva.objects.filter(
        fechaReserva=fecha_actual, mesa_asignadaa='37', hora__lt='16:01').count()
    mesa38 = nuevaReserva.objects.filter(
        fechaReserva=fecha_actual, mesa_asignadaa='38', hora__lt='16:01').count()
    mesa39 = nuevaReserva.objects.filter(
        fechaReserva=fecha_actual, mesa_asignadaa='39', hora__lt='16:01').count()
    tota_plaza9 = mesa35 + mesa36 + mesa37 + mesa38 + mesa39

    # plaza10
    mesajp = nuevaReserva.objects.filter(
        fechaReserva=fecha_actual, mesa_asignadaa='JP', hora__lt='16:01').count()
    tota_plaza10 = mesajp

    # plaza11
    mesabelua = nuevaReserva.objects.filter(
        fechaReserva=fecha_actual, mesa_asignadaa='BELUA', hora__lt='16:01').count()
    tota_plaza11 = mesabelua

    

    
    
    
    # plaza1c
    mesa1c = nuevaReserva.objects.filter(
        fechaReserva=fecha_actual, mesa_asignadaa='1', hora__gt='16:01').count()
    mesa2c = nuevaReserva.objects.filter(
        fechaReserva=fecha_actual, mesa_asignadaa='2', hora__gt='16:01').count()
    mesa3c = nuevaReserva.objects.filter(
        fechaReserva=fecha_actual, mesa_asignadaa='3', hora__gt='16:01').count()
    mesa6c = nuevaReserva.objects.filter(
        fechaReserva=fecha_actual, mesa_asignadaa='6', hora__gt='16:01').count()
    # mesa6Ac = nuevaReserva.objects.filter(
    #     fechaReserva=fecha_actual, mesa_asignadaa='6A', hora__gt='16:01').count()
    tota_plaza1c = mesa1c + mesa2c + mesa3c + mesa6c

    # plaza2c
    mesa4c = nuevaReserva.objects.filter(
        fechaReserva=fecha_actual, mesa_asignadaa='4', hora__gt='16:01').count()
    mesa5c = nuevaReserva.objects.filter(
        fechaReserva=fecha_actual, mesa_asignadaa='5', hora__gt='16:01').count()
    # mesa10c = nuevaReserva.objects.filter(
    #     fechaReserva=fecha_actual, mesa_asignadaa='10', hora__gt='16:01').count()
    mesa11c = nuevaReserva.objects.filter(
        fechaReserva=fecha_actual, mesa_asignadaa='11', hora__gt='16:01').count()
    mesa12c = nuevaReserva.objects.filter(
        fechaReserva=fecha_actual, mesa_asignadaa='12', hora__gt='16:01').count()
    tota_plaza2c = mesa4c + mesa5c + mesa11c + mesa12c


    # plaza3c
    mesa12ac = nuevaReserva.objects.filter(
        fechaReserva=fecha_actual, mesa_asignadaa='12A', hora__gt='16:01').count()
    mesa14c = nuevaReserva.objects.filter(
        fechaReserva=fecha_actual, mesa_asignadaa='14', hora__gt='16:01').count()
    mesa15c = nuevaReserva.objects.filter(
        fechaReserva=fecha_actual, mesa_asignadaa='15', hora__gt='16:01').count()
    mesa21c = nuevaReserva.objects.filter(
        fechaReserva=fecha_actual, mesa_asignadaa='21', hora__gt='16:01').count()

    tota_plaza3c = mesa12ac + mesa14c + mesa15c + mesa21c

    # plaza4c
    mesa6Ac = nuevaReserva.objects.filter(
        fechaReserva=fecha_actual, mesa_asignadaa='6A', hora__gt='16:01').count()
    mesa9c = nuevaReserva.objects.filter(
        fechaReserva=fecha_actual, mesa_asignadaa='9', hora__gt='16:01').count()
    mesa10c = nuevaReserva.objects.filter(
        fechaReserva=fecha_actual, mesa_asignadaa='10', hora__gt='16:01').count()
    mesa16c = nuevaReserva.objects.filter(
        fechaReserva=fecha_actual, mesa_asignadaa='16', hora__gt='16:01').count()
    # mesa34c = nuevaReserva.objects.filter(
    #     fechaReserva=fecha_actual, mesa_asignadaa='34', hora__gt='16:01').count()
    tota_plaza4c = mesa6Ac + mesa9c + mesa10c + mesa16c 

    # plaza5c
    mesa8c = nuevaReserva.objects.filter(
        fechaReserva=fecha_actual, mesa_asignadaa='8', hora__gt='16:01').count()
    mesa17c = nuevaReserva.objects.filter(
        fechaReserva=fecha_actual, mesa_asignadaa='17', hora__gt='16:01').count()
    mesa18c = nuevaReserva.objects.filter(
        fechaReserva=fecha_actual, mesa_asignadaa='18', hora__gt='16:01').count()
    mesa34c = nuevaReserva.objects.filter(
        fechaReserva=fecha_actual, mesa_asignadaa='34', hora__gt='16:01').count()

    tota_plaza5c = mesa8c + mesa17c + mesa18c + mesa34c

    # plaza6c
    mesa19c = nuevaReserva.objects.filter(
        fechaReserva=fecha_actual, mesa_asignadaa='19', hora__gt='16:01').count()
    mesa20c = nuevaReserva.objects.filter(
        fechaReserva=fecha_actual, mesa_asignadaa='20', hora__gt='16:01').count()
    mesa24c = nuevaReserva.objects.filter(
        fechaReserva=fecha_actual, mesa_asignadaa='24', hora__gt='16:01').count()
    mesa26c = nuevaReserva.objects.filter(
        fechaReserva=fecha_actual, mesa_asignadaa='26', hora__gt='16:01').count()

    tota_plaza6c = mesa19c + mesa20c + mesa24c + mesa26c

    # plaza8c
    mesa27c = nuevaReserva.objects.filter(
        fechaReserva=fecha_actual, mesa_asignadaa='27', hora__gt='16:01').count()
    mesa28c = nuevaReserva.objects.filter(
        fechaReserva=fecha_actual, mesa_asignadaa='28', hora__gt='16:01').count()
    mesa29c = nuevaReserva.objects.filter(
        fechaReserva=fecha_actual, mesa_asignadaa='29', hora__gt='16:01').count()
    mesab3c = nuevaReserva.objects.filter(
        fechaReserva=fecha_actual, mesa_asignadaa='B3', hora__gt='16:01').count()

    tota_plaza8c = mesa27c + mesa28c + mesa29c + mesab3c

    # plaza7c
    mesa22c = nuevaReserva.objects.filter(
        fechaReserva=fecha_actual, mesa_asignadaa='22', hora__gt='16:01').count()
    mesa23c = nuevaReserva.objects.filter(
        fechaReserva=fecha_actual, mesa_asignadaa='23', hora__gt='16:01').count()
    mesa30c = nuevaReserva.objects.filter(
        fechaReserva=fecha_actual, mesa_asignadaa='30', hora__gt='16:01').count()
    mesa31c = nuevaReserva.objects.filter(
        fechaReserva=fecha_actual, mesa_asignadaa='31', hora__gt='16:01').count()
    mesa32c = nuevaReserva.objects.filter(
        fechaReserva=fecha_actual, mesa_asignadaa='32', hora__gt='16:01').count()
    tota_plaza7c = mesa22c + mesa23c + mesa30c + mesa31c + mesa32c

    # plaza9c
    mesa35c = nuevaReserva.objects.filter(
        fechaReserva=fecha_actual, mesa_asignadaa='35', hora__gt='16:01').count()
    mesa36c = nuevaReserva.objects.filter(
        fechaReserva=fecha_actual, mesa_asignadaa='36', hora__gt='16:01').count()
    mesa37c = nuevaReserva.objects.filter(
        fechaReserva=fecha_actual, mesa_asignadaa='37', hora__gt='16:01').count()
    mesa38c = nuevaReserva.objects.filter(
        fechaReserva=fecha_actual, mesa_asignadaa='38', hora__gt='16:01').count()
    mesa39c = nuevaReserva.objects.filter(
        fechaReserva=fecha_actual, mesa_asignadaa='39', hora__gt='16:01').count()
    tota_plaza9c = mesa35c + mesa36c + mesa37c + mesa38c + mesa39c

    # plaza10c
    mesajpc = nuevaReserva.objects.filter(
        fechaReserva=fecha_actual, mesa_asignadaa='JP', hora__gt='16:01').count()
    tota_plaza10c = mesajpc

    # plaza11c
    mesabeluac = nuevaReserva.objects.filter(
        fechaReserva=fecha_actual, mesa_asignadaa='BELUA', hora__gt='16:01').count()
    tota_plaza11c = mesabeluac

    

    ##########BUSCAR LAS MANUALES ALMUERZO#################        MODIFICAR LAS PLAZAS
    if plazaAlmuerzoMan.objects.filter(fecha_dia = fecha_actual, plaza = 'Plaza1').exists():
    #plaza1
        np1 = plazaAlmuerzoMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza1')
        resultnp1 = np1.n1
        np2 = plazaAlmuerzoMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza1')
        resultnp2 = np2.n2
        np3 = plazaAlmuerzoMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza1')
        resultnp3 = np3.n3
        np6 = plazaAlmuerzoMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza1')
        resultnp6 = np6.n4
        # np6A = plazaAlmuerzoMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza1')
        # resultnp6A = np6A.n5

        totalnp1 = int(resultnp1) + int(resultnp2) + int(resultnp3) + int(resultnp6)

    #plaza2
        np4 = plazaAlmuerzoMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza2')
        resultnp4 = np4.n1
        np5 = plazaAlmuerzoMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza2')
        resultnp5 = np5.n2
        # np10 = plazaAlmuerzoMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza2')
        # resultnp10 = np10.n3
        np11 = plazaAlmuerzoMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza2')
        resultnp11 = np11.n4
        np12 = plazaAlmuerzoMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza2')
        resultnp12 = np12.n5

        totalnp2 = int(resultnp4) + int(resultnp5) + int(resultnp11) + int(resultnp12)

    #plaza3
        np12A = plazaAlmuerzoMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza3')
        resultnp12A = np12A.n1
        np14 = plazaAlmuerzoMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza3')
        resultnp14 = np14.n2
        np15 = plazaAlmuerzoMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza3')
        resultnp15 = np15.n3
        np21 = plazaAlmuerzoMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza3')
        resultnp21 = np21.n4

        totalnp3 = int(resultnp12A) + int(resultnp14) + int(resultnp15) + int(resultnp21) 

    #plaza4
        np6A = plazaAlmuerzoMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza4')
        resultnp6A = np6A.n1
        np9 = plazaAlmuerzoMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza4')
        resultnp9 = np9.n2
        np10 = plazaAlmuerzoMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza4')
        resultnp10 = np10.n3
        np16 = plazaAlmuerzoMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza4')
        resultnp16 = np16.n4
        # np34 = plazaAlmuerzoMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza4')
        # resultnp34 = np34.n5

        totalnp4 = int(resultnp6A) + int(resultnp9) + int(resultnp10) + int(resultnp16) 

    #plaza5
        np8 = plazaAlmuerzoMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza5')
        resultnp8 = np8.n1
        np17 = plazaAlmuerzoMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza5')
        resultnp17 = np17.n2
        np18 = plazaAlmuerzoMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza5')
        resultnp18 = np18.n3
        np34 = plazaAlmuerzoMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza5')
        resultnp34 = np34.n4

        totalnp5 = int(resultnp8) + int(resultnp17) + int(resultnp18) + int(resultnp34) 

    #plaza6
        np19 = plazaAlmuerzoMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza6')
        resultnp19 = np19.n1
        np20 = plazaAlmuerzoMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza6')
        resultnp20 = np20.n2
        np24 = plazaAlmuerzoMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza6')
        resultnp24 = np24.n3
        np26 = plazaAlmuerzoMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza6')
        resultnp26 = np26.n4

        totalnp6 = int(resultnp19) + int(resultnp20) + int(resultnp24) + int(resultnp26) 

    #plaza8
        np27 = plazaAlmuerzoMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza8')
        resultnp27 = np27.n1
        np28 = plazaAlmuerzoMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza8')
        resultnp28 = np28.n2
        np29 = plazaAlmuerzoMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza8')
        resultnp29 = np29.n3
        npB3 = plazaAlmuerzoMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza8')
        resultnpB3 = npB3.n4

        totalnp8 = int(resultnp27) + int(resultnp28) + int(resultnp29) + int(resultnpB3) 

    #plaza7
        np22 = plazaAlmuerzoMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza7')
        resultnp22 = np22.n1
        np23 = plazaAlmuerzoMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza7')
        resultnp23 = np23.n2
        np30 = plazaAlmuerzoMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza7')
        resultnp30 = np30.n3
        np31 = plazaAlmuerzoMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza7')
        resultnp31 = np31.n4
        np32 = plazaAlmuerzoMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza7')
        resultnp32 = np32.n5

        totalnp7 = int(resultnp22) + int(resultnp23) + int(resultnp30) + int(resultnp31) + int(resultnp32)
    
    #plaza9
        np35 = plazaAlmuerzoMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza9')
        resultnp35 = np35.n1
        np36 = plazaAlmuerzoMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza9')
        resultnp36 = np36.n2
        np37 = plazaAlmuerzoMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza9')
        resultnp37 = np37.n3
        np38 = plazaAlmuerzoMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza9')
        resultnp38 = np38.n4
        np39 = plazaAlmuerzoMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza9')
        resultnp39 = np39.n5

        totalnp9 = int(resultnp35) + int(resultnp36) + int(resultnp37) + int(resultnp38) + int(resultnp39)

    #plaza10
        npJP = plazaAlmuerzoMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza10')
        resultnpJP = npJP.n1 

    #plaza11
        npBELUA = plazaAlmuerzoMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza11')
        resultnpBELUA = npBELUA.n1
    


    else:
          resultnp1 = resultnp2 = resultnp3 = resultnp6 = resultnp6A = totalnp1 = resultnp4 = resultnp5 = resultnp10 \
            = resultnp11 = resultnp12 = totalnp2 = resultnp12A = resultnp14 = resultnp15 = resultnp16 = totalnp3 = resultnp12A \
            = resultnp14 = resultnp15 = resultnp16 = totalnp3 = resultnp17 = resultnp18 = resultnp8 = resultnp9= resultnp34\
            = totalnp4 = resultnp26 = resultnp19 = resultnp20 = resultnp21 = resultnp24\
            = totalnp5 = resultnp22 = resultnp30 = resultnp31 = resultnp32\
            = totalnp6 = resultnp27 = resultnp28 = resultnp29 = resultnpB3\
            = totalnp7 = resultnp35 = resultnp36 = resultnp37 = resultnp38 = resultnp39\
            = totalnp8 = resultnpJP = resultnpBELUA =  totalnp11 = totalnp9 = 0  
 
    ##########BUSCAR LAS MANUALES CENA#################   REEMPLAZAR POR LOS NUMEROS DE MESA SEGUN PLANTILLA 
    if plazaCenaMan.objects.filter(fecha_dia = fecha_actual, plaza = 'Plaza1').exists():
    #plaza1
        np1c = plazaCenaMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza1')
        resultnp1c = np1c.n1
        np2c = plazaCenaMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza1')
        resultnp2c = np2c.n2
        np3c = plazaCenaMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza1')
        resultnp3c = np3c.n3
        np6c = plazaCenaMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza1')
        resultnp6c = np6c.n4
        # np6Ac = plazaCenaMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza1')
        # resultnp6Ac = np6Ac.n5

        totalnp1c = int(resultnp1c) + int(resultnp2c) + int(resultnp3c) + int(resultnp6c)

    #plaza2
        np4c = plazaCenaMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza2')
        resultnp4c = np4c.n1
        np5c = plazaCenaMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza2')
        resultnp5c = np5c.n2
        # np10c = plazaCenaMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza2')
        # resultnp10c = np10c.n3
        np11c = plazaCenaMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza2')
        resultnp11c = np11c.n4
        np12c = plazaCenaMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza2')
        resultnp12c = np12c.n5

        totalnp2c = int(resultnp4c) + int(resultnp5c) + + int(resultnp11c) + int(resultnp12c)

    #plaza3
        np12Ac = plazaCenaMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza3')
        resultnp12Ac = np12Ac.n1
        np14c = plazaCenaMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza3')
        resultnp14c = np14c.n2
        np15c = plazaCenaMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza3')
        resultnp15c = np15c.n3
        np21c = plazaCenaMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza3')
        resultnp21c = np21c.n4

        totalnp3c = int(resultnp12Ac) + int(resultnp14c) + int(resultnp15c) + int(resultnp21c) 

    #plaza4
        np6Ac = plazaCenaMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza4')
        resultnp6Ac = np6Ac.n1
        np9c = plazaCenaMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza4')
        resultnp9c = np9c.n2
        np10c = plazaCenaMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza4')
        resultnp10c = np10c.n3
        np16c = plazaCenaMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza4')
        resultnp16c = np16c.n4
        # np34c = plazaCenaMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza4')
        # resultnp34c = np34c.n5

        totalnp4c = int(resultnp6Ac) + int(resultnp9c) + int(resultnp10c) + int(resultnp16c) 

    #plaza5
        np8c = plazaCenaMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza5')
        resultnp8c = np8c.n1
        np17c = plazaCenaMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza5')
        resultnp17c = np17c.n2
        np18c = plazaCenaMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza5')
        resultnp18c = np18c.n3
        np34c = plazaCenaMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza5')
        resultnp34c = np34c.n4

        totalnp5c = int(resultnp8c) + int(resultnp17c) + int(resultnp18c) + int(resultnp34c) 

    #plaza6
        np19c = plazaCenaMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza6')
        resultnp19c = np19c.n1
        np20c = plazaCenaMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza6')
        resultnp20c = np20c.n2
        np24c = plazaCenaMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza6')
        resultnp24c = np24c.n3
        np26c = plazaCenaMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza6')
        resultnp26c = np26c.n4

        totalnp6c = int(resultnp19c) + int(resultnp20c) + int(resultnp24c) + int(resultnp26c) 

    #plaza8
        np27c = plazaCenaMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza8')
        resultnp27c = np27c.n1
        np28c = plazaCenaMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza8')
        resultnp28c = np28c.n2
        np29c = plazaCenaMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza8')
        resultnp29c = np29c.n3
        npB3c = plazaCenaMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza8')
        resultnpB3c = npB3c.n4

        totalnp8c = int(resultnp27c) + int(resultnp28c) + int(resultnp29c) + int(resultnpB3c) 

    #plaza7
        np22c = plazaCenaMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza7')
        resultnp22c = np22c.n1
        np23c = plazaCenaMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza7')
        resultnp23c = np23c.n2
        np30c = plazaCenaMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza7')
        resultnp30c = np30c.n3
        np31c = plazaCenaMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza7')
        resultnp31c = np31c.n4
        np32c = plazaCenaMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza7')
        resultnp32c = np32c.n5

        totalnp7c = int(resultnp22c) + int(resultnp23c) + int(resultnp30c) + int(resultnp31c) + int(resultnp32c)
    
    #plaza9
        np35c = plazaCenaMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza9')
        resultnp35c = np35c.n1
        np36c = plazaCenaMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza9')
        resultnp36c = np36c.n2
        np37c = plazaCenaMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza9')
        resultnp37c = np37c.n3
        np38c = plazaCenaMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza9')
        resultnp38c = np38c.n4
        np39c = plazaCenaMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza9')
        resultnp39c = np39c.n5

        totalnp9c = int(resultnp35c) + int(resultnp36c) + int(resultnp37c) + int(resultnp38c) + int(resultnp39c)

    #plaza10
        npJPc = plazaCenaMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza10')
        resultnpJPc = npJPc.n1 

    #plaza11
        npBELUAc = plazaCenaMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza11')
        resultnpBELUAc = npBELUAc.n1
    
    
    else:
          resultnp1c = resultnp2c = resultnp3c = resultnp6c = resultnp6Ac = totalnp1c = resultnp4c = resultnp5c = resultnp10c \
            = resultnp11c = resultnp12c = totalnp2c = resultnp12Ac = resultnp14c = resultnp15c = resultnp16c = totalnp3c = resultnp12Ac \
            = resultnp14c = resultnp15c = resultnp16c = totalnp3c = resultnp17c = resultnp18c = resultnp8c = resultnp9c = resultnp34c\
            = totalnp4c = resultnp26c = resultnp19c = resultnp20c = resultnp21c = resultnp24c\
            = totalnp5c = resultnp22c = resultnp30c = resultnp31c = resultnp32c\
            = totalnp6c = resultnp27c = resultnp28c = resultnp29c = resultnpB3c\
            = totalnp7c = resultnp35c = resultnp36c = resultnp37c = resultnp38c = resultnp39c\
            = totalnp8c = resultnpJPc = resultnpBELUAc =  totalnp11c =  totalnp9c = 0


    return render(request, 'plazasfds.html', {'listaMozos': mozoz_p,
                                            'meza1': mesa1, 'meza1n':resultnp1, 'meza2': mesa2,'meza2n':resultnp2, 'meza3': mesa3, 'meza3n':resultnp3,'meza6': mesa6, 'meza6n':resultnp6, 'meza6A': mesa6A, 'meza6An':resultnp6A,'totalplaza1': tota_plaza1, 'totalnp1':totalnp1,'mozoNombre': mozoplaza1, 'estado': estado,
                                          'meza4': mesa4, 'meza4n': resultnp4, 'meza5': mesa5,'meza5n': resultnp5, 'meza10': mesa10, 'meza10n': resultnp10, 'meza11': mesa11, 'meza11n': resultnp11, 'meza12': mesa12, 'meza12n': resultnp12,'totalplaza2': tota_plaza2, 'totalnp2':totalnp2, 'mozoNombre2': mozoplaza2,
                                          'meza12a': mesa12a, 'meza12An': resultnp12A,'meza24': mesa24,'meza24n': resultnp24, 'meza14': mesa14, 'meza14n': resultnp14, 'meza15': mesa15, 'meza15n': resultnp15, 'meza16': mesa16, 'meza16n': resultnp16, 'totalplaza3': tota_plaza3, 'totalnp3':totalnp3, 'mozoNombre3': mozoplaza3,
                                          'meza17': mesa17, 'meza17n': resultnp17, 'meza18': mesa18, 'meza18n': resultnp18, 'meza8': mesa8, 'meza8n': resultnp8, 'meza9': mesa9, 'meza9n': resultnp9, 'meza34': mesa34, 'meza34n': resultnp34,'totalplaza4': tota_plaza4, 'totalnp4':totalnp4, 'mozoNombre4': mozoplaza4,
                                          'meza26': mesa26, 'meza26n': resultnp26, 'meza19': mesa19, 'meza19n': resultnp19, 'meza20': mesa20, 'meza20n': resultnp20, 'meza21': mesa21, 'meza21n': resultnp21,'totalplaza5': tota_plaza5, 'totalnp5':totalnp5, 'mozoNombre5': mozoplaza5, 
                                          'meza22': mesa22, 'meza22n': resultnp22, 'meza30': mesa30, 'meza30n': resultnp30, 'meza31': mesa31, 'meza31n': resultnp31, 'meza32': mesa32, 'meza32n': resultnp32, 'totalplaza6': tota_plaza6, 'totalnp6':totalnp6, 'mozoNombre6': mozoplaza6,
                                          'meza27': mesa27, 'meza27n': resultnp27, 'meza28': mesa28, 'meza28n': resultnp28, 'meza29': mesa29, 'meza29n': resultnp29, 'mezab3': mesab3, 'mezaB3n': resultnpB3, 'totalplaza7': tota_plaza7, 'totalnp7':totalnp7, 'mozoNombre7': mozoplaza7, 
                                          'meza35': mesa35, 'meza35n': resultnp35, 'meza36': mesa36, 'meza36n': resultnp36, 'meza37': mesa37, 'meza37n': resultnp37, 'meza38': mesa38, 'meza38n': resultnp38, 'meza39': mesa39, 'meza39n': resultnp39,'totalplaza8': tota_plaza8, 'totalnp8':totalnp8, 'mozoNombre8': mozoplaza8, 'mozoNombre11': mozoplaza11,
                                          'mezajp': mesajp, 'mezaJPn': resultnpJP,'totalplaza9': tota_plaza9, 'mozoNombre9': mozoplaza9, 'totalnp9': totalnp9, 
                                          'mezabelua': mesabelua, 'mezaBELUAn': resultnpBELUA, 'totalplaza10': tota_plaza10, 'mozoNombre10': mozoplaza10, 
                                          'estado': estado, 'anfialmuerzo': nombre_anfi, 'listaAnfitrionas':anfitrionas, 'horaActual':hora_actual, 'fecha_hoy':fecha_actual,
                                          ############cena############
                                          'listaMozosc': mozoz_pc,
                                            'meza1c': mesa1c, 'meza1nc':resultnp1c, 'meza2c': mesa2c,'meza2nc':resultnp2c, 'meza3c': mesa3c, 'meza3nc':resultnp3c,'meza6c': mesa6c, 'meza6nc':resultnp6c, 'meza6Ac': mesa6Ac, 'meza6Anc':resultnp6Ac,'totalplaza1c': tota_plaza1c, 'totalnp1c':totalnp1c,'mozoNombrec': mozoplaza1c, 'estadoc': estadoc,
                                          'meza4c': mesa4c, 'meza4nc': resultnp4c, 'meza5c': mesa5c,'meza5nc': resultnp5c, 'meza10c': mesa10c, 'meza10nc': resultnp10c, 'meza11c': mesa11c, 'meza11nc': resultnp11c, 'meza12c': mesa12c, 'meza12nc': resultnp12c,'totalplaza2c': tota_plaza2c, 'totalnp2c':totalnp2c, 'mozoNombre2c': mozoplaza2c,
                                          'meza12ac': mesa12ac, 'meza12Anc': resultnp12Ac,'meza24c': mesa24c,'meza24nc': resultnp24c, 'meza14c': mesa14c, 'meza14nc': resultnp14c, 'meza15c': mesa15c, 'meza15nc': resultnp15c, 'meza16c': mesa16c, 'meza16nc': resultnp16c, 'totalplaza3c': tota_plaza3c, 'totalnp3c':totalnp3c, 'mozoNombre3c': mozoplaza3c,
                                          'meza17c': mesa17c, 'meza17nc': resultnp17c, 'meza18c': mesa18c, 'meza18nc': resultnp18c, 'meza8c': mesa8c, 'meza8nc': resultnp8c, 'meza9c': mesa9c, 'meza9nc': resultnp9c, 'meza34c': mesa34c, 'meza34nc': resultnp34c,'totalplaza4c': tota_plaza4c, 'totalnp4c':totalnp4c, 'mozoNombre4c': mozoplaza4c,
                                          'meza26c': mesa26c, 'meza26nc': resultnp26c, 'meza19c': mesa19c, 'meza19nc': resultnp19c, 'meza20c': mesa20c, 'meza20nc': resultnp20c, 'meza21c': mesa21c, 'meza21nc': resultnp21c,'totalplaza5c': tota_plaza5c, 'totalnp5c':totalnp5c, 'mozoNombre5c': mozoplaza5c, 
                                          'meza22c': mesa22c, 'meza22nc': resultnp22c, 'meza30c': mesa30c, 'meza30nc': resultnp30c, 'meza31c': mesa31c, 'meza31nc': resultnp31c, 'meza32c': mesa32c, 'meza32nc': resultnp32c, 'totalplaza6c': tota_plaza6c, 'totalnp6c':totalnp6c, 'mozoNombre6c': mozoplaza6c,
                                          'meza27c': mesa27c, 'meza27nc': resultnp27c, 'meza28c': mesa28c, 'meza28nc': resultnp28c, 'meza29c': mesa29c, 'meza29nc': resultnp29c, 'mezab3c': mesab3c, 'mezaB3nc': resultnpB3c, 'totalplaza7c': tota_plaza7c, 'totalnp7c':totalnp7c, 'mozoNombre7c': mozoplaza7c, 'mozoNombre11c': mozoplaza11c,
                                          'meza35c': mesa35c, 'meza35nc': resultnp35c, 'meza36c': mesa36, 'meza36nc': resultnp36c, 'meza37c': mesa37c, 'meza37nc': resultnp37c, 'meza38c': mesa38c, 'meza38nc': resultnp38c, 'meza39c': mesa39c, 'meza39nc': resultnp39c,'totalplaza8c': tota_plaza8c, 'totalnp8c':totalnp8c, 'mozoNombre8c': mozoplaza8c,
                                          'mezajpc': mesajpc, 'mezaJPnc': resultnpJPc,'totalplaza9c': tota_plaza9c, 'mozoNombre9c': mozoplaza9c, 'totalnp9c':totalnp9c, 'totalplaza11c':tota_plaza11c,
                                          'mezabeluac': mesabeluac, 'mezaBELUAnc': resultnpBELUAc, 'totalplaza10c': tota_plaza10c, 'mozoNombre10c': mozoplaza10c, 
                                          'estadoc': estadoc, 'anfialmuerzoc': nombre_anfic, 'listaAnfitrionasc':anfitrionasc})

def controlMesaManCenfds(request, nmesa):
     
    fecha_actual = datetime.now().date()

    if not plazaCena.objects.filter(fecha_dia=fecha_actual).exists():
         return redirect('calculadia')
         

    if not  plazaCenaMan.objects.filter(fecha_dia=fecha_actual).exists():

        # plaza1
        guardap1 =  plazaCenaMan()
        guardap1.n1 = 0
        guardap1.n2 = 0
        guardap1.n3 = 0
        guardap1.n4 = 0
        guardap1.n5 = 0
        guardap1.plaza = "Plaza1"
        guardap1.save()
        # plaza2
        guardap2 =  plazaCenaMan()
        guardap2.n1 = 0
        guardap2.n2 = 0
        guardap2.n3 = 0
        guardap2.n4 = 0
        guardap2.n5 = 0
        guardap2.plaza = "Plaza2"
        guardap2.save()
        # plaza3
        guardap3 =  plazaCenaMan()
        guardap3.n1 = 0
        guardap3.n2 = 0
        guardap3.n3 = 0
        guardap3.n4 = 0
        guardap3.n5 = 0
        guardap3.plaza = "Plaza3"
        guardap3.save()
        # plaza4
        guardap4 =  plazaCenaMan()
        guardap4.n1 = 0
        guardap4.n2 = 0
        guardap4.n3 = 0
        guardap4.n4 = 0
        guardap4.n5 = 0
        guardap4.plaza = "Plaza4"
        guardap4.save()
        # plaza5
        guardap5 =  plazaCenaMan()
        guardap5.n1 = 0
        guardap5.n2 = 0
        guardap5.n3 = 0
        guardap5.n4 = 0
        guardap5.n5 = 0
        guardap5.plaza = "Plaza5"
        guardap5.save()
        # plaza6
        guardap6 =  plazaCenaMan()
        guardap6.n1 = 0
        guardap6.n2 = 0
        guardap6.n3 = 0
        guardap6.n4 = 0
        guardap6.n5 = 0
        guardap6.plaza = "Plaza6"
        guardap6.save()
        # plaza7
        guardap7 =  plazaCenaMan()
        guardap7.n1 = 0
        guardap7.n2 = 0
        guardap7.n3 = 0
        guardap7.n4 = 0
        guardap7.n5 = 0
        guardap7.plaza = "Plaza7"
        guardap7.save()
        # plaza8
        guardap8 =  plazaCenaMan()
        guardap8.n1 = 0
        guardap8.n2 = 0
        guardap8.n3 = 0
        guardap8.n4 = 0
        guardap8.n5 = 0
        guardap8.plaza = "Plaza8"
        guardap8.save()
        # plaza9
        guardap9 =  plazaCenaMan()
        guardap9.n1 = 0
        guardap9.n2 = 0
        guardap9.n3 = 0
        guardap9.n4 = 0
        guardap9.n5 = 0
        guardap9.plaza = "Plaza9"
        guardap9.save()
        # plaza10
        guardap10 =  plazaCenaMan()
        guardap10.n1 = 0
        guardap10.n2 = 0
        guardap10.n3 = 0
        guardap10.n4 = 0
        guardap10.n5 = 0
        guardap10.plaza = "Plaza10"
        guardap10.save()
        # plaza11
        guardap10 =  plazaCenaMan()
        guardap10.n1 = 0
        guardap10.n2 = 0
        guardap10.n3 = 0
        guardap10.n4 = 0
        guardap10.n5 = 0
        guardap10.plaza = "Plaza11"
        guardap10.save()

        return redirect('calculadia')

    n_mesa = nmesa
    #plaza1
    if n_mesa == "1":
           guardan1 =  plazaCenaMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza1')
           guardan1.n1 = int(guardan1.n1) + 1
           guardan1.save()
    if n_mesa == "2":
           guardan1 =  plazaCenaMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza1')
           guardan1.n2 = int(guardan1.n2) + 1
           guardan1.save()
    if n_mesa == "3":
           guardan1 =  plazaCenaMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza1')
           guardan1.n3 = int(guardan1.n3) + 1
           guardan1.save()
    if n_mesa == "6":
           guardan1 =  plazaCenaMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza1')
           guardan1.n4 = int(guardan1.n4) + 1
           guardan1.save()
    # if n_mesa == "6A":
    #        guardan1 =  plazaCenaMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza1')
    #        guardan1.n5 = int(guardan1.n5) + 1
    #        guardan1.save()
    
    #plaza2
    if n_mesa == "4":
           guardan1 =  plazaCenaMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza2')
           guardan1.n1 = int(guardan1.n1) + 1
           guardan1.save()
    if n_mesa == "5":
           guardan1 =  plazaCenaMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza2')
           guardan1.n2 = int(guardan1.n2) + 1
           guardan1.save()
    # if n_mesa == "10":
    #        guardan1 =  plazaCenaMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza2')
    #        guardan1.n3 = int(guardan1.n3) + 1
    #        guardan1.save()
    if n_mesa == "11":
           guardan1 =  plazaCenaMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza2')
           guardan1.n4 = int(guardan1.n4) + 1
           guardan1.save()
    if n_mesa == "12":
           guardan1 =  plazaCenaMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza2')
           guardan1.n5 = int(guardan1.n5) + 1
           guardan1.save()
        #plaza3
    if n_mesa == "12A":
           guardan1 =  plazaCenaMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza3')
           guardan1.n1 = int(guardan1.n1) + 1
           guardan1.save()
    if n_mesa == "14":
           guardan1 =  plazaCenaMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza3')
           guardan1.n2 = int(guardan1.n2) + 1
           guardan1.save()
    if n_mesa == "15":
           guardan1 =  plazaCenaMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza3')
           guardan1.n3 = int(guardan1.n3) + 1
           guardan1.save()
    if n_mesa == "21":
           guardan1 =  plazaCenaMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza3')
           guardan1.n4 = int(guardan1.n4) + 1
           guardan1.save()
    #plaza4
    if n_mesa == "6A":
           guardan1 =  plazaCenaMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza4')
           guardan1.n1 = int(guardan1.n1) + 1
           guardan1.save()
    if n_mesa == "9":
           guardan1 =  plazaCenaMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza4')
           guardan1.n2 = int(guardan1.n2) + 1
           guardan1.save()
    if n_mesa == "10":
           guardan1 =  plazaCenaMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza4')
           guardan1.n3 = int(guardan1.n3) + 1
           guardan1.save()
    if n_mesa == "16":
           guardan1 =  plazaCenaMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza4')
           guardan1.n4 = int(guardan1.n4) + 1
           guardan1.save()
    # if n_mesa == "34":
    #        guardan1 =  plazaCenaMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza4')
    #        guardan1.n5 = int(guardan1.n5) + 1
    #        guardan1.save()
    #plaza5
    if n_mesa == "8":
           guardan1 =  plazaCenaMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza5')
           guardan1.n1 = int(guardan1.n1) + 1
           guardan1.save()
    if n_mesa == "17":
           guardan1 =  plazaCenaMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza5')
           guardan1.n2 = int(guardan1.n2) + 1
           guardan1.save()
    if n_mesa == "18":
           guardan1 =  plazaCenaMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza5')
           guardan1.n3 = int(guardan1.n3) + 1
           guardan1.save()
    if n_mesa == "34":
           guardan1 =  plazaCenaMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza5')
           guardan1.n4 = int(guardan1.n4) + 1
           guardan1.save()
    #plaza6
    if n_mesa == "19":
           guardan1 =  plazaCenaMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza6')
           guardan1.n1 = int(guardan1.n1) + 1
           guardan1.save()
    if n_mesa == "20":
           guardan1 =  plazaCenaMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza6')
           guardan1.n2 = int(guardan1.n2) + 1
           guardan1.save()
    if n_mesa == "24":
           guardan1 =  plazaCenaMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza6')
           guardan1.n3 = int(guardan1.n3) + 1
           guardan1.save()
    if n_mesa == "26":
           guardan1 =  plazaCenaMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza6')
           guardan1.n4 = int(guardan1.n4) + 1
           guardan1.save()
    #plaza7
    if n_mesa == "27":
           guardan1 =  plazaCenaMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza8')
           guardan1.n1 = int(guardan1.n1) + 1
           guardan1.save()
    if n_mesa == "28":
           guardan1 =  plazaCenaMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza8')
           guardan1.n2 = int(guardan1.n2) + 1
           guardan1.save()
    if n_mesa == "29":
           guardan1 =  plazaCenaMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza8')
           guardan1.n3 = int(guardan1.n3) + 1
           guardan1.save()
    if n_mesa == "B3":
           guardan1 =  plazaCenaMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza8')
           guardan1.n4 = int(guardan1.n4) + 1
           guardan1.save()
    #plaza8
    if n_mesa == "22":
           guardan1 =  plazaCenaMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza7')
           guardan1.n1 = int(guardan1.n1) + 1
           guardan1.save()
    if n_mesa == "23":
           guardan1 =  plazaCenaMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza7')
           guardan1.n2 = int(guardan1.n2) + 1
           guardan1.save()
    if n_mesa == "30":
           guardan1 =  plazaCenaMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza7')
           guardan1.n3 = int(guardan1.n3) + 1
           guardan1.save()
    if n_mesa == "31":
           guardan1 =  plazaCenaMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza7')
           guardan1.n4 = int(guardan1.n4) + 1
           guardan1.save()
    if n_mesa == "32":
           guardan1 =  plazaCenaMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza7')
           guardan1.n5 = int(guardan1.n5) + 1
           guardan1.save()
    
    #plaza9
    if n_mesa == "35":
           guardan1 =  plazaCenaMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza9')
           guardan1.n1 = int(guardan1.n1) + 1
           guardan1.save()
    if n_mesa == "36":
           guardan1 =  plazaCenaMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza9')
           guardan1.n2 = int(guardan1.n2) + 1
           guardan1.save()
    if n_mesa == "37":
           guardan1 =  plazaCenaMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza9')
           guardan1.n3 = int(guardan1.n3) + 1
           guardan1.save()
    if n_mesa == "38":
           guardan1 =  plazaCenaMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza9')
           guardan1.n4 = int(guardan1.n4) + 1
           guardan1.save()
    if n_mesa == "39":
           guardan1 =  plazaCenaMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza9')
           guardan1.n5 = int(guardan1.n5) + 1
           guardan1.save()
    #plaza9
    if n_mesa == "JP":
           guardan1 =  plazaCenaMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza10')
           guardan1.n1 = int(guardan1.n1) + 1
           guardan1.save()
    #plaza10
    if n_mesa == "BELUA":
           guardan1 = plazaCenaMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza11')
           guardan1.n1 = int(guardan1.n1) + 1
           guardan1.save()
   
    return redirect('calculadia')

def controlMesaManAlmfds(request, nmesa):

    fecha_actual = datetime.now().date()

    if not plazaAlmuerzo.objects.filter(fecha_dia=fecha_actual).exists():
         return redirect('calculadia')
         

    if not plazaAlmuerzoMan.objects.filter(fecha_dia=fecha_actual).exists():

        # plaza1
        guardap1 = plazaAlmuerzoMan()
        guardap1.n1 = 0
        guardap1.n2 = 0
        guardap1.n3 = 0
        guardap1.n4 = 0
        guardap1.n5 = 0
        guardap1.plaza = "Plaza1"
        guardap1.save()
        # plaza2
        guardap2 = plazaAlmuerzoMan()
        guardap2.n1 = 0
        guardap2.n2 = 0
        guardap2.n3 = 0
        guardap2.n4 = 0
        guardap2.n5 = 0
        guardap2.plaza = "Plaza2"
        guardap2.save()
        # plaza3
        guardap3 = plazaAlmuerzoMan()
        guardap3.n1 = 0
        guardap3.n2 = 0
        guardap3.n3 = 0
        guardap3.n4 = 0
        guardap3.n5 = 0
        guardap3.plaza = "Plaza3"
        guardap3.save()
        # plaza4
        guardap4 = plazaAlmuerzoMan()
        guardap4.n1 = 0
        guardap4.n2 = 0
        guardap4.n3 = 0
        guardap4.n4 = 0
        guardap4.n5 = 0
        guardap4.plaza = "Plaza4"
        guardap4.save()
        # plaza5
        guardap5 = plazaAlmuerzoMan()
        guardap5.n1 = 0
        guardap5.n2 = 0
        guardap5.n3 = 0
        guardap5.n4 = 0
        guardap5.n5 = 0
        guardap5.plaza = "Plaza5"
        guardap5.save()
        # plaza6
        guardap6 = plazaAlmuerzoMan()
        guardap6.n1 = 0
        guardap6.n2 = 0
        guardap6.n3 = 0
        guardap6.n4 = 0
        guardap6.n5 = 0
        guardap6.plaza = "Plaza6"
        guardap6.save()
        # plaza7
        guardap7 = plazaAlmuerzoMan()
        guardap7.n1 = 0
        guardap7.n2 = 0
        guardap7.n3 = 0
        guardap7.n4 = 0
        guardap7.n5 = 0
        guardap7.plaza = "Plaza7"
        guardap7.save()
        # plaza8
        guardap8 = plazaAlmuerzoMan()
        guardap8.n1 = 0
        guardap8.n2 = 0
        guardap8.n3 = 0
        guardap8.n4 = 0
        guardap8.n5 = 0
        guardap8.plaza = "Plaza8"
        guardap8.save()
        # plaza9
        guardap9 = plazaAlmuerzoMan()
        guardap9.n1 = 0
        guardap9.n2 = 0
        guardap9.n3 = 0
        guardap9.n4 = 0
        guardap9.n5 = 0
        guardap9.plaza = "Plaza9"
        guardap9.save()
        # plaza10
        guardap10 = plazaAlmuerzoMan()
        guardap10.n1 = 0
        guardap10.n2 = 0
        guardap10.n3 = 0
        guardap10.n4 = 0
        guardap10.n5 = 0
        guardap10.plaza = "Plaza10"
        guardap10.save()
        # plaza11
        guardap11 = plazaAlmuerzoMan()
        guardap11.n1 = 0
        guardap11.n2 = 0
        guardap11.n3 = 0
        guardap11.n4 = 0
        guardap11.n5 = 0
        guardap11.plaza = "Plaza11"
        guardap11.save()

        return redirect('calculadia')

    n_mesa = nmesa
    #plaza1
    if n_mesa == "1":
           guardan1 = plazaAlmuerzoMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza1')
           guardan1.n1 = int(guardan1.n1) + 1
           guardan1.save()
    if n_mesa == "2":
           guardan1 = plazaAlmuerzoMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza1')
           guardan1.n2 = int(guardan1.n2) + 1
           guardan1.save()
    if n_mesa == "3":
           guardan1 = plazaAlmuerzoMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza1')
           guardan1.n3 = int(guardan1.n3) + 1
           guardan1.save()
    if n_mesa == "6":
           guardan1 = plazaAlmuerzoMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza1')
           guardan1.n4 = int(guardan1.n4) + 1
           guardan1.save()
    # if n_mesa == "6A":
    #        guardan1 = plazaAlmuerzoMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza1')
    #        guardan1.n5 = int(guardan1.n5) + 1
    #        guardan1.save()
    #plaza2
    if n_mesa == "4":
           guardan1 = plazaAlmuerzoMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza2')
           guardan1.n1 = int(guardan1.n1) + 1
           guardan1.save()
    if n_mesa == "5":
           guardan1 = plazaAlmuerzoMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza2')
           guardan1.n2 = int(guardan1.n2) + 1
           guardan1.save()
    # if n_mesa == "10":
    #        guardan1 = plazaAlmuerzoMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza2')
    #        guardan1.n3 = int(guardan1.n3) + 1
    #        guardan1.save()
    if n_mesa == "11":
           guardan1 = plazaAlmuerzoMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza2')
           guardan1.n4 = int(guardan1.n4) + 1
           guardan1.save()
    if n_mesa == "12":
           guardan1 = plazaAlmuerzoMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza2')
           guardan1.n5 = int(guardan1.n5) + 1
           guardan1.save()
        #plaza3
    if n_mesa == "12A":
           guardan1 = plazaAlmuerzoMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza3')
           guardan1.n1 = int(guardan1.n1) + 1
           guardan1.save()
    if n_mesa == "14":
           guardan1 = plazaAlmuerzoMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza3')
           guardan1.n2 = int(guardan1.n2) + 1
           guardan1.save()
    if n_mesa == "15":
           guardan1 = plazaAlmuerzoMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza3')
           guardan1.n3 = int(guardan1.n3) + 1
           guardan1.save()
    if n_mesa == "21":
           guardan1 = plazaAlmuerzoMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza3')
           guardan1.n4 = int(guardan1.n4) + 1
           guardan1.save()
    #plaza4
    if n_mesa == "6A":
           guardan1 = plazaAlmuerzoMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza4')
           guardan1.n1 = int(guardan1.n1) + 1
           guardan1.save()
    if n_mesa == "9":
           guardan1 = plazaAlmuerzoMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza4')
           guardan1.n2 = int(guardan1.n2) + 1
           guardan1.save()
    if n_mesa == "10":
           guardan1 = plazaAlmuerzoMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza4')
           guardan1.n3 = int(guardan1.n3) + 1
           guardan1.save()
    if n_mesa == "16":
           guardan1 = plazaAlmuerzoMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza4')
           guardan1.n4 = int(guardan1.n4) + 1
           guardan1.save()
    # if n_mesa == "34":
    #        guardan1 = plazaAlmuerzoMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza4')
    #        guardan1.n5 = int(guardan1.n5) + 1
    #        guardan1.save()
    #plaza5
    if n_mesa == "8":
           guardan1 = plazaAlmuerzoMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza5')
           guardan1.n1 = int(guardan1.n1) + 1
           guardan1.save()
    if n_mesa == "17":
           guardan1 = plazaAlmuerzoMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza5')
           guardan1.n2 = int(guardan1.n2) + 1
           guardan1.save()
    if n_mesa == "18":
           guardan1 = plazaAlmuerzoMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza5')
           guardan1.n3 = int(guardan1.n3) + 1
           guardan1.save()
    if n_mesa == "34":
           guardan1 = plazaAlmuerzoMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza5')
           guardan1.n4 = int(guardan1.n4) + 1
           guardan1.save()
    #plaza6
    if n_mesa == "19":
           guardan1 = plazaAlmuerzoMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza6')
           guardan1.n1 = int(guardan1.n1) + 1
           guardan1.save()
    if n_mesa == "20":
           guardan1 = plazaAlmuerzoMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza6')
           guardan1.n2 = int(guardan1.n2) + 1
           guardan1.save()
    if n_mesa == "24":
           guardan1 = plazaAlmuerzoMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza6')
           guardan1.n3 = int(guardan1.n3) + 1
           guardan1.save()
    if n_mesa == "26":
           guardan1 = plazaAlmuerzoMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza6')
           guardan1.n4 = int(guardan1.n4) + 1
           guardan1.save()
    # if n_mesa == "24":
    #        guardan1 = plazaAlmuerzoMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza6')
    #        guardan1.n4 = int(guardan1.n4) + 1
    #        guardan1.save()
    #plaza7
    if n_mesa == "27":
           guardan1 = plazaAlmuerzoMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza8')
           guardan1.n1 = int(guardan1.n1) + 1
           guardan1.save()
    if n_mesa == "28":
           guardan1 = plazaAlmuerzoMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza8')
           guardan1.n2 = int(guardan1.n2) + 1
           guardan1.save()
    if n_mesa == "29":
           guardan1 = plazaAlmuerzoMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza8')
           guardan1.n3 = int(guardan1.n3) + 1
           guardan1.save()
    if n_mesa == "B3":
           guardan1 = plazaAlmuerzoMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza8')
           guardan1.n4 = int(guardan1.n4) + 1
           guardan1.save()
    #plaza8
    if n_mesa == "22":
           guardan1 = plazaAlmuerzoMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza7')
           guardan1.n1 = int(guardan1.n1) + 1
           guardan1.save()
    if n_mesa == "23":
           guardan1 = plazaAlmuerzoMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza7')
           guardan1.n2 = int(guardan1.n2) + 1
           guardan1.save()
    if n_mesa == "30":
           guardan1 = plazaAlmuerzoMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza7')
           guardan1.n3 = int(guardan1.n3) + 1
           guardan1.save()
    if n_mesa == "31":
           guardan1 = plazaAlmuerzoMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza7')
           guardan1.n4 = int(guardan1.n4) + 1
           guardan1.save()
    if n_mesa == "32":
           guardan1 = plazaAlmuerzoMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza7')
           guardan1.n5 = int(guardan1.n5) + 1
           guardan1.save()
    
    #plaza9
    if n_mesa == "35":
           guardan1 = plazaAlmuerzoMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza9')
           guardan1.n1 = int(guardan1.n1) + 1
           guardan1.save()
    if n_mesa == "36":
           guardan1 = plazaAlmuerzoMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza9')
           guardan1.n2 = int(guardan1.n2) + 1
           guardan1.save()
    if n_mesa == "37":
           guardan1 = plazaAlmuerzoMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza9')
           guardan1.n3 = int(guardan1.n3) + 1
           guardan1.save()
    if n_mesa == "38":
           guardan1 = plazaAlmuerzoMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza9')
           guardan1.n4 = int(guardan1.n4) + 1
           guardan1.save()
    if n_mesa == "39":
           guardan1 = plazaAlmuerzoMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza9')
           guardan1.n5 = int(guardan1.n5) + 1
           guardan1.save()

    #plaza10
    if n_mesa == "JP":
           guardan1 = plazaAlmuerzoMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza10')
           guardan1.n1 = int(guardan1.n1) + 1
           guardan1.save()
    #plaza11
    if n_mesa == "BELUA":
           guardan1 = plazaAlmuerzoMan.objects.get(fecha_dia = fecha_actual, plaza = 'Plaza11')
           guardan1.n1 = int(guardan1.n1) + 1
           guardan1.save()

    return redirect('calculadia')

def estadisticas(request):
        
        fecha_actual = datetime.now()
        
    #clientes atendidos por dia semana
        total_clientes_dia = nuevaReserva.objects.filter(estado_id=2, fechaReserva__year=fecha_actual.year,
    fechaReserva__month=fecha_actual.month).annotate(
        dia_semana=Case(
            When(fechaReserva__week_day=1, then=Value('Domingo')),  # Domingo
            When(fechaReserva__week_day=2, then=Value('Lunes')),  # Lunes
            When(fechaReserva__week_day=3, then=Value('Martes')),  # Martes
            When(fechaReserva__week_day=4, then=Value('Miercoles')),  # Miércoles
            When(fechaReserva__week_day=5, then=Value('Jueves')),  # Jueves
            When(fechaReserva__week_day=6, then=Value('Viernes')),  # Viernes
            When(fechaReserva__week_day=7, then=Value('Sabado')),  # Sábado
            output_field=CharField(),
        )
    ).values('dia_semana').annotate(
        total=Sum('cantidadPersonas')
    )

        etiquetas = [dia['dia_semana'] for dia in total_clientes_dia]
        valores = [cantidad['total'] for cantidad in total_clientes_dia]
        fig, ax = ptl.subplots()
        fig.set_facecolor('#000000')
        ax.pie(valores, labels=etiquetas,autopct='%1.1f%%',startangle=140, shadow=True, textprops={'color': 'white'})
        ax.axis('equal')
        ax.set_title('Reservas Atendidas por día de la Semana', fontweight='bold', fontdict={'color': 'white', 'fontsize': 16})

        buffer = BytesIO()
        ptl.savefig(buffer, format='png')
        buffer.seek(0)
        image_base64 = base64.b64encode(buffer.read()).decode()
        grafico1 = "data:image/png;base64," + image_base64

        # Origen de la reserva 

        por_origen_reserva = nuevaReserva.objects.filter(origen_reserva__isnull=False, fechaReserva__year=fecha_actual.year,
    fechaReserva__month=fecha_actual.month).values('origen_reserva').annotate(total_origen=Count('pk'))

        etiquetas1 = [origen['origen_reserva'] for origen in por_origen_reserva]
        valores1 = [cantidad['total_origen'] for cantidad in por_origen_reserva]
        fig1, ax1 = ptl.subplots()
        fig1.set_facecolor('#000000')
        ax1.pie(valores1, labels=etiquetas1,autopct='%1.1f%%',startangle=140, shadow=True, textprops={'color': 'white'})
        ax1.axis('equal')
        ax1.set_title('Distribución por Origen', fontweight='bold', fontdict={'color': 'white', 'fontsize': 16})

        buffer1 = BytesIO()
        ptl.savefig(buffer1, format='png')
        buffer1.seek(0)
        image_base641 = base64.b64encode(buffer1.read()).decode()
        grafico2 = "data:image/png;base64," + image_base641

    #     # ocupacion mesas 

        # por_mesa_reserva = nuevaReserva.objects.filter(estado_id=2, mesa_asignadaa__isnull=False).values('mesa_asignadaa').annotate(total_mesa=Count('pk'))

        # etiquetas2 = [mesa['mesa_asignadaa'] for mesa in por_mesa_reserva]
        # valores2 = [cantidad['total_mesa'] for cantidad in por_mesa_reserva]
        # fig2, ax2 = ptl.subplots()
        # ax2.pie(valores2, labels=etiquetas2,autopct='%1.1f%%',startangle=140, shadow=True, textprops={'color': 'white'})
        # ax2.axis('equal')
        # ax2.set_title('Ocupación de Mesas', fontweight='bold', fontdict={'color': 'white', 'fontsize': 16})
        # fig2.set_facecolor('#000000')
        # buffer2 = BytesIO()
        # ptl.savefig(buffer2, format='png')
        # buffer2.seek(0)
        # image_base642 = base64.b64encode(buffer2.read()).decode()
        # grafico3 = "data:image/png;base64," + image_base642

    #     ## por estado de la reserva

        por_estado_reserva = nuevaReserva.objects.filter(fechaReserva__year=fecha_actual.year,
    fechaReserva__month=fecha_actual.month).exclude(estado_id=1).annotate(
        estado_reser=Case(
            When(estado=2, then=Value('Atendido')),  
            When(estado=3, then=Value('Anulado')),  
            When(estado=4, then=Value('No Show')), 
            output_field=CharField(),
        )
    ).values('estado_reser').annotate(
        total=Count('pk')
    )
        etiquetas3 = [estado_reserva['estado_reser'] for estado_reserva in por_estado_reserva]
        valores3 = [cantidad['total'] for cantidad in por_estado_reserva]
        fig3, ax3 = ptl.subplots()
        fig3.set_facecolor('#000000')
        ax3.pie(valores3, labels=etiquetas3,autopct='%1.1f%%',startangle=140, shadow=True, textprops={'color': 'white'} )
        ax3.axis('equal')
        ax3.set_title('Distribución por Estado', fontweight='bold', fontdict={'color': 'white', 'fontsize': 16})

        buffer3 = BytesIO()
        ptl.savefig(buffer3, format='png')
        buffer3.seek(0)
        image_base643 = base64.b64encode(buffer3.read()).decode()
        grafico4 = "data:image/png;base64," + image_base643

    #     # por almuerzo o cena  atendidas

        por_hora_reserva = nuevaReserva.objects.filter(estado_id=2, fechaReserva__year=fecha_actual.year,
    fechaReserva__month=fecha_actual.month).annotate(
        hora_reserva=Case(
            When(hora__lt='16:01', then=Value('Almuerzo')),  
            When(hora__gt='16:01', then=Value('Cena')), 
            output_field=CharField(),
        )
    ).values('hora_reserva').annotate(
        total=Count('pk')
    )
        etiquetas4 = [hora_reserva['hora_reserva'] for hora_reserva in por_hora_reserva]
        valores4 = [cantidad['total'] for cantidad in por_hora_reserva]
        fig4, ax4 = ptl.subplots()
        fig4.set_facecolor('#000000')
        ax4.pie(valores4, labels=etiquetas4,autopct='%1.1f%%',startangle=140, shadow=True, textprops={'color': 'white'} )
        ax4.axis('equal')
        ax4.set_title('Reservas Atendidas Según Horario', fontweight='bold', fontdict={'color': 'white', 'fontsize': 16})

        buffer4 = BytesIO()
        ptl.savefig(buffer4, format='png')
        buffer4.seek(0)
        image_base644 = base64.b64encode(buffer4.read()).decode()
        grafico5 = "data:image/png;base64," + image_base644

        # reservas con incidencia 

        cuenta_incidencias = incidencia.objects.filter(id_reservacion__isnull = False).count()
        cuenta_reservas = nuevaReserva.objects.filter(estado = 2).count()

        etiquetas5 = ['Incidencias','Reservas']
        valores5 = [cuenta_incidencias, cuenta_reservas]
        fig5, ax5 = ptl.subplots()
        fig5.set_facecolor('#000000')

        ax5.pie(valores5, labels=etiquetas5,autopct='%1.1f%%',startangle=140, shadow=True, textprops={'color': 'white'} )

        ax5.axis('equal')
        ax5.set_title('Atenciones con Incidencias', fontweight='bold', fontdict={'color': 'white', 'fontsize': 16})

        buffer5 = BytesIO()
        ptl.savefig(buffer5, format='png')
        buffer5.seek(0)
        image_base645 = base64.b64encode(buffer5.read()).decode()
        grafico6 = "data:image/png;base64," + image_base645



        return render(request, 'estadisticas.html', {'graf':grafico1,'graf6':grafico6, 'graf2':grafico2, 'graf4':grafico4, 'graf5':grafico5})

def guardaPlazaAlmfds(request):
    fecha_actual = datetime.now().date()
    hora_actual = datetime.now().strftime('%H:%M')

    if plazaAlmuerzo.objects.filter(fecha_dia=fecha_actual).exists():
        return HttpResponse('Ya esta guardado')

    # if hora_actual < '16:01':
    #     return HttpResponse('No puede guardar la plaza hasta finalizar el turno')

    if request.method == 'POST':

        mozo = request.POST.get('mozoplaza1')
        m1 = request.POST.get('m1')
        r1 = request.POST.get('r1')
        m2 = request.POST.get('m2')
        r2 = request.POST.get('r2')
        m3 = request.POST.get('m3')
        r3 = request.POST.get('r3')
        m6 = request.POST.get('m6')
        r6 = request.POST.get('r6')
        # m6a = request.POST.get('m6A')
        # r6a = request.POST.get('r6A')
        total = request.POST.get('totalpz1')
        plaza = 'Plaza1'
        anfi = request.POST.get('anfitriona')

        plaza_dia = plazaAlmuerzo(m1=m1, r1=r1, m2=m2, r2=r2, m3=m3, r3=r3,
                                   m4=m6, r4=r6, total=total, plaza=plaza, mozo_nombre=mozo, nombre_anfitriona=anfi)

        plaza_dia.save()

        mozo = request.POST.get('mozoplaza2')
        m4 = request.POST.get('m4')
        r4 = request.POST.get('r4')
        m5 = request.POST.get('m5')
        r5 = request.POST.get('r5')
        # m10 = request.POST.get('m10')
        # r10 = request.POST.get('r10')
        m11 = request.POST.get('m11')
        r11 = request.POST.get('r11')
        m12 = request.POST.get('m12')
        r12 = request.POST.get('r12')
        total = request.POST.get('totalpz2')
        plaza = 'Plaza2'
        anfi = request.POST.get('anfitriona')

        plaza_dia = plazaAlmuerzo(m1=m4, r1=r4, m2=m5, r2=r5, m3=m11,
                                   r3=r11, m4=m12, r4=r12, total=total, plaza=plaza, mozo_nombre=mozo, nombre_anfitriona=anfi)

        plaza_dia.save()

        mozo = request.POST.get('mozoplaza3')
        m12a = request.POST.get('m12A')
        r12 = request.POST.get('r12')
        m14 = request.POST.get('m14')
        r14 = request.POST.get('r14')
        m15 = request.POST.get('m15')
        r15 = request.POST.get('r15')
        m21 = request.POST.get('m21')
        r21 = request.POST.get('r21')
        total = request.POST.get('totalpz3')
        plaza = 'Plaza3'
        anfi = request.POST.get('anfitriona')

        plaza_dia = plazaAlmuerzo(m1=m12a, r1=r12, m2=m14, r2=r14, m3=m15, r3=r15, m4=m21,
                                   r4=r21, total=total, plaza=plaza, mozo_nombre=mozo, nombre_anfitriona=anfi)

        plaza_dia.save()

        mozo = request.POST.get('mozoplaza4')
        m6a = request.POST.get('m6A')
        r6a = request.POST.get('r6A')
        m10 = request.POST.get('m10')
        r10 = request.POST.get('r10')
        # m18 = request.POST.get('m18')
        # r18 = request.POST.get('r18')
        # m8 = request.POST.get('m8')
        # r8 = request.POST.get('r8')
        m9 = request.POST.get('m9')
        r9 = request.POST.get('r9')
        m16 = request.POST.get('m16')
        r16 = request.POST.get('r16')
        # m34 = request.POST.get('m34')
        # r34 = request.POST.get('r34')
        total = request.POST.get('totalpz4')
        plaza = 'Plaza4'
        anfi = request.POST.get('anfitriona')

        plaza_dia = plazaAlmuerzo(m1=m6a, r1=r6a, m2=m9, r2=r9, m3=m10, r3=r10, m4=m16,
                                   r4=r16, total=total, plaza=plaza, mozo_nombre=mozo, nombre_anfitriona=anfi)

        plaza_dia.save()

        mozo = request.POST.get('mozoplaza5')
        m8 = request.POST.get('m8')
        r8 = request.POST.get('r8')
        m17 = request.POST.get('m17')
        r17 = request.POST.get('r17')
        m18 = request.POST.get('m18')
        r18 = request.POST.get('r18')
        m34 = request.POST.get('m34')
        r34 = request.POST.get('r34')
        # m26 = request.POST.get('m26')
        # r26 = request.POST.get('r26')
        # m19 = request.POST.get('m19')
        # r19 = request.POST.get('r19')
        # m20 = request.POST.get('m20')
        # r20 = request.POST.get('r20')
        # m21 = request.POST.get('m21')
        # r21 = request.POST.get('r21')
        total = request.POST.get('totalpz5')
        plaza = 'Plaza5'
        anfi = request.POST.get('anfitriona')

        plaza_dia = plazaAlmuerzo(m1=m8, r1=r8, m2=m17, r2=r17, m3=m18, r3=r18, m4=m34,
                                   r4=r34, total=total, plaza=plaza, mozo_nombre=mozo, nombre_anfitriona=anfi)

        plaza_dia.save()

        mozo = request.POST.get('mozoplaza6')
        m19 = request.POST.get('m19')
        r19 = request.POST.get('r19')
        m20 = request.POST.get('m20')
        r20 = request.POST.get('r20')
        m24 = request.POST.get('m24')
        r24 = request.POST.get('r24')
        m26 = request.POST.get('m26')
        r26 = request.POST.get('r26')
        # m22 = request.POST.get('m22')
        # r22 = request.POST.get('r22')
        # m30 = request.POST.get('m30')
        # r30 = request.POST.get('r30')
        # m31 = request.POST.get('m31')
        # r31 = request.POST.get('r31')
        # m32 = request.POST.get('m32')
        # r32 = request.POST.get('r32')
        total = request.POST.get('totalpz6')
        plaza = 'Plaza6'
        anfi = request.POST.get('anfitriona')

        plaza_dia = plazaAlmuerzo(m1=m19, r1=r19, m2=m20, r2=r20, m3=m24, r3=r24, m4=m26,
                                   r4=r26, total=total, plaza=plaza, mozo_nombre=mozo, nombre_anfitriona=anfi)

        plaza_dia.save()

        mozo = request.POST.get('mozoplaza7')
        m27 = request.POST.get('m27')
        r27 = request.POST.get('r27')
        m28 = request.POST.get('m28')
        r28 = request.POST.get('r28')
        m29 = request.POST.get('m29')
        r29 = request.POST.get('r29')
        mb3 = request.POST.get('mb3')
        rb3 = request.POST.get('rb3')
        total = request.POST.get('totalpz7')
        plaza = 'Plaza7'
        anfi = request.POST.get('anfitriona')

        plaza_dia = plazaAlmuerzo(m1=m27, r1=r27, m2=m28, r2=r28, m3=m29, r3=r29, m4=mb3,
                                   r4=rb3, total=total, plaza=plaza, mozo_nombre=mozo, nombre_anfitriona=anfi)

        plaza_dia.save()

        mozo = request.POST.get('mozoplaza8')
        m22 = request.POST.get('m22')
        r22 = request.POST.get('r22')
        m23 = request.POST.get('m23')
        r23 = request.POST.get('r23')
        m30 = request.POST.get('m30')
        r30 = request.POST.get('r30')
        m31 = request.POST.get('m31')
        r31 = request.POST.get('r31')
        m32 = request.POST.get('m32')
        r32 = request.POST.get('r32')
        # m27 = request.POST.get('m27')
        # r27 = request.POST.get('r27')
        # m28 = request.POST.get('m28')
        # r28 = request.POST.get('r28')
        # m29 = request.POST.get('m29')
        # r29 = request.POST.get('r29')
        # mb3 = request.POST.get('mb3')
        # rb3 = request.POST.get('rb3')
        total = request.POST.get('totalpz8')
        plaza = 'Plaza8'
        anfi = request.POST.get('anfitriona')

        plaza_dia = plazaAlmuerzo(m1=m22, r1=r22, m2=m23, r2=r23, m3=m30, r3=r30, m4=m31,
                                   r4=r31, m5=m32, r5=r32, total=total, plaza=plaza, mozo_nombre=mozo, nombre_anfitriona=anfi)

        plaza_dia.save()

        mozo = request.POST.get('mozoplaza9')
        m35 = request.POST.get('m35')
        r35 = request.POST.get('r35')
        m36 = request.POST.get('m36')
        r36 = request.POST.get('r36')
        m37 = request.POST.get('m37')
        r37 = request.POST.get('r37')
        m38 = request.POST.get('m38')
        r38 = request.POST.get('r38')
        m39 = request.POST.get('m39')
        r39 = request.POST.get('r39')
        total = request.POST.get('totalpz9')
        plaza = 'Plaza9'
        anfi = request.POST.get('anfitriona')

        plaza_dia = plazaAlmuerzo(m1=m35, r1=r35, m2=m36, r2=r36, m3=m37, r3=r37, m4=m38,
                                   r4=r38, m5=m39, r5=r39, total=total, plaza=plaza, mozo_nombre=mozo, nombre_anfitriona=anfi)

        plaza_dia.save()

        
        mozo = request.POST.get('mozoplaza10') + ', ' + \
                                request.POST.get('mozoplaza10b') + ' y ' + \
                                request.POST.get('mozoplaza10cm')
        mjp = request.POST.get('mjp')
        rjp = request.POST.get('rjp')
        total = request.POST.get('totalpz10')
        plaza = 'Plaza10'
        anfi = request.POST.get('anfitriona')

        plaza_dia = plazaAlmuerzo(
            m1=mjp, r1=rjp, total=total, plaza=plaza, mozo_nombre=mozo, nombre_anfitriona=anfi)

        plaza_dia.save()

        mozo = request.POST.get('mozoplaza11') + ', ' + \
                                request.POST.get('mozoplaza11b') + ' y ' + \
                                request.POST.get('mozoplaza11cm')
                                
        mbelua = request.POST.get('mbelua')
        rbelua = request.POST.get('rbelua')
        total = request.POST.get('totalpz11')
        plaza = 'Plaza11'
        anfi = request.POST.get('anfitriona')

        plaza_dia = plazaAlmuerzo(m1=mbelua, r1=rbelua, total=total,
                                  plaza=plaza, mozo_nombre=mozo, nombre_anfitriona=anfi)

        plaza_dia.save()


        
        zona_horaria_peru = pytz.timezone('America/Lima')
        fecha_actual = datetime.now(zona_horaria_peru)
        es_fds = fecha_actual.weekday() in [5, 6]

        if es_fds:
            return controlMesaManAlmfds(request, 1)
        else:
            return controlMesaManAlm(request, 1)

def guardaPlazaCenafds(request):
    fecha_actual = datetime.now().date()
    hora_actual = datetime.now().strftime('%H:%M')

    if plazaCena.objects.filter(fecha_dia=fecha_actual).exists():
        return HttpResponse('Ya esta guardado')

    # if hora_actual < '16:01':
    #     return HttpResponse('No puede guardar la plaza hasta finalizar el turno')

    if request.method == 'POST':

        mozo = request.POST.get('mozoplaza1c')
        m1 = request.POST.get('m1c')
        r1 = request.POST.get('r1c')
        m2 = request.POST.get('m2c')
        r2 = request.POST.get('r2c')
        m3 = request.POST.get('m3c')
        r3 = request.POST.get('r3c')
        m6 = request.POST.get('m6c')
        r6 = request.POST.get('r6c')
        # m6a = request.POST.get('m6A')
        # r6a = request.POST.get('r6A')
        total = request.POST.get('totalpz1c')
        plaza = 'Plaza1'
        anfi = request.POST.get('anfitrionac')

        plaza_dia = plazaCena(m1=m1, r1=r1, m2=m2, r2=r2, m3=m3, r3=r3,
                                   m4=m6, r4=r6, total=total, plaza=plaza, mozo_nombre=mozo, nombre_anfitriona=anfi)

        plaza_dia.save()

        mozo = request.POST.get('mozoplaza2c')
        m4 = request.POST.get('m4c')
        r4 = request.POST.get('r4c')
        m5 = request.POST.get('m5c')
        r5 = request.POST.get('r5c')
        # m10 = request.POST.get('m10')
        # r10 = request.POST.get('r10')
        m11 = request.POST.get('m11c')
        r11 = request.POST.get('r11c')
        m12 = request.POST.get('m12c')
        r12 = request.POST.get('r12c')
        total = request.POST.get('totalpz2c')
        plaza = 'Plaza2'
        anfi = request.POST.get('anfitrionac')

        plaza_dia = plazaCena(m1=m4, r1=r4, m2=m5, r2=r5, m3=m11,
                                   r3=r11, m4=m12, r4=r12, total=total, plaza=plaza, mozo_nombre=mozo, nombre_anfitriona=anfi)

        plaza_dia.save()

        mozo = request.POST.get('mozoplaza3c')
        m12a = request.POST.get('m12Ac')
        r12 = request.POST.get('r12c')
        m14 = request.POST.get('m14c')
        r14 = request.POST.get('r14c')
        m15 = request.POST.get('m15c')
        r15 = request.POST.get('r15c')
        m21 = request.POST.get('m21c')
        r21 = request.POST.get('r21c')
        total = request.POST.get('totalpz3c')
        plaza = 'Plaza3'
        anfi = request.POST.get('anfitrionac')

        plaza_dia = plazaCena(m1=m12a, r1=r12, m2=m14, r2=r14, m3=m15, r3=r15, m4=m21,
                                   r4=r21, total=total, plaza=plaza, mozo_nombre=mozo, nombre_anfitriona=anfi)

        plaza_dia.save()

        mozo = request.POST.get('mozoplaza4c')
        m6a = request.POST.get('m6Ac')
        r6a = request.POST.get('r6Ac')
        m10 = request.POST.get('m10c')
        r10 = request.POST.get('r10c')
        # m18 = request.POST.get('m18')
        # r18 = request.POST.get('r18')
        # m8 = request.POST.get('m8')
        # r8 = request.POST.get('r8')
        m9 = request.POST.get('m9c')
        r9 = request.POST.get('r9c')
        m16 = request.POST.get('m16c')
        r16 = request.POST.get('r16c')
        # m34 = request.POST.get('m34')
        # r34 = request.POST.get('r34')
        total = request.POST.get('totalpz4c')
        plaza = 'Plaza4'
        anfi = request.POST.get('anfitrionac')

        plaza_dia = plazaCena(m1=m6a, r1=r6a, m2=m9, r2=r9, m3=m10, r3=r10, m4=m16,
                                   r4=r16, total=total, plaza=plaza, mozo_nombre=mozo, nombre_anfitriona=anfi)

        plaza_dia.save()

        mozo = request.POST.get('mozoplaza5c')
        m8 = request.POST.get('m8c')
        r8 = request.POST.get('r8c')
        m17 = request.POST.get('m17c')
        r17 = request.POST.get('r17c')
        m18 = request.POST.get('m18c')
        r18 = request.POST.get('r18c')
        m34 = request.POST.get('m34c')
        r34 = request.POST.get('r34c')
        # m26 = request.POST.get('m26')
        # r26 = request.POST.get('r26')
        # m19 = request.POST.get('m19')
        # r19 = request.POST.get('r19')
        # m20 = request.POST.get('m20')
        # r20 = request.POST.get('r20')
        # m21 = request.POST.get('m21')
        # r21 = request.POST.get('r21')
        total = request.POST.get('totalpz5c')
        plaza = 'Plaza5'
        anfi = request.POST.get('anfitrionac')

        plaza_dia = plazaCena(m1=m8, r1=r8, m2=m17, r2=r17, m3=m18, r3=r18, m4=m34,
                                   r4=r34, total=total, plaza=plaza, mozo_nombre=mozo, nombre_anfitriona=anfi)

        plaza_dia.save()

        mozo = request.POST.get('mozoplaza6c')
        m19 = request.POST.get('m19c')
        r19 = request.POST.get('r19c')
        m20 = request.POST.get('m20c')
        r20 = request.POST.get('r20c')
        m24 = request.POST.get('m24c')
        r24 = request.POST.get('r24c')
        m26 = request.POST.get('m26c')
        r26 = request.POST.get('r26c')
        # m22 = request.POST.get('m22')
        # r22 = request.POST.get('r22')
        # m30 = request.POST.get('m30')
        # r30 = request.POST.get('r30')
        # m31 = request.POST.get('m31')
        # r31 = request.POST.get('r31')
        # m32 = request.POST.get('m32')
        # r32 = request.POST.get('r32')
        total = request.POST.get('totalpz6c')
        plaza = 'Plaza6'
        anfi = request.POST.get('anfitrionac')

        plaza_dia = plazaCena(m1=m19, r1=r19, m2=m20, r2=r20, m3=m24, r3=r24, m4=m26,
                                   r4=r26, total=total, plaza=plaza, mozo_nombre=mozo, nombre_anfitriona=anfi)

        plaza_dia.save()

        mozo = request.POST.get('mozoplaza7c')
        m27 = request.POST.get('m27c')
        r27 = request.POST.get('r27c')
        m28 = request.POST.get('m28c')
        r28 = request.POST.get('r28c')
        m29 = request.POST.get('m29c')
        r29 = request.POST.get('r29c')
        mb3 = request.POST.get('mb3c')
        rb3 = request.POST.get('rb3c')
        total = request.POST.get('totalpz7c')
        plaza = 'Plaza7'
        anfi = request.POST.get('anfitrionac')

        plaza_dia = plazaCena(m1=m27, r1=r27, m2=m28, r2=r28, m3=m29, r3=r29, m4=mb3,
                                   r4=rb3, total=total, plaza=plaza, mozo_nombre=mozo, nombre_anfitriona=anfi)

        plaza_dia.save()

        mozo = request.POST.get('mozoplaza8c')
        m22 = request.POST.get('m22c')
        r22 = request.POST.get('r22c')
        m23 = request.POST.get('m23c')
        r23 = request.POST.get('r23c')
        m30 = request.POST.get('m30c')
        r30 = request.POST.get('r30c')
        m31 = request.POST.get('m31c')
        r31 = request.POST.get('r31c')
        m32 = request.POST.get('m32c')
        r32 = request.POST.get('r32c')
        # m27 = request.POST.get('m27')
        # r27 = request.POST.get('r27')
        # m28 = request.POST.get('m28')
        # r28 = request.POST.get('r28')
        # m29 = request.POST.get('m29')
        # r29 = request.POST.get('r29')
        # mb3 = request.POST.get('mb3')
        # rb3 = request.POST.get('rb3')
        total = request.POST.get('totalpz8c')
        plaza = 'Plaza8'
        anfi = request.POST.get('anfitrionac')

        plaza_dia = plazaCena(m1=m22, r1=r22, m2=m23, r2=r23, m3=m30, r3=r30, m4=m31,
                                   r4=r31, m5=m32, r5=r32, total=total, plaza=plaza, mozo_nombre=mozo, nombre_anfitriona=anfi)

        plaza_dia.save()

        mozo = request.POST.get('mozoplaza9c')
        m35 = request.POST.get('m35c')
        r35 = request.POST.get('r35c')
        m36 = request.POST.get('m36c')
        r36 = request.POST.get('r36c')
        m37 = request.POST.get('m37c')
        r37 = request.POST.get('r37c')
        m38 = request.POST.get('m38c')
        r38 = request.POST.get('r38c')
        m39 = request.POST.get('m39c')
        r39 = request.POST.get('r39c')
        total = request.POST.get('totalpz9c')
        plaza = 'Plaza9'
        anfi = request.POST.get('anfitrionac')

        plaza_dia = plazaCena(m1=m35, r1=r35, m2=m36, r2=r36, m3=m37, r3=r37, m4=m38,
                                   r4=r38, m5=m39, r5=r39, total=total, plaza=plaza, mozo_nombre=mozo, nombre_anfitriona=anfi)

        plaza_dia.save()

        
        mozo = request.POST.get('mozoplaza10c') + ', ' + \
                                request.POST.get('mozoplaza10bc') + ' y ' + \
                                request.POST.get('mozoplaza10bcn')
        mjp = request.POST.get('mjpc')
        rjp = request.POST.get('rjpc')
        total = request.POST.get('totalpz10c')
        plaza = 'Plaza10'
        anfi = request.POST.get('anfitrionac')

        plaza_cena = plazaCena(m1=mjp, r1=rjp, total=total,
                               plaza=plaza, mozo_nombre=mozo, nombre_anfitriona=anfi)

        plaza_cena.save()

        mozo = request.POST.get('mozoplaza11c') + ', ' + \
                                request.POST.get('mozoplaza11bc') + ' y ' + \
                                request.POST.get('mozoplaza11bcn') 
        mbelua = request.POST.get('mbeluac')
        rbelua = request.POST.get('rbeluac')
        total = request.POST.get('totalpz11c')
        plaza = 'Plaza11'
        anfi = request.POST.get('anfitrionac')

        plaza_cena = plazaCena(m1=mbelua, r1=rbelua, total=total,
                               plaza=plaza, mozo_nombre=mozo, nombre_anfitriona=anfi)

        plaza_cena.save()

        

        
        zona_horaria_peru = pytz.timezone('America/Lima')
        fecha_actual = datetime.now(zona_horaria_peru)
        es_fds = fecha_actual.weekday() in [5, 6]

        if es_fds:
            return controlMesaManCenfds(request, 1)
        else:
            return controlMesaManAlm(request, 1)